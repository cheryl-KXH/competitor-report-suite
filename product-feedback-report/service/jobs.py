from __future__ import annotations

from datetime import date, datetime
import re
import traceback
import sys
import shutil
import tempfile
import time
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from service import dingtalk_docs, dingtalk_table
from service.config import load_configs, output_root
from scripts.delivery.generate_tables import generate_delivery_tables
from scripts.delivery.prepare_product_menu import (
    _crawl_date,
    _primary_brand,
    _safe_filename as safe_delivery_filename,
    prepare_product_menu,
)
from scripts.delivery.processing import (
    find_delivery_rows,
    jd_delivery_metrics,
    platform_delivery_metrics,
    read_annotation,
    tracked_product_statuses,
)
from scripts.reporting.generate_report import generate_report
from scripts.reporting.html import (
    delivery_report_from_metrics,
    jd_report_from_metrics,
    social_report_from_summaries,
)
from scripts.social.generate_tables import generate_consumer_feedback_tables
from scripts.social.processing import (
    normalize_social_product_key,
    social_cleaned_workbook_products,
    summarize_social_cleaned_workbook,
)


CONSUMER_FEEDBACK_TABLE_ID = "hvcu7Bw"
DELIVERY_OUTPUT_PLATFORMS = (
    ("meituanData", "美团"),
    ("elemeData", "饿了么"),
    ("jdData", "京东"),
)
DELIVERY_OUTPUT_FIELDS = ("meituanData", "elemeData", "jdData")
REPORT_OUTPUT_FIELDS = (*DELIVERY_OUTPUT_FIELDS, "report", "folder")
SOCIAL_INPUT_PLATFORMS = (("socialCleanedRawData", "社媒CleanedRawData"),)
SOCIAL_INPUT_FIELDS = tuple(key for key, _ in SOCIAL_INPUT_PLATFORMS)
AI_TABLE_ATTACHMENT_URL_ATTEMPTS = 10
AI_TABLE_ATTACHMENT_URL_DELAY_SECONDS = 1.0


class ArchivedSocialSourceUnreadable(RuntimeError):
    def __init__(self, filename: str) -> None:
        self.filename = filename
        super().__init__(
            f"[{filename}]已归档为钉钉文档链接导致读取失败，"
            "请重新从本地上传原始文件，再进行报告生成。"
        )


def _status(configs: dict[str, dict[str, Any]], key: str, fallback: str) -> str:
    return configs.get("report_rules", {}).get("statuses", {}).get(key, fallback)


def _task_input_dir(configs: dict[str, dict[str, Any]], record: dingtalk_table.TaskRecord) -> Path:
    value = record.cells.get("deliveryData") or record.cells.get("allData")
    raw_dir = output_root(configs) / record.record_id / "raw_data"
    linked_dir = dingtalk_docs.download_linked_folder(configs.get("dingtalk_docs", {}), value, raw_dir)
    if linked_dir and linked_dir.is_dir():
        return linked_dir
    raise RuntimeError("输入数据字段为空，或不是可识别的本机文件/钉钉文档附件。")


def _task_delivery_input_dir(configs: dict[str, dict[str, Any]], record: dingtalk_table.TaskRecord) -> Path:
    value = record.cells.get("deliveryData")
    if not _has_dingtalk_node_link(value):
        raise RuntimeError("“外卖数据”字段缺少可识别的钉钉 xlsx 附件。")
    raw_dir = output_root(configs) / record.record_id / "raw_data"
    linked_dir = dingtalk_docs.download_linked_folder(
        configs.get("dingtalk_docs", {}), value, raw_dir
    )
    if linked_dir and linked_dir.is_dir():
        return linked_dir
    raise RuntimeError("“外卖数据”字段不是可下载的钉钉 xlsx 附件。")


def _task_docs_folder_id(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    keys: tuple[str, ...] = ("deliveryData", "productMenu", *DELIVERY_OUTPUT_FIELDS),
) -> str | None:
    docs_config = configs.get("dingtalk_docs", {})
    for key in keys:
        value = record.cells.get(key)
        if not value:
            continue
        folder_id = dingtalk_docs.folder_id_for_linked_node(docs_config, value)
        if folder_id:
            return folder_id
    return None


def _value_items(value: Any) -> list[Any]:
    if isinstance(value, list):
        return [item for item in value if item]
    return [value] if value else []


def _has_dingtalk_node_link(value: Any) -> bool:
    return any(
        dingtalk_docs.node_id_from_url(dingtalk_docs.extract_link(item))
        for item in _value_items(value)
    )


def _local_item_name(item: Any) -> str:
    if isinstance(item, dict):
        for key in ("filename", "fileName", "name"):
            if item.get(key):
                return str(item[key]).strip()
    return ""


def _single_ai_table_attachment(
    value: Any, field_label: str, *, require_url: bool = True
) -> tuple[str, str] | None:
    items = _value_items(value)
    attachments = [item for item in items if isinstance(item, dict) and item.get("resourceId")]
    if not attachments:
        return None
    if len(items) != 1 or len(attachments) != 1:
        raise RuntimeError(f"“{field_label}”字段仅支持一个本地上传的 xlsx 附件。")
    item = attachments[0]
    name = _local_item_name(item)
    if not name:
        raise RuntimeError(f"“{field_label}”本地上传附件缺少文件名。")
    if Path(name).name != name:
        raise RuntimeError(f"“{field_label}”本地上传附件文件名无效：{name}")
    if Path(name).suffix.lower() != ".xlsx":
        raise RuntimeError(f"“{field_label}”字段只支持 xlsx 文件：{name}")
    # The current AI Table MCP returns the signed download URL in `url` while
    # `resourceUrl` can be a relative metadata endpoint such as /core/api/....
    resource_url = item.get("url") or item.get("resourceUrl")
    if isinstance(resource_url, list):
        resource_url = resource_url[0] if resource_url else ""
    url = str(resource_url or "").strip()
    if require_url and not url.startswith(("http://", "https://")):
        raise RuntimeError(f"“{field_label}”本地上传附件缺少可下载 URL：{name}")
    return name, url if url.startswith(("http://", "https://")) else ""


def _wait_for_ai_table_attachment_url(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    *,
    attempts: int = AI_TABLE_ATTACHMENT_URL_ATTEMPTS,
    delay_seconds: float = AI_TABLE_ATTACHMENT_URL_DELAY_SECONDS,
) -> tuple[str, str] | None:
    for attempt in range(attempts):
        attachment = _single_ai_table_attachment(
            record.cells.get(field_key), field_label, require_url=False
        )
        if attachment is None:
            return None
        if attachment[1]:
            return attachment
        if attempt + 1 < attempts:
            if delay_seconds > 0:
                time.sleep(delay_seconds)
            refreshed = dingtalk_table.fetch_record(configs, record.record_id)
            record.cells[field_key] = refreshed.cells.get(field_key)
    name = _single_ai_table_attachment(
        record.cells.get(field_key), field_label, require_url=False
    )[0]
    raise RuntimeError(
        f"“{field_label}”本地上传附件在等待后仍缺少可下载 URL：{name}。"
        "请稍后重试，或确认钉钉 AI 表 MCP 能返回附件 resourceUrl。"
    )


def _download_ai_table_attachment(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    attachment: tuple[str, str],
    target_dir: Path | None = None,
) -> Path:
    name, url = attachment
    if target_dir is None:
        task_root = output_root(configs) / record.record_id
        if field_key == "deliveryData":
            target_dir = task_root / "raw_data"
        elif field_key in SOCIAL_INPUT_FIELDS:
            target_dir = task_root / "social_inputs" / field_key
        else:
            target_dir = task_root
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / name
    temporary = target.with_name(f".{target.name}.download")
    try:
        request = urllib.request.Request(url)
        with urllib.request.urlopen(request, timeout=120) as response:
            content = response.read()
        if not content:
            raise RuntimeError(f"“{field_label}”本地上传附件下载结果为空：{name}")
        temporary.write_bytes(content)
        temporary.replace(target)
    except Exception as exc:
        if isinstance(exc, RuntimeError):
            raise
        raise RuntimeError(f"“{field_label}”本地上传附件下载失败：{name}：{exc}") from exc
    finally:
        if temporary.exists():
            temporary.unlink()
    return target


def _download_uploaded_xlsx(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    target_dir: Path,
    *,
    required: bool = True,
) -> Path | None:
    value = record.cells.get(field_key)
    if not _value_items(value):
        if required:
            raise RuntimeError(f"“{field_label}”字段为空，请先从 AI 表上传一个 xlsx 附件。")
        return None
    if _has_dingtalk_node_link(value):
        raise RuntimeError(
            f"“{field_label}”已是钉钉文档链接，不属于未归档的 AI 表上传附件。"
            "如需重新生成，请先重新上传原始 xlsx。"
        )
    attachment = _wait_for_ai_table_attachment_url(
        configs, record, field_key, field_label
    )
    if attachment is None:
        raise RuntimeError(
            f"“{field_label}”只支持 AI 表“+”上传的单个 xlsx 附件。"
        )
    return _download_ai_table_attachment(
        configs,
        record,
        field_key,
        field_label,
        attachment,
        target_dir=target_dir,
    )


def _single_local_xlsx(value: Any, field_label: str) -> Path:
    items = _value_items(value)
    if not items:
        raise RuntimeError(f"“{field_label}”字段为空，请先放入一个 xlsx 文件。")
    candidates: dict[Path, Path] = {}
    search_roots = (Path.home() / "Desktop", Path.cwd())
    for item in items:
        link = dingtalk_docs.extract_link(item)
        if dingtalk_docs.node_id_from_url(link):
            continue
        if link.startswith(("http://", "https://")):
            raise RuntimeError(f"“{field_label}”字段包含无法识别的非钉钉链接：{link}")
        raw_path = link.replace("file://", "", 1) if link.startswith("file://") else link
        path = Path(raw_path) if raw_path else None
        explicit_path = bool(
            path
            and (
                link.startswith("file://")
                or path.is_absolute()
                or path.parent != Path(".")
            )
        )
        if explicit_path and path and path.exists():
            resolved = path.resolve()
            candidates[resolved] = path
            continue
        name = _local_item_name(item) or raw_path
        if not name:
            continue
        for root in search_roots:
            direct = root / name
            if direct.exists():
                candidates[direct.resolve()] = direct
    if not candidates:
        raise RuntimeError(f"“{field_label}”字段未找到可上传的本机 xlsx 文件。")
    if len(candidates) != 1:
        names = "、".join(sorted(path.name for path in candidates.values()))
        raise RuntimeError(f"“{field_label}”字段匹配到多个本机文件：{names}。仅支持单个 xlsx。")
    path = next(iter(candidates.values()))
    if path.is_dir():
        raise RuntimeError(f"“{field_label}”字段指向本机目录，仅支持单个 xlsx 文件。")
    if path.suffix.lower() != ".xlsx":
        raise RuntimeError(f"“{field_label}”字段只支持 xlsx 文件：{path.name}")
    return path


def _task_brand(record: dingtalk_table.TaskRecord) -> str:
    value = record.cells.get("brand")
    if isinstance(value, dict):
        value = value.get("name") or value.get("value") or value.get("text")
    if isinstance(value, list):
        value = value[0] if value else ""
        if isinstance(value, dict):
            value = value.get("name") or value.get("value") or value.get("text")
    return str(value or "").strip()


def _task_products(record: dingtalk_table.TaskRecord) -> list[str]:
    def text_values(value: Any) -> list[str]:
        if isinstance(value, dict):
            for key in ("value", "name", "text"):
                if key in value:
                    return text_values(value[key])
            return []
        if isinstance(value, list):
            values: list[str] = []
            for item in value:
                values.extend(text_values(item))
            return values
        text = str(value or "").strip()
        return re.split(r"[,，、]+", text) if text else []

    products: list[str] = []
    seen: set[str] = set()
    for item in text_values(record.cells.get("productName")):
        product = item.strip()
        if product and product not in seen:
            seen.add(product)
            products.append(product)
    return products


def _task_report_date(record: dingtalk_table.TaskRecord) -> date:
    value = record.cells.get("launchDate")
    if isinstance(value, dict):
        value = value.get("value") or value.get("date") or value.get("text")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        raise RuntimeError("“报告日期”为空，无法创建钉钉归档目录。")
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError as exc:
        raise RuntimeError(f"“报告日期”格式无效：{text}") from exc


def _task_folder_name(record: dingtalk_table.TaskRecord) -> str:
    brand = _task_brand(record)
    if not brand:
        raise RuntimeError("“竞品品牌”为空，无法创建钉钉归档目录。")
    products = _task_products(record)
    if not products:
        raise RuntimeError("“关注新品”为空，无法创建钉钉归档目录。")
    clean_brand = re.sub(r'[\\/:*?"<>|\r\n]+', "_", brand).strip()
    clean_products = [
        cleaned
        for item in products
        if (cleaned := re.sub(r'[\\/:*?"<>|\r\n]+', "_", item).strip())
    ]
    if not clean_brand or not clean_products:
        raise RuntimeError("“竞品品牌”或“关注新品”无法生成有效的钉钉文件夹名称。")
    return f"{clean_brand}：{'、'.join(clean_products)} {_task_report_date(record):%Y%m%d}"


def _ensure_local_upload_folder(
    configs: dict[str, dict[str, Any]], record: dingtalk_table.TaskRecord
) -> str:
    report_date = _task_report_date(record)
    docs_config = configs.get("dingtalk_docs", {})
    brand = _clean_folder_text(_task_brand(record))
    products = [
        cleaned
        for item in _task_products(record)
        if (cleaned := _clean_folder_text(item))
    ]
    if not brand:
        raise RuntimeError("“竞品品牌”为空，无法创建钉钉归档目录。")
    if not products:
        raise RuntimeError("“关注新品”为空，无法创建钉钉归档目录。")
    date_suffix = report_date.strftime("%Y%m%d")
    month_id, _ = dingtalk_docs.ensure_local_upload_month_folder(
        docs_config, report_date.year, report_date.month
    )
    matches = [
        (folder_id, name)
        for folder_id, name in dingtalk_docs.child_folders(docs_config, month_id)
        if brand in name
        and all(product in name for product in products)
        and name.endswith(f" {date_suffix}")
    ]
    if len(matches) == 1:
        return matches[0][0]
    if len(matches) > 1:
        names = "、".join(name for _, name in matches)
        raise RuntimeError(
            f"同一月份存在多个同时包含品牌、产品名和报告日期的文件夹：{names}。"
        )
    folder_id, _ = dingtalk_docs.ensure_child_folder(
        docs_config, month_id, _task_folder_name(record)
    )
    return folder_id


def _clean_folder_text(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|\r\n]+', "_", value).strip()


def _social_cleaned_input_filename(brand: str, product: str, report_date: date) -> str:
    parts = [_clean_folder_text(value) for value in (brand, product)]
    if not all(parts):
        raise RuntimeError("品牌或新品无法生成社媒清洗数据归档文件名。")
    return f"{parts[0]}-{report_date:%Y%m%d}-{parts[1]}-社媒清洗数据.xlsx"


def _ensure_dingtalk_input_attachment(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    folder_id: str | None,
) -> str | None:
    value = record.cells.get(field_key)
    if _has_dingtalk_node_link(value):
        items = _value_items(value)
        node_count = sum(
            bool(dingtalk_docs.node_id_from_url(dingtalk_docs.extract_link(item)))
            for item in items
        )
        if len(items) != 1 or node_count != 1:
            raise RuntimeError(f"“{field_label}”字段仅支持一个钉钉附件或文件夹链接。")
        docs_config = configs.get("dingtalk_docs", {})
        source_folder = dingtalk_docs.folder_id_for_linked_node(docs_config, value)
        target_folder = folder_id or source_folder
        if not target_folder or source_folder == target_folder:
            return target_folder
        target_dir = output_root(configs) / record.record_id
        if field_key == "deliveryData":
            target_dir /= "raw_data"
        local_path = dingtalk_docs.download_linked_file(
            docs_config, value, target_dir
        )
        if not local_path or local_path.suffix.lower() != ".xlsx":
            raise RuntimeError(f"“{field_label}”字段不是可下载的 xlsx 附件。")
        url = dingtalk_docs.upload_file(docs_config, local_path, target_folder)
        _update_input_file_reference(configs, record, field_key, local_path.name, url)
        return target_folder
    ai_table_attachment = _single_ai_table_attachment(value, field_label)
    local_path = (
        _download_ai_table_attachment(
            configs, record, field_key, field_label, ai_table_attachment
        )
        if ai_table_attachment
        else _single_local_xlsx(value, field_label)
    )
    if not folder_id:
        folder_id = _ensure_local_upload_folder(configs, record)
    local_path = _cache_promoted_input(configs, record, field_key, local_path)
    url = dingtalk_docs.upload_file(configs.get("dingtalk_docs", {}), local_path, folder_id)
    _update_input_file_reference(configs, record, field_key, local_path.name, url)
    return folder_id


def _update_input_file_reference(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    name: str,
    url: str,
) -> None:
    cell_type = str(
        dingtalk_table.field_defs(configs).get(field_key, {}).get("cellType") or ""
    ).strip().lower()
    if cell_type == "link" or (not cell_type and field_key in dingtalk_table.LINK_KEYS):
        dingtalk_table.update_link_fields(
            configs, record.record_id, {field_key: (name, url)}
        )
        record.cells[field_key] = dingtalk_table.link_cell(name, url)
        return
    dingtalk_table.update_attachment_fields(
        configs, record.record_id, {field_key: (name, url)}
    )
    record.cells[field_key] = dingtalk_table.attachment_cell(name, url)


def _cache_promoted_input(
    configs: dict[str, dict[str, Any]], record: dingtalk_table.TaskRecord, field_key: str, source: Path
) -> Path:
    task_root = output_root(configs) / record.record_id
    target_dir = task_root / "raw_data" if field_key == "deliveryData" else task_root
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / source.name
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)
    return target


def _task_annotation_path(configs: dict[str, dict[str, Any]], record: dingtalk_table.TaskRecord) -> Path | None:
    value = record.cells.get("productMenu")
    if not _has_dingtalk_node_link(value):
        raise RuntimeError("“产品清单”字段缺少可识别的钉钉 xlsx 附件。")
    if value:
        downloaded = dingtalk_docs.download_linked_file(
            configs.get("dingtalk_docs", {}),
            value,
            output_root(configs) / record.record_id,
        )
        if downloaded:
            return downloaded
    candidates = sorted((output_root(configs) / record.record_id).glob("*产品清单.xlsx"))
    return candidates[-1] if candidates else None

def _consumer_feedback_configs(configs: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    scoped = {key: dict(value) for key, value in configs.items()}
    scoped["dingtalk"] = dict(configs["dingtalk"])
    scoped["dingtalk"]["tableId"] = CONSUMER_FEEDBACK_TABLE_ID
    scoped["field_mapping"] = {
        "fields": {
            "brand": {"fieldId": "01ZM8y7"},
            "productName": {"fieldId": "mHe1U1b"},
            "launchDate": {"fieldId": "mKUEya0"},
            "day30": {"fieldId": "SSrz1N1"},
            "socialCleanedRawData": {
                "fieldId": "Q4mbLWx",
                "cellType": "attachment",
            },
            "report": {"fieldId": "amWTton", "cellType": "link"},
            "feedback": {"fieldId": "j8IgB7P"},
        }
    }
    return scoped


def _social_task_date(record: dingtalk_table.TaskRecord, key: str, label: str) -> date:
    value = record.cells.get(key)
    while isinstance(value, (dict, list)):
        if isinstance(value, list):
            value = value[0] if value else None
        else:
            value = value.get("value") or value.get("date") or value.get("text")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        timestamp = float(value)
        if timestamp > 10_000_000_000:
            timestamp /= 1000
        return datetime.fromtimestamp(timestamp).date()
    text = str(value or "").strip()
    if not text:
        raise RuntimeError(f"“{label}”为空，无法生成社媒评论统计表。")
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-")
    try:
        return datetime.fromisoformat(normalized.replace("Z", "+00:00")).date()
    except ValueError as exc:
        raise RuntimeError(f"“{label}”格式无效：{text}") from exc


def _download_linked_social_file(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    target_dir: Path | None = None,
) -> Path:
    directory = target_dir or (
        output_root(configs) / record.record_id / "social_inputs" / field_key
    )
    downloaded = dingtalk_docs.download_linked_file(
        configs.get("dingtalk_docs", {}), record.cells.get(field_key), directory
    )
    if not downloaded or downloaded.suffix.lower() != ".xlsx":
        raise RuntimeError(f"“{field_label}”字段不是可下载的 xlsx 附件。")
    return downloaded


def _download_social_input_source(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    field_key: str,
    field_label: str,
    target_dir: Path | None = None,
) -> Path | None:
    """Download a social input without archiving or changing AI Table cells."""
    value = record.cells.get(field_key)
    if not _value_items(value):
        return None
    if _has_dingtalk_node_link(value):
        try:
            items = _value_items(value)
            node_count = sum(
                bool(
                    dingtalk_docs.node_id_from_url(
                        dingtalk_docs.extract_link(item)
                    )
                )
                for item in items
            )
            if len(items) != 1 or node_count != 1:
                raise RuntimeError(
                    f"“{field_label}”字段仅支持一个钉钉 xlsx 附件。"
                )
            return _download_linked_social_file(
                configs,
                record,
                field_key,
                field_label,
                target_dir=target_dir,
            )
        except Exception as exc:
            filename = (
                _local_item_name(items[0]) if items else f"{field_label}.xlsx"
            )
            raise ArchivedSocialSourceUnreadable(filename) from exc

    attachment = _wait_for_ai_table_attachment_url(
        configs, record, field_key, field_label
    )
    if attachment is None:
        raise RuntimeError(
            f"“{field_label}”字段只支持钉钉文档节点或 AI 表“+”上传的单个 xlsx 附件。"
        )
    return _download_ai_table_attachment(
        configs,
        record,
        field_key,
        field_label,
        attachment,
        target_dir=target_dir,
    )


def _existing_field_ids(configs: dict[str, dict[str, Any]]) -> set[str]:
    dingtalk = configs["dingtalk"]
    data = dingtalk_table.call_table_tool(
        dingtalk,
        "get_fields",
        {"baseId": dingtalk["baseId"], "tableId": dingtalk["tableId"]},
    )
    fields = data.get("data", {}).get("fields") or data.get("data") or []
    return {str(field.get("fieldId") or field.get("id")) for field in fields if isinstance(field, dict)}


def _filter_links_for_existing_fields(configs: dict[str, dict[str, Any]], links: dict[str, tuple[str, str]]) -> dict[str, tuple[str, str]]:
    existing = _existing_field_ids(configs)
    return {key: value for key, value in links.items() if dingtalk_table.field_id(configs, key) in existing}


def build_progress_callback(configs: dict[str, dict[str, Any]], record_id: str):
    def update(message: str) -> None:
        try:
            dingtalk_table.update_feedback(configs, record_id, message)
        except Exception as exc:
            print(f"WARNING: 进度回写失败：{message} {type(exc).__name__}: {exc}", file=sys.stderr)

    return update


def _empty_delivery_platforms(outputs: dict[str, Path]) -> list[str]:
    empty: list[str] = []
    for key, platform in DELIVERY_OUTPUT_PLATFORMS:
        path = outputs[key]
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            has_product = any(
                str(row[0] or "").strip()
                for row in wb.active.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True)
            )
        finally:
            wb.close()
        if not has_product:
            empty.append(platform)
    return empty


def _delivery_completion_message(empty_platforms: list[str]) -> str:
    if not empty_platforms:
        return "外卖数据统计完毕"
    return f"外卖数据统计完毕，{'、'.join(empty_platforms)}无数据"


def _format_report_completion_feedback(warnings: list[str]) -> str:
    unique_warnings = list(dict.fromkeys(warnings))
    social_no_data_products: list[str] = []
    consumer_missing_products: list[str] = []
    delivery_warnings: list[str] = []
    other_warnings: list[str] = []
    platform_missing_suffixes = tuple(
        f"{platform}无数据" for platform in ("大众点评", "微博", "小红书", "抖音", "B站")
    )

    for warning in unique_warnings:
        social_match = re.fullmatch(r"(.+)社媒无数据", warning)
        if social_match:
            social_no_data_products.append(social_match.group(1))
            continue
        consumer_match = re.fullmatch(r"(.+)：消费者反馈统计暂无法获取", warning)
        if consumer_match:
            consumer_missing_products.append(consumer_match.group(1))
            continue
        if warning.endswith(platform_missing_suffixes):
            continue
        if warning.endswith("外卖无数据"):
            delivery_warnings.append(warning)
            continue
        other_warnings.append(warning)

    segments = delivery_warnings
    if social_no_data_products:
        product_labels = "".join(f"[{product}]" for product in social_no_data_products)
        segments.append(f"{product_labels}社媒无数据")
    social_no_data_set = set(social_no_data_products)
    consumer_only = [
        product
        for product in consumer_missing_products
        if product not in social_no_data_set
    ]
    if consumer_only:
        product_labels = "".join(f"[{product}]" for product in consumer_only)
        segments.append(f"{product_labels}消费者反馈统计暂无法获取")
    segments.extend(other_warnings)
    return "报告已生成" if not segments else f"报告已生成：{'；'.join(segments)}"


def _prepare_delivery_run_inputs(
    configs: dict[str, dict[str, Any]],
    record: dingtalk_table.TaskRecord,
    workspace: Path,
    *,
    progress_callback=None,
    pre_annotation_callback=None,
    auto_menu_callback=None,
) -> tuple[Path, Path, list[Any]]:
    delivery_dir = workspace / "delivery"
    delivery_path = _download_uploaded_xlsx(
        configs, record, "deliveryData", "外卖数据", delivery_dir
    )
    assert delivery_path is not None
    raw_rows = find_delivery_rows(delivery_dir)
    if not raw_rows:
        raise RuntimeError("外卖数据附件中没有可识别的原始数据。")

    if pre_annotation_callback:
        pre_annotation_callback()

    menu_value = record.cells.get("productMenu")
    manual_attachment = _single_ai_table_attachment(
        menu_value, "产品清单", require_url=False
    )
    auto_annotation_path: Path | None = None
    if not manual_attachment or auto_menu_callback:
        if progress_callback:
            message = "3/6 正在自动标注在售不满30天的产品上新日期"
            if manual_attachment:
                message = "3/6 正在生成产品清单样件并读取人工配置的产品清单"
            progress_callback(message)
        auto_annotation_path = prepare_product_menu(
            record.record_id,
            delivery_dir,
            workspace / "auto_product_menu",
            report_date=_task_report_date(record),
        )
        if auto_menu_callback:
            auto_menu_callback(auto_annotation_path)
    if manual_attachment:
        if progress_callback and not auto_menu_callback:
            progress_callback("3/6 正在读取人工配置的产品清单")
        annotation_path = _download_uploaded_xlsx(
            configs,
            record,
            "productMenu",
            "产品清单",
            workspace / "product_menu",
        )
        assert annotation_path is not None
    else:
        assert auto_annotation_path is not None
        annotation_path = auto_annotation_path
    return delivery_dir, annotation_path, raw_rows


def _matching_social_records(
    configs: dict[str, dict[str, Any]],
    *,
    brand: str,
    report_date: date,
    products: list[str],
) -> dict[str, tuple[str, dingtalk_table.TaskRecord, date]]:
    requested = {normalize_social_product_key(item): item for item in products}
    matches: dict[str, tuple[str, dingtalk_table.TaskRecord, date]] = {}
    for record in dingtalk_table.fetch_records(configs):
        if _task_brand(record) != brand:
            continue
        try:
            day30 = _social_task_date(record, "day30", "30日")
        except RuntimeError:
            continue
        if day30 != report_date:
            continue
        launch_date = _social_task_date(record, "launchDate", "上市日期")
        for product in _task_products(record):
            key = normalize_social_product_key(product)
            if key not in requested:
                continue
            if key in matches:
                raise RuntimeError(
                    f"新品“{requested[key]}”匹配到多条消费者反馈记录。"
                )
            matches[key] = (requested[key], record, launch_date)
    return matches


def _collect_social_inputs(
    configs: dict[str, dict[str, Any]],
    matches: dict[str, tuple[str, dingtalk_table.TaskRecord, date]],
    workspace: Path,
    *,
    skip_archived: bool = False,
) -> dict[str, Path]:
    assignments: dict[str, Path] = {}
    unreadable_archived_files: list[str] = []
    field_key, field_label = SOCIAL_INPUT_PLATFORMS[0]
    for product_key, (product, record, _) in matches.items():
        value = record.cells.get(field_key)
        if not _value_items(value):
            continue
        if skip_archived and _has_dingtalk_node_link(value):
            continue
        try:
            source_path = _download_social_input_source(
                configs,
                record,
                field_key,
                field_label,
                target_dir=(
                    workspace / "social_sources" / record.record_id / field_key
                ),
            )
        except ArchivedSocialSourceUnreadable as exc:
            unreadable_archived_files.append(exc.filename)
            continue
        assert source_path is not None
        source_products = social_cleaned_workbook_products(source_path)
        unknown = [
            source_product
            for key, source_product in source_products.items()
            if key != product_key
        ]
        if unknown:
            raise RuntimeError(
                f"“{field_label}”文件包含不属于当前新品“{product}”的数据："
                f"{'、'.join(unknown)}。"
            )
        if source_products and product_key not in source_products:
            raise RuntimeError(
                f"“{field_label}”文件不包含当前新品“{product}”。"
            )
        assignments[product_key] = source_path

    if unreadable_archived_files:
        filenames = "、".join(
            f"[{name}]" for name in dict.fromkeys(unreadable_archived_files)
        )
        raise RuntimeError(
            f"{filenames}已归档为钉钉文档链接导致读取失败，"
            "请重新从本地上传以上原始文件，再进行报告生成。"
        )

    return assignments


def _generate_social_outputs(
    configs: dict[str, dict[str, Any]],
    matches: dict[str, tuple[str, dingtalk_table.TaskRecord, date]],
    cleaned_files: dict[str, Path],
    output_dir: Path,
) -> tuple[dict[str, Path], dict[str, Any]]:
    outputs: dict[str, Path] = {}
    models: dict[str, Any] = {}
    for product_key, (product, record, launch_date) in matches.items():
        cleaned_file = cleaned_files.get(product_key)
        if cleaned_file is None:
            continue
        end_date = _social_task_date(record, "day30", "30日")
        period = f"{launch_date.month}.{launch_date.day}-{end_date.month}.{end_date.day}"
        summaries = summarize_social_cleaned_workbook(cleaned_file)
        output = generate_consumer_feedback_tables(
            record.record_id,
            cleaned_file,
            brand=_task_brand(record),
            product=product,
            start_date=f"{launch_date.month}.{launch_date.day}",
            end_date=f"{end_date.month}.{end_date.day}",
            output_dir=output_dir / record.record_id,
        )
        outputs[product_key] = output
        models[product] = social_report_from_summaries(
            title=f"{product} {period} 第三方平台评价反馈",
            period=period,
            summaries=summaries,
        )
    return outputs, models


def run_prepare_product_menu(record_id: str) -> dict[str, Any]:
    configs = load_configs()
    progress_callback = build_progress_callback(configs, record_id)
    try:
        dingtalk_table.mark_status(configs, record_id, _status(configs, "productMenuRunning", "产品清单生成中"))
        record = dingtalk_table.fetch_record(configs, record_id)
        folder_id = _ensure_local_upload_folder(configs, record)
        dingtalk_table.clear_attachment_fields(configs, record_id, ("productMenu",))
        record.cells["productMenu"] = []
        progress_callback("1/4 正在读取外卖数据")
        with tempfile.TemporaryDirectory(prefix="feedback-product-menu-") as tmp:
            input_dir = Path(tmp) / "delivery"
            _download_uploaded_xlsx(
                configs, record, "deliveryData", "外卖数据", input_dir
            )
            output = prepare_product_menu(
                record_id,
                input_dir,
                output_root(configs) / record_id,
                progress_callback=progress_callback,
                report_date=_task_report_date(record),
            )
        progress_callback("4/4 正在上传到钉钉文档")
        url = dingtalk_docs.upload_file(configs.get("dingtalk_docs", {}), output, folder_id)
        dingtalk_table.mark_links(
            configs,
            record_id,
            {"productMenu": (output.name, url)},
            _status(configs, "waitingAnnotation", "待标注"),
        )
        progress_callback("已提取产品清单，请确认在售不满30日的新品上新日期标注无遗漏后，再进行外卖数据统计")
        return {"ok": True, "stage": "prepare-product-menu", "recordId": record_id, "output": str(output), "url": url}
    except Exception as exc:
        message = f"{type(exc).__name__}: {exc}"
        try:
            progress_callback(f"生成失败：{message}")
            dingtalk_table.mark_failed(configs, record_id, message)
        except Exception:
            message += "\n\n回写失败状态也失败：\n" + traceback.format_exc()
        raise RuntimeError(message) from exc


def run_generate_delivery_tables(record_id: str) -> dict[str, Any]:
    configs = load_configs()
    progress_callback = build_progress_callback(configs, record_id)
    try:
        dingtalk_table.mark_status(configs, record_id, _status(configs, "dataTablesRunning", "外卖数据统计中"))
        record = dingtalk_table.fetch_record(configs, record_id)
        folder_id = _ensure_local_upload_folder(configs, record)
        dingtalk_table.clear_link_fields(configs, record_id, DELIVERY_OUTPUT_FIELDS)
        for key in DELIVERY_OUTPUT_FIELDS:
            record.cells[key] = ""
        progress_callback("1/3 正在读取 AI 表上传的外卖数据和可选产品清单")
        with tempfile.TemporaryDirectory(prefix="feedback-delivery-") as tmp:
            input_dir, annotation_path, _ = _prepare_delivery_run_inputs(
                configs,
                record,
                Path(tmp),
                progress_callback=progress_callback,
            )
            progress_callback("2/3 正在生成外卖数表")
            outputs = generate_delivery_tables(
                record_id,
                input_dir,
                annotation_path,
                output_root(configs) / record_id / "delivery_tables",
            )
        empty_platforms = _empty_delivery_platforms(outputs)
        empty_keys = {
            key for key, platform in DELIVERY_OUTPUT_PLATFORMS if platform in empty_platforms
        }
        generated_outputs = {
            key: path for key, path in outputs.items() if key not in empty_keys
        }
        for key in empty_keys:
            path = outputs.get(key)
            if path and path.exists():
                path.unlink()
        progress_callback("3/3 正在上传外卖数表到钉钉文档")
        links: dict[str, tuple[str, str]] = {}
        for key, path in generated_outputs.items():
            links[key] = (path.name, dingtalk_docs.upload_file(configs.get("dingtalk_docs", {}), path, folder_id))
        writable_links = _filter_links_for_existing_fields(configs, links) if links else {}
        dingtalk_table.mark_links(
            configs,
            record_id,
            writable_links,
            _status(configs, "waitingReport", "外卖数据已生成"),
        )
        progress_callback(_delivery_completion_message(empty_platforms))
        return {
            "ok": True,
            "stage": "generate-delivery-tables",
            "recordId": record_id,
            "outputs": {key: str(path) for key, path in generated_outputs.items()},
            "emptyPlatforms": empty_platforms,
        }
    except Exception as exc:
        message = f"{type(exc).__name__}: {exc}"
        try:
            dingtalk_table.mark_failed(configs, record_id, message)
        except Exception:
            message += "\n\n回写失败状态也失败：\n" + traceback.format_exc()
        raise RuntimeError(message) from exc


def run_finalize_report(record_id: str) -> dict[str, Any]:
    configs = load_configs()
    progress_callback = build_progress_callback(configs, record_id)
    try:
        progress_callback("1/5 正在启动归档")
        record = dingtalk_table.fetch_record(configs, record_id)
        if not record.cells.get("report"):
            raise RuntimeError("最终报告链接为空，请先生成并检查报告。")
        brand = _task_brand(record)
        report_date = _task_report_date(record)
        products = _task_products(record)
        if not brand or not products:
            raise RuntimeError("归档缺少竞品品牌或关注新品。")
        progress_callback("2/5 正在校验并整理待归档文件")
        report_url = dingtalk_docs.extract_link(record.cells.get("report"))
        social_configs = _consumer_feedback_configs(configs)
        social_matches = _matching_social_records(
            social_configs,
            brand=brand,
            report_date=report_date,
            products=products,
        )
        progress_callback("3/5 正在创建目录并上传原始文件")
        folder_id = _ensure_local_upload_folder(configs, record)
        folder_url = dingtalk_docs.node_url(folder_id)

        main_uploads: dict[str, tuple[str, str]] = {}
        social_uploads: dict[str, dict[str, tuple[str, str]]] = {}
        archived_files: list[str] = []
        crawl_date = report_date
        with tempfile.TemporaryDirectory(prefix="feedback-finalize-") as tmp:
            workspace = Path(tmp)
            delivery_value = record.cells.get("deliveryData")
            if not _has_dingtalk_node_link(delivery_value):
                delivery_path = _download_uploaded_xlsx(
                    configs,
                    record,
                    "deliveryData",
                    "外卖数据",
                    workspace / "delivery",
                )
                assert delivery_path is not None
                raw_rows = find_delivery_rows(delivery_path.parent)
                if not raw_rows:
                    raise RuntimeError("外卖数据附件中没有可识别的原始数据。")
                crawl_date = _crawl_date(raw_rows)
                normalized = delivery_path.with_name(
                    safe_delivery_filename(
                        f"{_primary_brand(raw_rows)}-{crawl_date:%Y%m%d}-外卖清洗数据.xlsx"
                    )
                )
                if normalized != delivery_path:
                    shutil.copy2(delivery_path, normalized)
                url = dingtalk_docs.upload_file(
                    configs.get("dingtalk_docs", {}), normalized, folder_id
                )
                main_uploads["deliveryData"] = (normalized.name, url)
                archived_files.append(normalized.name)
            else:
                current_name = _local_item_name(_value_items(delivery_value)[0])
                match = re.search(r"(20\d{6})", current_name)
                if match:
                    crawl_date = datetime.strptime(match.group(1), "%Y%m%d").date()

            menu_value = record.cells.get("productMenu")
            if _value_items(menu_value) and not _has_dingtalk_node_link(menu_value):
                menu_path = _download_uploaded_xlsx(
                    configs,
                    record,
                    "productMenu",
                    "产品清单",
                    workspace / "product_menu",
                )
                assert menu_path is not None
                normalized = menu_path.with_name(
                    safe_delivery_filename(
                        f"{brand}-{crawl_date:%Y%m%d}-产品清单.xlsx"
                    )
                )
                if normalized != menu_path:
                    shutil.copy2(menu_path, normalized)
                url = dingtalk_docs.upload_file(
                    configs.get("dingtalk_docs", {}), normalized, folder_id
                )
                main_uploads["productMenu"] = (normalized.name, url)
                archived_files.append(normalized.name)

            social_files = _collect_social_inputs(
                social_configs,
                social_matches,
                workspace,
                skip_archived=True,
            )
            field_key, _ = SOCIAL_INPUT_PLATFORMS[0]
            for product_key, path in social_files.items():
                product, social_record, _ = social_matches[product_key]
                if _has_dingtalk_node_link(social_record.cells.get(field_key)):
                    continue
                normalized = path.with_name(
                    _social_cleaned_input_filename(brand, product, report_date)
                )
                if normalized != path:
                    shutil.copy2(path, normalized)
                url = dingtalk_docs.upload_file(
                    configs.get("dingtalk_docs", {}), normalized, folder_id
                )
                social_uploads.setdefault(social_record.record_id, {})[
                    field_key
                ] = (normalized.name, url)
                archived_files.append(normalized.name)

        progress_callback("4/5 正在回写归档链接")
        if main_uploads:
            dingtalk_table.update_attachment_fields(
                configs, record_id, main_uploads
            )
        for social_record_id, links in social_uploads.items():
            dingtalk_table.update_attachment_fields(
                social_configs, social_record_id, links
            )
        dingtalk_table.update_link_fields(
            configs,
            record_id,
            {
                "folder": (
                    _task_folder_name(record),
                    folder_url,
                )
            },
        )
        message = "5/5 原始文件归档完成"
        if archived_files:
            message += f"：{'、'.join(archived_files)}"
        else:
            message += "，所有输入已是钉钉文档链接"
        progress_callback(message)
        return {
            "ok": True,
            "status": "completed",
            "recordId": record_id,
            "message": "原始文件归档完成",
            "reportUrl": report_url,
            "folderUrl": folder_url,
            "archivedFileCount": len(archived_files),
            "archivedFileNames": "、".join(archived_files),
        }
    except Exception as exc:
        message = f"原始文件归档失败：{type(exc).__name__}: {exc}"
        try:
            progress_callback(message)
        except Exception:
            message += "\n\n回写失败说明也失败：\n" + traceback.format_exc()
        raise RuntimeError(message) from exc


def run_generate_report(record_id: str) -> dict[str, Any]:
    configs = load_configs()
    progress_callback = build_progress_callback(configs, record_id)
    current_stage = "读取当前行信息"
    try:
        dingtalk_table.mark_status(configs, record_id, _status(configs, "reportRunning", "报告生成中"))
        progress_callback("1/6 正在读取品牌、新品及报告日期")
        record = dingtalk_table.fetch_record(configs, record_id)
        brand = _task_brand(record)
        if not brand:
            raise RuntimeError("“竞品品牌”为空，无法生成报告。")
        report_date = _task_report_date(record)
        grouped_products = _task_products(record)
        if not grouped_products:
            raise RuntimeError("“关注新品”为空，无法生成报告。")
        social_configs = _consumer_feedback_configs(configs)
        folder_id = _ensure_local_upload_folder(configs, record)
        dingtalk_table.clear_link_fields(configs, record_id, REPORT_OUTPUT_FIELDS)
        warnings: list[str] = []
        with tempfile.TemporaryDirectory(prefix="feedback-report-") as tmp:
            workspace = Path(tmp)
            current_stage = "读取原始附件"
            progress_callback("2/6 正在读取上传的外卖及社媒数据")
            social_matches = _matching_social_records(
                social_configs,
                brand=brand,
                report_date=report_date,
                products=grouped_products,
            )
            for product in grouped_products:
                if normalize_social_product_key(product) not in social_matches:
                    warnings.append(f"{product}社媒无数据")
            social_files: dict[str, Path] = {}

            def download_social_sources() -> None:
                nonlocal social_files
                social_files = _collect_social_inputs(
                    social_configs, social_matches, workspace
                )
                for product_key, (product, _, _) in social_matches.items():
                    if product_key not in social_files:
                        warnings.append(f"{product}社媒无数据")

            input_dir, annotation_path, raw_rows = _prepare_delivery_run_inputs(
                configs,
                record,
                workspace,
                progress_callback=progress_callback,
                pre_annotation_callback=download_social_sources,
                auto_menu_callback=lambda path: dingtalk_docs.upload_file(
                    configs.get("dingtalk_docs", {}), path, folder_id
                ),
            )

            current_stage = "生成数表"
            progress_callback("4/6 正在进行外卖及社媒数据统计")
            annotations = read_annotation(annotation_path)
            delivery_statuses = tracked_product_statuses(
                annotations, brand, grouped_products
            )
            meituan_metrics = platform_delivery_metrics(
                raw_rows, annotations, "美团"
            )
            eleme_metrics = platform_delivery_metrics(
                raw_rows, annotations, "饿了么"
            )
            jd_metrics = jd_delivery_metrics(raw_rows, annotations)
            delivery_model = delivery_report_from_metrics(
                meituan_metrics, eleme_metrics, grouped_products
            )
            jd_model = jd_report_from_metrics(jd_metrics)
            delivery_outputs = generate_delivery_tables(
                record_id,
                input_dir,
                annotation_path,
                workspace / "outputs" / record_id / "delivery_tables",
            )
            empty_platforms = _empty_delivery_platforms(delivery_outputs)
            delivery_links: dict[str, tuple[str, str]] = {}
            for key, label in DELIVERY_OUTPUT_PLATFORMS:
                path = delivery_outputs[key]
                if label in empty_platforms:
                    if path.exists():
                        path.unlink()
                    if label == "京东":
                        warnings.append("京东外卖无数据")
                    continue
                delivery_links[key] = (
                    path.name,
                    dingtalk_docs.upload_file(
                        configs.get("dingtalk_docs", {}), path, folder_id
                    ),
                )
            writable_delivery_links = (
                _filter_links_for_existing_fields(configs, delivery_links)
                if delivery_links
                else {}
            )
            if writable_delivery_links:
                dingtalk_table.mark_links(
                    configs, record_id, writable_delivery_links
                )

            social_outputs, social_models = _generate_social_outputs(
                social_configs,
                social_matches,
                social_files,
                workspace / "outputs" / record_id / "consumer_feedback_tables",
            )
            launch_dates = {
                product: launch_date
                for product, _, launch_date in social_matches.values()
            }
            for product_key, output in social_outputs.items():
                product, social_record, _ = social_matches[product_key]
                url = dingtalk_docs.upload_file(
                    configs.get("dingtalk_docs", {}), output, folder_id
                )
                dingtalk_table.mark_links(
                    social_configs,
                    social_record.record_id,
                    {"report": (output.name, url)},
                )

            current_stage = "生成报告"
            progress_callback("5/6 正在生成跟踪反馈报告")
            result = generate_report(
                record_id,
                brand,
                grouped_products,
                report_date=report_date,
                social_paths={},
                launch_dates=launch_dates,
                delivery_statuses=delivery_statuses,
                configs=configs,
                delivery_report=delivery_model,
                jd_report=jd_model,
                social_reports=social_models,
                output_dir=workspace / "outputs" / record_id,
            )
            current_stage = "上传报告"
            progress_callback("6/6 正在上传至钉钉文档")
            if not result.pdf_path or not result.pdf_path.is_file():
                raise RuntimeError("报告生成未产生 PDF 文件。")
            dingtalk_docs.upload_file(
                configs.get("dingtalk_docs", {}), result.path, folder_id
            )
            url = dingtalk_docs.upload_file(
                configs.get("dingtalk_docs", {}), result.pdf_path, folder_id
            )
            html_file_name = result.path.name
            output_file_name = result.pdf_path.name
        dingtalk_table.mark_links(
            configs,
            record_id,
            {
                "report": (output_file_name, url),
            },
            _status(configs, "done", "已生成"),
        )
        all_warnings = list(dict.fromkeys([*warnings, *result.warnings]))
        feedback = _format_report_completion_feedback(all_warnings)
        progress_callback(feedback)
        return {
            "ok": True,
            "stage": "generate-report",
            "recordId": record_id,
            "output": None,
            "outputFile": output_file_name,
            "htmlFile": html_file_name,
            "url": url,
            "warnings": all_warnings,
        }
    except Exception as exc:
        detail = str(exc).strip() or type(exc).__name__
        feedback = f"报告生成失败（{current_stage}）：{detail}"
        message = f"{type(exc).__name__}: {exc}"
        try:
            dingtalk_table.mark_failed(configs, record_id, feedback)
        except Exception:
            message += "\n\n回写失败状态也失败：\n" + traceback.format_exc()
        raise RuntimeError(message) from exc
