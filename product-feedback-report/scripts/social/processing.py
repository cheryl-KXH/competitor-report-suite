from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


SOCIAL_RAW_PLATFORMS = (
    ("weibo", "微博"),
    ("xiaohongshu", "小红书"),
    ("douyin", "抖音"),
    ("bilibili", "B站"),
)
@dataclass(frozen=True)
class PlatformSummary:
    key: str
    label: str
    positive_tags: tuple[tuple[str, int], ...]
    negative_tags: tuple[tuple[str, int], ...]
    positive_users: int
    negative_users: int


def _text(value: Any) -> str:
    return str(value or "").strip()


def _header_index(header: list[Any], name: str) -> int | None:
    normalize = lambda value: re.sub(
        r"\s+", "", unicodedata.normalize("NFKC", _text(value))
    )
    normalized = [normalize(value) for value in header]
    target = normalize(name)
    return next((index for index, value in enumerate(normalized) if value == target), None)


def normalize_social_product_key(value: Any) -> str:
    """Normalize product labels without allowing fuzzy/substring matching."""
    text = unicodedata.normalize("NFKC", _text(value)).casefold()
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", "", text)


def social_workbook_products(path: Path) -> dict[str, str]:
    """Return normalized product keys and their first source labels."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        header = [cell.value for cell in worksheet[1]]
        product_index = _header_index(header, "产品名称")
        if product_index is None:
            raise RuntimeError(f"社媒文件缺少“产品名称”列：{path.name}")

        products: dict[str, str] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(_text(value) for value in row):
                continue
            raw_product = row[product_index] if product_index < len(row) else None
            label = _text(raw_product)
            key = normalize_social_product_key(label)
            if not key:
                raise RuntimeError(
                    f"社媒文件第 {row_number} 行“产品名称”为空：{path.name}"
                )
            products.setdefault(key, label)
        if not products:
            raise RuntimeError(f"社媒文件没有可处理的数据行：{path.name}")
        return products
    finally:
        workbook.close()


def social_cleaned_workbook_products(path: Path) -> dict[str, str]:
    """Return products found across the four raw-data sheets in a cleaned workbook."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        products: dict[str, str] = {}
        missing_sheets = [label for _, label in SOCIAL_RAW_PLATFORMS if label not in workbook.sheetnames]
        if missing_sheets:
            raise RuntimeError(
                f"社媒清洗数据缺少工作表：{'、'.join(missing_sheets)}（文件：{path.name}）"
            )
        for _, label in SOCIAL_RAW_PLATFORMS:
            worksheet = workbook[label]
            header = [cell.value for cell in worksheet[1]]
            product_index = _header_index(header, "产品名称")
            if product_index is None:
                raise RuntimeError(
                    f"社媒清洗数据的“{label}”工作表缺少“产品名称”列：{path.name}"
                )
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(_text(value) for value in row):
                    continue
                raw_product = row[product_index] if product_index < len(row) else None
                product = _text(raw_product)
                product_key = normalize_social_product_key(product)
                if not product_key:
                    raise RuntimeError(
                        f"社媒清洗数据的“{label}”工作表第 {row_number} 行"
                        f"“产品名称”为空：{path.name}"
                    )
                products.setdefault(product_key, product)
        return products
    finally:
        workbook.close()


def split_social_workbook(
    path: Path,
    *,
    product_names: dict[str, str],
    output_dir: Path,
    output_names: dict[str, str],
) -> dict[str, Path]:
    """Create one formatting-preserving workbook per product in the source file."""
    source_products = social_workbook_products(path)
    unknown = [source_products[key] for key in source_products if key not in product_names]
    if unknown:
        raise RuntimeError(
            f"社媒文件包含无法匹配到同品牌同一“30日”记录的产品：{'、'.join(unknown)}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    outputs: dict[str, Path] = {}
    for product_key in source_products:
        workbook = load_workbook(path, data_only=False)
        try:
            worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
            header = [cell.value for cell in worksheet[1]]
            product_index = _header_index(header, "产品名称")
            if product_index is None:
                raise RuntimeError(f"社媒文件缺少“产品名称”列：{path.name}")
            for row_number in range(worksheet.max_row, 1, -1):
                row_values = [cell.value for cell in worksheet[row_number]]
                if not any(_text(value) for value in row_values):
                    worksheet.delete_rows(row_number)
                    continue
                raw_product = (
                    row_values[product_index]
                    if product_index < len(row_values)
                    else None
                )
                if normalize_social_product_key(raw_product) != product_key:
                    worksheet.delete_rows(row_number)
            output_path = output_dir / output_names[product_key]
            workbook.save(output_path)
            outputs[product_key] = output_path
        finally:
            workbook.close()
    return outputs


def summarize_social_rows(key: str, label: str, rows: Iterable[Iterable[Any]]) -> PlatformSummary:
    values = [list(row) for row in rows]
    if not values:
        return PlatformSummary(key, label, (), (), 0, 0)
    header = values[0]
    sentiment_index = _header_index(header, "情感识别")
    if sentiment_index is None:
        raise RuntimeError(f"{label}数据缺少“情感识别”列。")

    label_columns: list[int] = []
    normalized = [_text(value).replace(" ", "") for value in header]
    for index, column_name in enumerate(normalized):
        if re.fullmatch(r"评价\d+-对应标签", column_name):
            label_columns.append(index)
    if not label_columns:
        raise RuntimeError(f"{label}数据缺少“评价n-对应标签”列。")

    positive_tags: Counter[str] = Counter()
    negative_tags: Counter[str] = Counter()
    positive_users = 0
    negative_users = 0
    for row in values[1:]:
        sentiment = _text(row[sentiment_index] if sentiment_index < len(row) else "")
        if sentiment == "正向":
            positive_users += 1
        elif sentiment == "负向":
            negative_users += 1
        for tag_index in label_columns:
            tag = _text(row[tag_index] if tag_index < len(row) else "")
            if not tag:
                continue
            if sentiment == "正向":
                positive_tags[tag] += 1
            elif sentiment == "负向":
                negative_tags[tag] += 1

    sort_tags = lambda counter: tuple(sorted(counter.items(), key=lambda item: (-item[1], item[0])))
    return PlatformSummary(
        key,
        label,
        sort_tags(positive_tags),
        sort_tags(negative_tags),
        positive_users,
        negative_users,
    )


def summarize_social_file(key: str, label: str, path: Path) -> PlatformSummary:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        return summarize_social_rows(key, label, worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _count_value(value: Any, *, label: str) -> int:
    if isinstance(value, bool):
        raise RuntimeError(f"{label}不是有效整数：{value}")
    if isinstance(value, (int, float)) and float(value).is_integer():
        return int(value)
    text = _text(value).replace(",", "")
    try:
        parsed = float(text)
    except ValueError as exc:
        raise RuntimeError(f"{label}不是有效整数：{value}") from exc
    if not parsed.is_integer():
        raise RuntimeError(f"{label}不是有效整数：{value}")
    return int(parsed)


def summarize_dianping_rows(rows: Iterable[Iterable[Any]]) -> PlatformSummary:
    values = [list(row) for row in rows]
    header_row = None
    indexes: tuple[int, int, int, int] | None = None
    for row_number, row in enumerate(values):
        normalized = [_text(value).replace(" ", "") for value in row]
        try:
            positive_index = normalized.index("大众点评好评")
            negative_index = normalized.index("大众点评差评")
        except ValueError:
            continue
        positive_count_index = next(
            (index for index in range(positive_index + 1, len(normalized)) if normalized[index] == "评论数"),
            None,
        )
        negative_count_index = next(
            (index for index in range(negative_index + 1, len(normalized)) if normalized[index] == "评论数"),
            None,
        )
        if positive_count_index is not None and negative_count_index is not None:
            header_row = row_number
            indexes = (
                positive_index,
                positive_count_index,
                negative_index,
                negative_count_index,
            )
            break
    if header_row is None or indexes is None:
        raise RuntimeError("大众点评工作表缺少统计表头。")

    positive_tags: Counter[str] = Counter()
    negative_tags: Counter[str] = Counter()
    positive_users: int | None = None
    negative_users: int | None = None
    positive_index, positive_count_index, negative_index, negative_count_index = indexes
    for row in values[header_row + 1 :]:
        padded = row + [None] * max(0, negative_count_index + 1 - len(row))
        positive_label = _text(padded[positive_index])
        negative_label = _text(padded[negative_index])
        if positive_label == "好评用户数" and negative_label == "差评用户数":
            positive_users = _count_value(
                padded[positive_count_index], label="大众点评好评用户数"
            )
            negative_users = _count_value(
                padded[negative_count_index], label="大众点评差评用户数"
            )
            break
        if positive_label and positive_label != "/":
            positive_tags[positive_label] += _count_value(
                padded[positive_count_index], label=f"大众点评好评标签“{positive_label}”评论数"
            )
        if negative_label and negative_label != "/":
            negative_tags[negative_label] += _count_value(
                padded[negative_count_index], label=f"大众点评差评标签“{negative_label}”评论数"
            )
    if positive_users is None or negative_users is None:
        raise RuntimeError("大众点评工作表缺少“好评用户数/差评用户数”汇总行。")

    sort_tags = lambda counter: tuple(sorted(counter.items(), key=lambda item: (-item[1], item[0])))
    return PlatformSummary(
        "dianping",
        "大众点评",
        sort_tags(positive_tags),
        sort_tags(negative_tags),
        positive_users,
        negative_users,
    )


def summarize_social_cleaned_workbook(path: Path) -> list[PlatformSummary]:
    """Read the prepared Dianping table and four raw-platform sheets from one workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        required_sheets = ["大众点评", *(label for _, label in SOCIAL_RAW_PLATFORMS)]
        missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
        if missing_sheets:
            raise RuntimeError(
                f"社媒清洗数据缺少工作表：{'、'.join(missing_sheets)}（文件：{path.name}）"
            )
        summaries = [
            summarize_dianping_rows(workbook["大众点评"].iter_rows(values_only=True))
        ]
        summaries.extend(
            summarize_social_rows(
                key,
                label,
                workbook[label].iter_rows(values_only=True),
            )
            for key, label in SOCIAL_RAW_PLATFORMS
        )
        return summaries
    finally:
        workbook.close()


def empty_platform_summary(key: str, label: str) -> PlatformSummary:
    return PlatformSummary(key, label, (), (), 0, 0)


def safe_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n]+', "_", value).strip().strip(".")
    return cleaned or "社媒评论统计"


def safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", value).strip().strip("'")
    return (cleaned or "社媒评论统计")[:31]


def _apply_table_style(worksheet, cell_range: str, *, fill: PatternFill | None = None, bold: bool = False) -> None:
    thin = Side(style="thin", color="404040")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="微软雅黑", size=11, bold=bold)
            if fill:
                cell.fill = fill


def build_social_feedback_workbook(
    *,
    brand: str,
    product: str,
    start_date: str,
    end_date: str,
    summaries: list[PlatformSummary],
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{safe_filename(brand)}-{safe_filename(product)}-社媒评论统计.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_sheet_name(product)
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="D9D7D7")
    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = f"{product} {start_date}-{end_date} 第三方平台评价反馈"
    worksheet["E1"] = "好评率"
    worksheet["E2"] = "总计"
    worksheet["E3"] = "好评用户数"
    worksheet["E4"] = "差评用户数"

    positive_users = sum(summary.positive_users for summary in summaries)
    negative_users = sum(summary.negative_users for summary in summaries)
    total_users = positive_users + negative_users
    worksheet["F1"] = positive_users / total_users if total_users else 0
    worksheet["F2"] = total_users
    worksheet["F3"] = positive_users
    worksheet["F4"] = negative_users
    worksheet["F1"].number_format = "0%"
    for row in range(2, 5):
        worksheet[f"F{row}"].number_format = "#,##0"

    _apply_table_style(worksheet, "A1:D1", fill=header_fill, bold=True)
    _apply_table_style(worksheet, "E1:F4")
    for cell in worksheet["A1:D1"][0]:
        cell.font = Font(name="微软雅黑", size=13, bold=True)
    for cell in worksheet["E"][:4]:
        cell.font = Font(name="微软雅黑", size=11, color="000000")

    current_row = 2
    for summary in summaries:
        header_row = current_row
        worksheet.cell(header_row, 1, f"{summary.label}好评")
        worksheet.cell(header_row, 2, "评论数")
        worksheet.cell(header_row, 3, f"{summary.label}差评")
        worksheet.cell(header_row, 4, "评论数")
        _apply_table_style(worksheet, f"A{header_row}:D{header_row}", fill=header_fill, bold=True)
        current_row += 1

        detail_count = max(len(summary.positive_tags), len(summary.negative_tags))
        for index in range(detail_count):
            positive = summary.positive_tags[index] if index < len(summary.positive_tags) else None
            negative = summary.negative_tags[index] if index < len(summary.negative_tags) else None
            worksheet.cell(current_row, 1, positive[0] if positive else "")
            worksheet.cell(current_row, 2, positive[1] if positive else "")
            worksheet.cell(current_row, 3, negative[0] if negative else "")
            worksheet.cell(current_row, 4, negative[1] if negative else "")
            current_row += 1
        if detail_count:
            _apply_table_style(worksheet, f"A{header_row + 1}:D{current_row - 1}")

        worksheet.cell(current_row, 1, "好评用户数")
        worksheet.cell(current_row, 2, summary.positive_users)
        worksheet.cell(current_row, 3, "差评用户数")
        worksheet.cell(current_row, 4, summary.negative_users)
        _apply_table_style(worksheet, f"A{current_row}:D{current_row}", bold=True)
        current_row += 1

    last_row = current_row - 1
    for column in ("B", "D"):
        for cell in worksheet[column][1:last_row]:
            if isinstance(cell.value, int):
                cell.number_format = "#,##0"
    widths = {"A": 27, "B": 11, "C": 27, "D": 11, "E": 15, "F": 11}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for row in range(1, last_row + 1):
        worksheet.row_dimensions[row].height = 26

    worksheet.freeze_panes = "A2"
    worksheet.print_area = f"A1:F{last_row}"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    workbook.save(output_path)
    return output_path
