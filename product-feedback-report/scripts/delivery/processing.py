from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DELIVERY_HEADERS = ["平台", "品牌", "店铺名称", "分类描述", "宝贝名称", "原价", "现价", "宝贝月销", "抓取日期"]
RECENT_LAUNCH_DAYS = 32
SALES_WINDOW_DAYS = 30
RECENT_LAUNCH_HEADER = "近32日上新日期"
LEGACY_RECENT_LAUNCH_HEADER = "近30日上新日期"
ANNOTATION_HEADERS = ["平台", "品牌", "产品", RECENT_LAUNCH_HEADER]


@dataclass(frozen=True)
class RawSaleRow:
    platform: str
    brand: str
    store: str
    category: str
    product: str
    monthly_sales: float
    crawl_date: date


@dataclass(frozen=True)
class AnnotationRow:
    platform: str
    brand: str
    product: str
    recent_launch_date: date | None


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            return None
    return None


def format_chinese_date(value: date | None) -> str:
    return value.strftime("%Y年%m月%d日") if value else ""


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_bool(value: Any) -> bool:
    text = normalize_text(value).lower()
    return text in {"是", "y", "yes", "true", "1", "新", "上新"}


def numeric(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[,，]", "", str(value))
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def find_header_map(row: Iterable[Any], expected: list[str]) -> dict[str, int]:
    cells = [normalize_text(cell) for cell in row]
    mapping: dict[str, int] = {}
    for name in expected:
        if name in cells:
            mapping[name] = cells.index(name)
    return mapping


def looks_like_delivery_sheet(ws) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        mapping = find_header_map(row, ["平台", "品牌", "宝贝名称", "宝贝月销"])
        if {"平台", "品牌", "宝贝名称", "宝贝月销"}.issubset(mapping):
            return True
    return False


def iter_delivery_rows(path: Path) -> list[RawSaleRow]:
    rows: list[RawSaleRow] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if not looks_like_delivery_sheet(ws):
            continue
        header_map: dict[str, int] | None = None
        for row in ws.iter_rows(values_only=True):
            if header_map is None:
                candidate = find_header_map(row, DELIVERY_HEADERS)
                if {"平台", "品牌", "宝贝名称", "宝贝月销", "抓取日期"}.issubset(candidate):
                    header_map = candidate
                continue
            crawl_date = parse_date(row[header_map["抓取日期"]])
            product = normalize_text(row[header_map["宝贝名称"]])
            brand = normalize_text(row[header_map["品牌"]])
            platform = normalize_text(row[header_map["平台"]])
            if not product or not brand or not platform or not crawl_date:
                continue
            rows.append(
                RawSaleRow(
                    platform=platform,
                    brand=brand,
                    store=normalize_text(row[header_map.get("店铺名称", -1)]) if "店铺名称" in header_map else "",
                    category=normalize_text(row[header_map.get("分类描述", -1)]) if "分类描述" in header_map else "",
                    product=product,
                    monthly_sales=numeric(row[header_map["宝贝月销"]]),
                    crawl_date=crawl_date,
                )
            )
    return rows


def find_workbooks(input_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in input_dir.rglob("*.xlsx")
        if not path.name.startswith(("~$", ".~")) and not path.name.startswith(".")
    )


def find_delivery_rows(input_dir: Path) -> list[RawSaleRow]:
    all_rows: list[RawSaleRow] = []
    for path in find_workbooks(input_dir):
        all_rows.extend(iter_delivery_rows(path))
    return all_rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        for row_idx in range(1, min(ws.max_row, 300) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                display_width = sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in str(value))
                max_len = max(max_len, min(display_width, 42))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    ws.freeze_panes = "A2"


def save_workbook(wb: Workbook, path: Path) -> Path:
    ensure_dir(path.parent)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(path)
    return path


def write_product_menu(rows: list[RawSaleRow], output_path: Path, recent_launch_dates: dict[tuple[str, str], date] | None = None) -> Path:
    recent_launch_dates = recent_launch_dates or {}
    unique = sorted({(row.platform, row.brand, row.product) for row in rows})
    wb = Workbook()
    ws = wb.active
    ws.title = "产品清单及上新日期"
    ws.append(ANNOTATION_HEADERS)
    for platform, brand, product in unique:
        ws.append([platform, brand, product, format_chinese_date(recent_launch_dates.get((brand, product)))])
    return save_workbook(wb, output_path)


def read_annotation(path: Path) -> list[AnnotationRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = find_header_map(header, ANNOTATION_HEADERS)
    if RECENT_LAUNCH_HEADER not in mapping and LEGACY_RECENT_LAUNCH_HEADER in [normalize_text(cell) for cell in header]:
        mapping[RECENT_LAUNCH_HEADER] = [normalize_text(cell) for cell in header].index(LEGACY_RECENT_LAUNCH_HEADER)
    required = {"平台", "品牌", "产品", RECENT_LAUNCH_HEADER}
    missing = required - set(mapping)
    if missing:
        raise RuntimeError(f"标注菜单缺少字段：{', '.join(sorted(missing))}")
    rows: list[AnnotationRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = normalize_text(row[mapping["产品"]])
        if not product:
            continue
        rows.append(
            AnnotationRow(
                platform=normalize_text(row[mapping["平台"]]),
                brand=normalize_text(row[mapping["品牌"]]),
                product=product,
                recent_launch_date=parse_date(row[mapping[RECENT_LAUNCH_HEADER]]),
            )
        )
    return rows


def _annotation_lookup(rows: list[AnnotationRow]) -> dict[tuple[str, str, str], AnnotationRow]:
    return {(_platform_group(row.platform), row.brand, row.product): row for row in rows}


def _platform_group(platform: str) -> str:
    if "饿了么" in platform:
        return "饿了么"
    if "美团" in platform:
        return "美团"
    if "京东" in platform:
        return "京东"
    return platform


def _latest_store_product_rows(raw_rows: list[RawSaleRow], platform_name: str) -> list[RawSaleRow]:
    latest: dict[tuple[str, str, str, str], RawSaleRow] = {}
    for row in raw_rows:
        platform = _platform_group(row.platform)
        if platform != platform_name:
            continue
        if not row.store:
            raise RuntimeError(
                f"{platform_name}外卖数据缺少门店名称：品牌={row.brand}，产品={row.product}，抓取日期={row.crawl_date:%Y-%m-%d}"
            )
        key = (platform, row.brand, row.store, row.product)
        previous = latest.get(key)
        if previous is None or row.crawl_date >= previous.crawl_date:
            latest[key] = row
    return list(latest.values())


def _store_sale_days(row: RawSaleRow, lookup: dict[tuple[str, str, str], AnnotationRow]) -> int:
    platform = _platform_group(row.platform)
    annotation = lookup.get((platform, row.brand, row.product))
    launch_date = annotation.recent_launch_date if annotation else None
    if launch_date is None:
        return SALES_WINDOW_DAYS
    difference = (row.crawl_date - launch_date).days
    if difference < 0:
        raise RuntimeError(
            "上架日期晚于抓取日期："
            f"平台={platform}，品牌={row.brand}，产品={row.product}，"
            f"上架日期={launch_date:%Y-%m-%d}，抓取日期={row.crawl_date:%Y-%m-%d}"
        )
    return min(difference + 1, SALES_WINDOW_DAYS)


def _platform_delivery_metrics(
    raw_rows: list[RawSaleRow], annotations: list[AnnotationRow], platform_name: str
) -> list[dict[str, Any]]:
    lookup = _annotation_lookup(annotations)
    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {"sales": 0.0, "sale_days": 0, "stores": 0}
    )
    for row in _latest_store_product_rows(raw_rows, platform_name):
        data = grouped[(row.brand, row.product)]
        data["sales"] += row.monthly_sales
        data["sale_days"] += _store_sale_days(row, lookup)
        data["stores"] += 1

    metrics: list[dict[str, Any]] = []
    for (brand, product), data in grouped.items():
        metrics.append(
            {
                "brand": brand,
                "product": product,
                "sales": data["sales"],
                "sale_days": data["sale_days"],
                "stores": data["stores"],
                "launch_date": (
                    lookup.get((platform_name, brand, product)).recent_launch_date
                    if lookup.get((platform_name, brand, product))
                    else None
                ),
                "daily_store_avg": data["sales"] / data["sale_days"],
            }
        )
    return metrics


def generate_delivery_summary(raw_rows: list[RawSaleRow], annotations: list[AnnotationRow], output_path: Path) -> Path:
    product_rows: dict[str, dict[str, Any]] = defaultdict(dict)
    for platform in ("美团", "饿了么"):
        for metric in _platform_delivery_metrics(raw_rows, annotations, platform):
            product_rows[metric["product"]][platform] = metric
    total_avg = sum(sum(p.get(platform, {}).get("daily_store_avg", 0.0) for platform in ("美团", "饿了么")) for p in product_rows.values())
    ranked = []
    for product, data in product_rows.items():
        meituan_avg = data.get("美团", {}).get("daily_store_avg", 0.0)
        eleme_avg = data.get("饿了么", {}).get("daily_store_avg", 0.0)
        total = meituan_avg + eleme_avg
        ranked.append((product, meituan_avg, eleme_avg, total, total / total_avg if total_avg else 0.0))
    ranked.sort(key=lambda item: item[3], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "美团&饿了么外卖数据"
    ws.append(["排名", "商品名称", "美团日店均", "饿了么日店均", "线上合计日店均", "销量占比"])
    for idx, (product, meituan_avg, eleme_avg, total, share) in enumerate(ranked, 1):
        ws.append([idx, product, round(meituan_avg, 1), round(eleme_avg, 1), round(total, 1), share])
    for cell in ws["F"][1:]:
        cell.number_format = "0.0%"
    return save_workbook(wb, output_path)


def generate_platform_delivery_summary(
    raw_rows: list[RawSaleRow],
    annotations: list[AnnotationRow],
    platform_name: str,
    output_path: Path,
) -> Path:
    metrics = _platform_delivery_metrics(raw_rows, annotations, platform_name)
    total_sales = sum(item["sales"] for item in metrics)
    metrics.sort(key=lambda item: (-item["sales"], item["product"], item["brand"]))

    wb = Workbook()
    ws = wb.active
    ws.title = f"{platform_name}外卖数据"
    ws.append(
        [
            "排名",
            "商品名称",
            "日店均销量",
            "总销量",
            "总销量占比",
            "总在售天数",
            "在售门店数",
            "上新日期",
        ]
    )
    previous_sales: float | None = None
    rank = 0
    for idx, item in enumerate(metrics, 1):
        if previous_sales is None or item["sales"] != previous_sales:
            rank = idx
        previous_sales = item["sales"]
        ws.append(
            [
                rank,
                item["product"],
                round(item["daily_store_avg"], 4),
                item["sales"],
                item["sales"] / total_sales if total_sales else 0.0,
                item["sale_days"],
                item["stores"],
                item["launch_date"],
            ]
        )
    for cell in ws["C"][1:]:
        cell.number_format = "0.0000"
    for cell in ws["D"][1:]:
        cell.number_format = "#,##0"
    for cell in ws["E"][1:]:
        cell.number_format = "0.00%"
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0"
    for cell in ws["G"][1:]:
        cell.number_format = "#,##0"
    for cell in ws["H"][1:]:
        cell.number_format = "yyyy-mm-dd"
    return save_workbook(wb, output_path)


def generate_jd_summary(raw_rows: list[RawSaleRow], annotations: list[AnnotationRow], output_path: Path) -> Path:
    lookup = _annotation_lookup(annotations)
    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: {"sales_sum": 0.0, "stores": 0})
    for row in _latest_store_product_rows(raw_rows, "京东"):
        data = grouped[(row.brand, row.product)]
        data["sales_sum"] += row.monthly_sales
        data["stores"] += 1
    metrics = [
        {
            "brand": brand,
            "product": product,
            "sales_sum": data["sales_sum"],
            "stores": data["stores"],
            "sales": data["sales_sum"] / data["stores"],
            "launch_date": (
                lookup.get(("京东", brand, product)).recent_launch_date
                if lookup.get(("京东", brand, product))
                else None
            ),
        }
        for (brand, product), data in grouped.items()
    ]
    total_sales = sum(item["sales"] for item in metrics)
    metrics.sort(key=lambda item: (-item["sales"], item["product"], item["brand"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "京东外卖数据"
    ws.append(["排名", "商品名称", "销量加总", "在售门店数", "总销量", "总销量占比", "上新日期"])
    previous_sales: float | None = None
    rank = 0
    for idx, item in enumerate(metrics, 1):
        if previous_sales is None or item["sales"] != previous_sales:
            rank = idx
        previous_sales = item["sales"]
        ws.append(
            [
                rank,
                item["product"],
                item["sales_sum"],
                item["stores"],
                round(item["sales"], 1),
                item["sales"] / total_sales if total_sales else 0.0,
                item["launch_date"],
            ]
        )
    for cell in ws["C"][1:]:
        cell.number_format = "#,##0"
    for cell in ws["D"][1:]:
        cell.number_format = "#,##0"
    for cell in ws["E"][1:]:
        cell.number_format = "0.0"
    for cell in ws["F"][1:]:
        cell.number_format = "0.00%"
    for cell in ws["G"][1:]:
        cell.number_format = "yyyy-mm-dd"
    return save_workbook(wb, output_path)

