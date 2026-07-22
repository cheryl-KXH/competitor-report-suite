"""参考原 PDF 结构生成自包含 HTML 跟踪报告。"""

from __future__ import annotations

import base64
import html
import mimetypes
import re
import unicodedata
import urllib.request
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SUITE_ROOT = ROOT.parent.resolve()
GENERIC_POSITIVE_TAGS = {"好喝，喜欢，推荐"}
GENERIC_NEGATIVE_TAGS = {"难喝，不喜欢，不推荐"}


@dataclass(frozen=True)
class ProductInfo:
    product_name: str
    series: str = ""
    selling_point: str = ""
    price: str = ""
    ingredients: str = ""
    image_urls: tuple[str, ...] = ()


@dataclass(frozen=True)
class PlatformSale:
    product: str
    daily_store_avg: float
    total_sales: float
    stores: int


@dataclass(frozen=True)
class CombinedSale:
    rank: int
    key: str
    product: str
    meituan_daily: float
    eleme_daily: float
    combined_daily: float
    total_sales: float
    share: float
    tracked: bool = False


@dataclass(frozen=True)
class DeliveryReport:
    all_rows: tuple[CombinedSale, ...]
    display_rows: tuple[CombinedSale, ...]
    tracked_rows: dict[str, CombinedSale]
    meituan_total_daily: float
    eleme_total_daily: float
    meituan_stores: int
    eleme_stores: int


@dataclass(frozen=True)
class JdSale:
    rank: int
    product: str
    total_sales: float
    share: float


@dataclass(frozen=True)
class SocialSection:
    label: str
    positive_tags: tuple[tuple[str, int], ...]
    negative_tags: tuple[tuple[str, int], ...]
    positive_users: int
    negative_users: int

    @property
    def positive_count(self) -> int:
        return sum(count for _, count in self.positive_tags)

    @property
    def negative_count(self) -> int:
        return sum(count for _, count in self.negative_tags)


@dataclass(frozen=True)
class SocialReport:
    title: str
    period: str
    sections: tuple[SocialSection, ...]
    positive_users: int
    negative_users: int
    positive_top: tuple[tuple[str, int], ...]
    negative_top: tuple[tuple[str, int], ...]

    @property
    def total_users(self) -> int:
        return self.positive_users + self.negative_users

    @property
    def positive_count(self) -> int:
        return sum(section.positive_count for section in self.sections)

    @property
    def negative_count(self) -> int:
        return sum(section.negative_count for section in self.sections)

    @property
    def total_count(self) -> int:
        return self.positive_count + self.negative_count

    @property
    def positive_rate(self) -> float:
        return self.positive_users / self.total_users if self.total_users else 0.0


@dataclass(frozen=True)
class ReportBuildResult:
    path: Path
    warnings: tuple[str, ...]
    pdf_path: Path | None = None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else 0.0


def _integer(value: Any) -> int:
    return int(round(_number(value)))


def normalize_product_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).casefold()
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", "", text)


def _header_map(values: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    return {_text(value).replace(" ", ""): index for index, value in enumerate(values)}


def read_platform_sales(path: Path, platform: str) -> tuple[dict[str, PlatformSale], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise RuntimeError(f"{platform}外卖数表为空：{path.name}") from exc
    columns = _header_map(header)
    required = {"商品名称", "日店均销量", "总销量", "在售门店数"}
    missing = required - set(columns)
    if missing:
        raise RuntimeError(f"{platform}外卖数表缺少字段：{'、'.join(sorted(missing))}")
    sales: dict[str, PlatformSale] = {}
    max_stores = 0
    for row in rows:
        product = _text(row[columns["商品名称"]])
        if not product:
            continue
        key = normalize_product_key(product)
        if key in sales:
            raise RuntimeError(f"{platform}外卖数表存在规范化后重名商品：{sales[key].product}、{product}")
        stores = _integer(row[columns["在售门店数"]])
        sales[key] = PlatformSale(
            product=product,
            daily_store_avg=_number(row[columns["日店均销量"]]),
            total_sales=_number(row[columns["总销量"]]),
            stores=stores,
        )
        max_stores = max(max_stores, stores)
    if not sales:
        raise RuntimeError(f"{platform}外卖数表没有有效商品数据：{path.name}")
    return sales, max_stores


def _tracked_matches(products: list[str], names: dict[str, str]) -> dict[str, str]:
    matched: dict[str, str] = {}
    for product in products:
        target = normalize_product_key(product)
        exact = [key for key in names if key == target]
        candidates = exact or [key for key in names if target and target in key]
        if len(candidates) > 1:
            labels = "、".join(names[key] for key in candidates)
            raise RuntimeError(f"关注新品“{product}”匹配到多个外卖商品：{labels}")
        if candidates:
            matched[product] = candidates[0]
    return matched


def combine_delivery_sales(
    meituan_path: Path, eleme_path: Path, tracked_products: list[str]
) -> DeliveryReport:
    meituan, meituan_stores = read_platform_sales(meituan_path, "美团")
    eleme, eleme_stores = read_platform_sales(eleme_path, "饿了么")
    return combine_delivery_models(
        meituan,
        eleme,
        tracked_products,
        meituan_stores=meituan_stores,
        eleme_stores=eleme_stores,
    )


def delivery_report_from_metrics(
    meituan_metrics: list[dict[str, Any]],
    eleme_metrics: list[dict[str, Any]],
    tracked_products: list[str],
) -> DeliveryReport:
    def convert(metrics: list[dict[str, Any]]) -> tuple[dict[str, PlatformSale], int]:
        sales: dict[str, PlatformSale] = {}
        max_stores = 0
        for item in metrics:
            product = _text(item.get("product"))
            if not product:
                continue
            key = normalize_product_key(product)
            stores = _integer(item.get("stores"))
            sales[key] = PlatformSale(
                product=product,
                daily_store_avg=_number(item.get("daily_store_avg")),
                total_sales=_number(item.get("sales")),
                stores=stores,
            )
            max_stores = max(max_stores, stores)
        return sales, max_stores

    meituan, meituan_stores = convert(meituan_metrics)
    eleme, eleme_stores = convert(eleme_metrics)
    if not meituan:
        raise RuntimeError("美团外卖数据没有有效商品。")
    if not eleme:
        raise RuntimeError("饿了么外卖数据没有有效商品。")
    return combine_delivery_models(
        meituan,
        eleme,
        tracked_products,
        meituan_stores=meituan_stores,
        eleme_stores=eleme_stores,
    )


def combine_delivery_models(
    meituan: dict[str, PlatformSale],
    eleme: dict[str, PlatformSale],
    tracked_products: list[str],
    *,
    meituan_stores: int,
    eleme_stores: int,
) -> DeliveryReport:
    keys = set(meituan) | set(eleme)
    names = {
        key: (meituan.get(key) or eleme[key]).product
        for key in keys
    }
    tracked_matches = _tracked_matches(tracked_products, names)
    tracked_keys = set(tracked_matches.values())
    total_sales = sum(
        (meituan.get(key).total_sales if key in meituan else 0.0)
        + (eleme.get(key).total_sales if key in eleme else 0.0)
        for key in keys
    )
    sortable: list[tuple[str, str, float, float, float, float]] = []
    for key in keys:
        mt = meituan.get(key)
        elm = eleme.get(key)
        mt_daily = mt.daily_store_avg if mt else 0.0
        elm_daily = elm.daily_store_avg if elm else 0.0
        combined_sales = (mt.total_sales if mt else 0.0) + (elm.total_sales if elm else 0.0)
        sortable.append((key, names[key], mt_daily, elm_daily, mt_daily + elm_daily, combined_sales))
    sortable.sort(key=lambda item: (-item[5], item[1]))
    ranked: list[CombinedSale] = []
    previous_sales: float | None = None
    previous_rank = 0
    for index, (key, product, mt_daily, elm_daily, combined_daily, combined_sales) in enumerate(sortable, 1):
        rank = previous_rank if previous_sales is not None and combined_sales == previous_sales else index
        previous_sales = combined_sales
        previous_rank = rank
        ranked.append(
            CombinedSale(
                rank=rank,
                key=key,
                product=product,
                meituan_daily=mt_daily,
                eleme_daily=elm_daily,
                combined_daily=combined_daily,
                total_sales=combined_sales,
                share=combined_sales / total_sales if total_sales else 0.0,
                tracked=key in tracked_keys,
            )
        )
    display_rows = tuple(row for row in ranked if row.rank <= 20 or row.tracked)
    by_key = {row.key: row for row in ranked}
    tracked_rows = {
        product: by_key[key]
        for product, key in tracked_matches.items()
    }
    return DeliveryReport(
        all_rows=tuple(ranked),
        display_rows=display_rows,
        tracked_rows=tracked_rows,
        meituan_total_daily=sum(item.daily_store_avg for item in meituan.values()),
        eleme_total_daily=sum(item.daily_store_avg for item in eleme.values()),
        meituan_stores=meituan_stores,
        eleme_stores=eleme_stores,
    )


def jd_report_from_metrics(
    metrics: list[dict[str, Any]],
) -> tuple[tuple[JdSale, ...], float]:
    total = sum(_number(item.get("sales")) for item in metrics)
    rows: list[JdSale] = []
    previous_sales: float | None = None
    previous_rank = 0
    for index, item in enumerate(
        sorted(metrics, key=lambda value: (-_number(value.get("sales")), _text(value.get("product")))),
        1,
    ):
        sales = _number(item.get("sales"))
        rank = previous_rank if previous_sales is not None and sales == previous_sales else index
        previous_sales = sales
        previous_rank = rank
        rows.append(
            JdSale(
                rank=rank,
                product=_text(item.get("product")),
                total_sales=sales,
                share=sales / total if total else 0.0,
            )
        )
    return tuple(rows), total


def read_jd_sales(path: Path | None) -> tuple[tuple[JdSale, ...], float]:
    if not path or not path.exists():
        return (), 0.0
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return (), 0.0
    columns = _header_map(header)
    required = {"排名", "商品名称", "总销量", "总销量占比"}
    if required - set(columns):
        raise RuntimeError(f"京东外卖数表缺少字段：{'、'.join(sorted(required - set(columns)))}")
    all_rows: list[JdSale] = []
    for row in rows:
        product = _text(row[columns["商品名称"]])
        if not product:
            continue
        all_rows.append(
            JdSale(
                rank=_integer(row[columns["排名"]]),
                product=product,
                total_sales=_number(row[columns["总销量"]]),
                share=_number(row[columns["总销量占比"]]),
            )
        )
    return tuple(all_rows), sum(row.total_sales for row in all_rows)


def _counter_top(counter: Counter[str], generic: set[str], limit: int = 3) -> tuple[tuple[str, int], ...]:
    filtered = [
        (label, count)
        for label, count in counter.items()
        if normalize_product_key(label) not in generic and label and count > 0
    ]
    filtered.sort(key=lambda item: (-item[1], item[0]))
    return tuple(filtered[:limit])


def read_social_report(path: Path) -> SocialReport:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        raise RuntimeError(f"社媒评论统计表为空：{path.name}")
    title = _text(values[0][0])
    period_match = re.search(r"(\d{1,2}\.\d{1,2}\s*-\s*\d{1,2}\.\d{1,2})", title)
    period = period_match.group(1).replace(" ", "") if period_match else ""
    sections: list[SocialSection] = []
    positive_counter: Counter[str] = Counter()
    negative_counter: Counter[str] = Counter()
    row_index = 1
    while row_index < len(values):
        row = list(values[row_index]) + [None] * 6
        positive_header = _text(row[0])
        negative_header = _text(row[2])
        if _text(row[1]) != "评论数" or _text(row[3]) != "评论数" or not positive_header.endswith("好评"):
            row_index += 1
            continue
        label = positive_header[:-2]
        positive_tags: list[tuple[str, int]] = []
        negative_tags: list[tuple[str, int]] = []
        positive_users = 0
        negative_users = 0
        row_index += 1
        while row_index < len(values):
            detail = list(values[row_index]) + [None] * 6
            if _text(detail[0]) == "好评用户数" and _text(detail[2]) == "差评用户数":
                positive_users = _integer(detail[1])
                negative_users = _integer(detail[3])
                row_index += 1
                break
            positive_label = _text(detail[0])
            negative_label = _text(detail[2])
            positive_count = _integer(detail[1])
            negative_count = _integer(detail[3])
            if positive_label and positive_label != "/" and positive_count > 0:
                positive_tags.append((positive_label, positive_count))
                positive_counter[positive_label] += positive_count
            if negative_label and negative_label != "/" and negative_count > 0:
                negative_tags.append((negative_label, negative_count))
                negative_counter[negative_label] += negative_count
            row_index += 1
        sections.append(
            SocialSection(
                label=label or negative_header[:-2],
                positive_tags=tuple(positive_tags),
                negative_tags=tuple(negative_tags),
                positive_users=positive_users,
                negative_users=negative_users,
            )
        )
    generic_positive = {normalize_product_key(value) for value in GENERIC_POSITIVE_TAGS}
    generic_negative = {normalize_product_key(value) for value in GENERIC_NEGATIVE_TAGS}
    return SocialReport(
        title=title,
        period=period,
        sections=tuple(sections),
        positive_users=sum(section.positive_users for section in sections),
        negative_users=sum(section.negative_users for section in sections),
        positive_top=_counter_top(positive_counter, generic_positive),
        negative_top=_counter_top(negative_counter, generic_negative),
    )


def social_report_from_summaries(
    *, title: str, period: str, summaries: list[Any]
) -> SocialReport:
    sections = tuple(
        SocialSection(
            label=_text(summary.label),
            positive_tags=tuple(summary.positive_tags),
            negative_tags=tuple(summary.negative_tags),
            positive_users=int(summary.positive_users),
            negative_users=int(summary.negative_users),
        )
        for summary in summaries
    )
    positive_counter: Counter[str] = Counter()
    negative_counter: Counter[str] = Counter()
    for section in sections:
        positive_counter.update(dict(section.positive_tags))
        negative_counter.update(dict(section.negative_tags))
    return SocialReport(
        title=title,
        period=period,
        sections=sections,
        positive_users=sum(section.positive_users for section in sections),
        negative_users=sum(section.negative_users for section in sections),
        positive_top=_counter_top(
            positive_counter,
            {normalize_product_key(value) for value in GENERIC_POSITIVE_TAGS},
        ),
        negative_top=_counter_top(
            negative_counter,
            {normalize_product_key(value) for value in GENERIC_NEGATIVE_TAGS},
        ),
    )


def _safe_config_path(value: str) -> Path | None:
    if not value:
        return None
    candidate = (ROOT / value).resolve()
    try:
        candidate.relative_to(SUITE_ROOT)
    except ValueError:
        return None
    return candidate if candidate.exists() and candidate.is_file() else None


def _file_data_uri(path: Path) -> str:
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return f"data:{mime};base64,{base64.b64encode(path.read_bytes()).decode('ascii')}"


def _remote_image_data_uri(url: str) -> str:
    if url.startswith("data:image/"):
        return url
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        content = response.read(15 * 1024 * 1024 + 1)
        if len(content) > 15 * 1024 * 1024:
            raise RuntimeError("产品图片超过 15MB")
        content_type = response.headers.get("Content-Type", "image/png").split(";")[0].strip()
        if not content_type.startswith("image/"):
            raise RuntimeError(f"产品图片类型无效：{content_type}")
    return f"data:{content_type};base64,{base64.b64encode(content).decode('ascii')}"


def _font_css(configs: dict[str, dict[str, Any]], warnings: list[str]) -> tuple[str, str]:
    layout = configs.get("html_layout", {})
    font_config = configs.get("font_files", {})
    fonts = layout.get("fonts", {})
    default_family = str(
        font_config.get("defaultFonts", {}).get("cssFamily")
        or fonts.get("fallbackFamily")
        or "Microsoft YaHei, Arial, sans-serif"
    )
    font_dir = str(font_config.get("fontDirectory") or "")
    css: list[str] = []
    for key, unicode_range in (
        (fonts.get("chineseFontFileKey", "chineseFont"), "U+2E80-2EFF, U+3000-303F, U+3400-4DBF, U+4E00-9FFF, U+F900-FAFF, U+FF00-FFEF"),
        (fonts.get("latinFontFileKey", "latinFont"), "U+0000-00FF, U+2000-206F, U+20A0-20CF"),
    ):
        filename = _text(font_config.get(key, {}).get("fileName"))
        path = _safe_config_path(str(Path(font_dir) / filename)) if filename else None
        if not path:
            warnings.append(f"HTML 字体文件不可访问：{filename or key}，已使用系统字体")
            continue
        font_format = "opentype" if path.suffix.lower() == ".otf" else "truetype"
        css.append(
            "@font-face{font-family:'ReportFont';"
            f"src:url('{_file_data_uri(path)}') format('{font_format}');"
            f"font-display:block;unicode-range:{unicode_range};}}"
        )
    family = f"'ReportFont', {default_family}" if css else default_family
    return "\n".join(css), family


def _escape(value: Any) -> str:
    return html.escape(_text(value), quote=True)


def _multiline(value: Any) -> str:
    text = _escape(value)
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    return text.replace("\n", "<br>")


def _format_count(value: float) -> str:
    return f"{value:,.0f}"


def _format_daily(value: float) -> str:
    return f"{value:.1f}"


def _format_percent(value: float) -> str:
    return f"{value:.1%}"


def _format_positive_rate(value: float) -> str:
    return f"{value:.0%}"


def _top_labels(values: tuple[tuple[str, int], ...]) -> str:
    return "、".join(label for label, _ in values) or "暂无明确高频观点"


def _price_html(value: str) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    text = "".join(line.strip() for line in text.split("\n") if line.strip())
    text = re.sub(r"\s*/\s*", "/", text).strip()
    escaped = _escape(text)
    pattern = re.compile(r"([(（])\s*(\d+(?:\.\d+)?)\s*元\s*([)）])")
    return pattern.sub(
        lambda match: (
            f'{match.group(1)}<span class="price-strike">'
            f'{match.group(2)}元</span>{match.group(3)}'
        ),
        escaped,
    )


def _delivery_html(brand: str, delivery: DeliveryReport) -> str:
    rows = []
    for row in delivery.display_rows:
        class_name = ' class="tracked-row"' if row.tracked else ""
        rows.append(
            f"<tr{class_name}><td>{_escape(row.product)}</td>"
            f"<td>{_format_daily(row.meituan_daily)}</td><td>{_format_daily(row.eleme_daily)}</td>"
            f"<td>{_format_daily(row.combined_daily)}</td><td>{_format_percent(row.share)}</td>"
            f'<td class="rank-cell">{row.rank}</td></tr>'
        )
    total_daily = delivery.meituan_total_daily + delivery.eleme_total_daily
    rows.append(
        "<tr class=\"total-row\"><td>总计</td>"
        f"<td>{_format_daily(delivery.meituan_total_daily)}</td>"
        f"<td>{_format_daily(delivery.eleme_total_daily)}</td>"
        f'<td>{_format_daily(total_daily)}</td><td>100.0%</td><td class="rank-cell">-</td></tr>'
    )
    return f"""
<div class="table-scroll">
<table class="sales-table">
<colgroup><col class="product-column"><col class="platform-column"><col class="platform-column">
<col class="combined-column"><col class="share-column"><col class="rank-column"></colgroup>
<thead><tr class="table-title-row"><th colspan="6">{_escape(brand)} 美团&amp;饿了么外卖销量表现</th></tr>
<tr><th rowspan="2">商品名称</th><th>美团</th><th>饿了么</th><th colspan="3">线上合计</th></tr>
<tr><th>日店均销量</th><th>日店均销量</th><th>日店均销量</th><th>销量占比</th><th class="rank-header">排名</th></tr></thead>
<tbody>{''.join(rows)}</tbody>
</table></div>
<p class="source-note">数据来源：根据{_escape(brand)}美团外卖 {_format_count(delivery.meituan_stores)} 家店铺和饿了么外卖 {_format_count(delivery.eleme_stores)} 家店铺月销数据计算</p>
"""


def _jd_html(
    brand: str,
    rows: tuple[JdSale, ...],
    total_sales: float,
    tracked_products: list[str],
) -> str:
    if not rows:
        return f'<p class="jd-missing-note">{_escape(brand)}京东外卖销售数据暂无法获取。</p>'
    names = {normalize_product_key(row.product): row.product for row in rows}
    tracked_keys = set(_tracked_matches(tracked_products, names).values())
    body_rows: list[str] = []
    for row in rows:
        product_key = normalize_product_key(row.product)
        if row.rank > 20 and product_key not in tracked_keys:
            continue
        class_name = ' class="tracked-row"' if product_key in tracked_keys else ""
        body_rows.append(
            f'<tr{class_name}><td class="rank-cell">{row.rank}</td><td>{_escape(row.product)}</td>'
            f"<td>{_format_count(row.total_sales)}</td><td>{_format_percent(row.share)}</td></tr>"
        )
    body = "".join(body_rows)
    body += f'<tr class="total-row"><td colspan="2">店铺合计</td><td>{_format_count(total_sales)}</td><td>100.0%</td></tr>'
    return f"""
<div class="table-scroll"><table class="jd-table">
<colgroup><col class="rank-column"><col class="product-column"><col class="sales-column"><col class="share-column"></colgroup>
<thead><tr class="table-title-row"><th colspan="4">{_escape(brand)} 京东外卖销量表现</th></tr>
<tr><th class="rank-header">排名</th><th>商品名称</th><th>京东外卖销量</th><th>销量占比</th></tr></thead>
<tbody>{body}</tbody></table></div>
<p class="jd-note">注：<strong>京东外卖销量</strong>显示为“<strong>品牌全国门店总量</strong>”，而<strong>美团/饿了么</strong>的销量为“<strong>单店独立销量</strong>”。因统计口径差异，三者不可直接对比。</p>
"""


def _product_info_html(product: str, info: ProductInfo | None, image_uri: str) -> str:
    if not info:
        return '<p class="missing-note">周报产品信息暂无法获取。</p>'
    image_html = (
        f'<img class="product-image" src="{image_uri}" alt="{_escape(product)} 产品外观">'
        if image_uri
        else '<span class="missing-inline">产品外观暂无法获取</span>'
    )
    values = (
        ("新品名称", info.product_name or product, "name-row"),
        ("产品系列归属", info.series, "series-row"),
        ("产品卖点介绍", info.selling_point, "long"),
        ("产品价格", info.price, "compact"),
        ("原料构成", info.ingredients, "ingredients-row"),
    )
    rows = "".join(
        f'<tr class="{kind}"><th>{label}</th><td>'
        f'{(_price_html(value) if label == "产品价格" else _multiline(value)) or "-"}</td></tr>'
        for label, value, kind in values
    )
    return f'<table class="product-info"><tbody>{rows}<tr class="image-row"><th>产品外观</th><td>{image_html}</td></tr></tbody></table>'


def _social_summary_html(social: SocialReport | None) -> str:
    if not social:
        return '<p class="missing-note">消费者反馈统计暂无法获取。</p>'
    period = f"（{_escape(social.period)}）" if social.period else ""
    if not social.total_users:
        return f'<p>上市30日{period}暂未获取到有效第三方评论。</p>'
    return (
        f'<p class="social-summary">上市30日{period}第三方评论共 {_format_count(social.total_users)} 条，'
        f'好评率为 {_format_positive_rate(social.positive_rate)}：<br>'
        f'好评（{_format_count(social.positive_users)} 条）主要提及关键词：{_escape(_top_labels(social.positive_top))}；<br>'
        f'差评（{_format_count(social.negative_users)} 条）主要提及关键词：{_escape(_top_labels(social.negative_top))}。</p>'
    )


def _social_detail_html(social: SocialReport | None) -> str:
    if not social:
        return '<p class="missing-note">消费者反馈详情暂无法获取。</p>'

    total_value = _format_count(social.total_users) if social.total_users else ""
    positive_user_value = _format_count(social.positive_users) if social.positive_users else ""
    negative_user_value = _format_count(social.negative_users) if social.negative_users else ""
    rate_value = _format_positive_rate(social.positive_rate) if social.total_users else ""
    detail_rows: list[str] = []
    colgroup = (
        '<colgroup><col class="positive-text"><col class="positive-count"><col class="negative-text">'
        '<col class="negative-count"><col class="kpi-label-column"><col class="kpi-value-column"></colgroup>'
    )

    for section_index, section in enumerate(social.sections):
        section_rows: list[str] = []
        if section_index == 0:
            section_rows.append(
                f'<tr class="feedback-title-row"><th class="feedback-title-cell" colspan="4">'
                f'{_escape(social.title)}</th><th class="kpi-label">好评率</th>'
                f'<td class="kpi-value">{rate_value}</td></tr>'
            )
        kpi_cells = (
            f'<th class="kpi-label">总计</th><td class="kpi-value">{total_value}</td>'
            if section_index == 0
            else '<td class="kpi-spacer" colspan="2"></td>'
        )
        section_rows.append(
            f'<tr class="platform-header"><th class="detail-header">{_escape(section.label)}好评</th>'
            f'<th class="detail-header count-header">评论数</th>'
            f'<th class="detail-header">{_escape(section.label)}差评</th>'
            f'<th class="detail-header count-header">评论数</th>{kpi_cells}</tr>'
        )
        detail_count = max(len(section.positive_tags), len(section.negative_tags))
        if section_index == 0:
            detail_count = max(detail_count, 2)

        for index in range(detail_count):
            positive = section.positive_tags[index] if index < len(section.positive_tags) else ("", 0)
            negative = section.negative_tags[index] if index < len(section.negative_tags) else ("", 0)
            if not section.positive_tags and not section.negative_tags:
                detail_cells = '<td colspan="4" class="empty-platform"></td>'
            else:
                detail_cells = (
                    f'<td>{_escape(positive[0])}</td><td>{positive[1] if positive[0] else ""}</td>'
                    f'<td>{_escape(negative[0])}</td><td>{negative[1] if negative[0] else ""}</td>'
                )

            if section_index == 0 and index == 0:
                detail_kpi_cells = (
                    f'<th class="kpi-label">好评用户数</th><td class="kpi-value">{positive_user_value}</td>'
                )
            elif section_index == 0 and index == 1:
                detail_kpi_cells = (
                    f'<th class="kpi-label">差评用户数</th><td class="kpi-value">{negative_user_value}</td>'
                )
            else:
                detail_kpi_cells = '<td class="kpi-spacer" colspan="2"></td>'
            section_rows.append(
                f'<tr>{detail_cells}{detail_kpi_cells}</tr>'
            )
        section_rows.append(
            f'<tr class="platform-total"><th>好评用户数</th><td>{section.positive_users}</td>'
            f'<th>差评用户数</th><td>{section.negative_users}</td>'
            '<td class="kpi-spacer" colspan="2"></td></tr>'
        )
        detail_rows.extend(section_rows)

    if not social.sections:
        section_rows = [
            f'<tr class="feedback-title-row"><th class="feedback-title-cell" colspan="4">'
            f'{_escape(social.title)}</th><th class="kpi-label">好评率</th>'
            f'<td class="kpi-value">{rate_value}</td></tr>',
                f'<tr><td colspan="4" class="empty-platform"></td>'
                f'<th class="kpi-label">总计</th><td class="kpi-value">{total_value}</td></tr>',
                f'<tr><td colspan="4" class="empty-platform"></td>'
                f'<th class="kpi-label">好评用户数</th><td class="kpi-value">{positive_user_value}</td></tr>',
                f'<tr><td colspan="4" class="empty-platform"></td>'
                f'<th class="kpi-label">差评用户数</th><td class="kpi-value">{negative_user_value}</td></tr>',
        ]
        detail_rows.extend(section_rows)

    return (
        '<div class="table-scroll feedback-detail-table-wrap">'
        f'<table class="feedback-table feedback-detail-table">{colgroup}'
        f'<tbody>{"".join(detail_rows)}</tbody></table></div>'
    )


def _mobile_zoom_script(page_width: int) -> str:
    return f"""
<script>
(function() {{
  var reportWidth = {page_width};
  var viewport = null;
  var scaleWrap = null;
  var report = null;
  var baseScale = 1;
  var userScale = 1;
  var pinchStartDistance = 0;
  var pinchStartScale = 1;
  var currentScale = 1;
  var panX = 0;
  var dragStartX = 0;
  var dragStartY = 0;
  var dragStartPanX = 0;
  var dragDirection = '';

  function isMobilePreview() {{
    return window.matchMedia('(max-width: 767px)').matches;
  }}

  function clamp(value, min, max) {{
    return Math.min(max, Math.max(min, value));
  }}

  function touchDistance(touches) {{
    var dx = touches[0].clientX - touches[1].clientX;
    var dy = touches[0].clientY - touches[1].clientY;
    return Math.sqrt(dx * dx + dy * dy);
  }}

  function applyScale() {{
    if (!viewport || !scaleWrap || !report) return;
    if (!isMobilePreview()) {{
      baseScale = 1;
      userScale = 1;
      currentScale = 1;
      panX = 0;
      viewport.style.width = reportWidth + 'px';
      viewport.style.height = '';
      scaleWrap.style.transform = '';
      return;
    }}
    var viewportWidth = Math.max(1, window.innerWidth || document.documentElement.clientWidth || reportWidth);
    baseScale = Math.min(1, viewportWidth / reportWidth);
    var scale = baseScale * userScale;
    currentScale = scale;
    var maxPan = Math.max(0, reportWidth - viewportWidth / scale);
    panX = clamp(panX, -maxPan, 0);
    scaleWrap.style.transform = 'scale(' + scale + ') translateX(' + panX + 'px)';
    viewport.style.width = '100vw';
    viewport.style.height = (report.offsetHeight * scale) + 'px';
  }}

  function setup() {{
    viewport = document.querySelector('.report-viewport');
    scaleWrap = document.querySelector('.report-scale');
    report = document.querySelector('.report');
    applyScale();
    window.addEventListener('resize', applyScale, {{ passive: true }});
    window.addEventListener('orientationchange', function() {{ setTimeout(applyScale, 200); }}, {{ passive: true }});
    if (!viewport) return;
    viewport.addEventListener('touchstart', function(event) {{
      if (event.touches.length === 2) {{
        pinchStartDistance = touchDistance(event.touches);
        pinchStartScale = userScale;
        dragDirection = '';
        event.preventDefault();
      }} else if (event.touches.length === 1 && userScale > 1) {{
        dragStartX = event.touches[0].clientX;
        dragStartY = event.touches[0].clientY;
        dragStartPanX = panX;
        dragDirection = '';
      }}
    }}, {{ passive: false }});
    viewport.addEventListener('touchmove', function(event) {{
      if (event.touches.length === 2 && pinchStartDistance) {{
        userScale = clamp(pinchStartScale * (touchDistance(event.touches) / pinchStartDistance), 1, 4);
        applyScale();
        event.preventDefault();
        return;
      }}
      if (event.touches.length !== 1 || userScale <= 1) return;
      var dx = event.touches[0].clientX - dragStartX;
      var dy = event.touches[0].clientY - dragStartY;
      if (!dragDirection && Math.max(Math.abs(dx), Math.abs(dy)) > 6) {{
        dragDirection = Math.abs(dx) > Math.abs(dy) ? 'horizontal' : 'vertical';
      }}
      if (dragDirection !== 'horizontal') return;
      panX = dragStartPanX + dx / currentScale;
      applyScale();
      event.preventDefault();
    }}, {{ passive: false }});
    viewport.addEventListener('touchend', function(event) {{
      if (event.touches.length < 2) pinchStartDistance = 0;
      if (event.touches.length === 0) dragDirection = '';
    }}, {{ passive: true }});
  }}

  if (document.readyState === 'loading') {{
    document.addEventListener('DOMContentLoaded', setup);
  }} else {{
    setup();
  }}
}})();
</script>
"""


def build_report_html(
    *,
    title: str,
    brand: str,
    products: list[str],
    report_date: date,
    meituan_path: Path | None,
    eleme_path: Path | None,
    jd_path: Path | None,
    social_paths: dict[str, Path],
    product_infos: dict[str, ProductInfo],
    launch_dates: dict[str, date | None],
    output_path: Path,
    configs: dict[str, dict[str, Any]],
    delivery_report: DeliveryReport | None = None,
    jd_report: tuple[tuple[JdSale, ...], float] | None = None,
    social_report_models: dict[str, SocialReport] | None = None,
    delivery_statuses: dict[str, str] | None = None,
) -> ReportBuildResult:
    warnings: list[str] = []
    delivery_statuses = delivery_statuses or {}
    if delivery_report is None:
        if not meituan_path or not eleme_path:
            raise RuntimeError("生成报告缺少美团或饿了么数据。")
        delivery = combine_delivery_sales(meituan_path, eleme_path, products)
    else:
        delivery = delivery_report
    jd_rows, jd_total = jd_report or read_jd_sales(jd_path)
    social_reports: dict[str, SocialReport] = dict(social_report_models or {})
    for product, path in social_paths.items():
        if product in social_reports:
            continue
        try:
            social_reports[product] = read_social_report(path)
        except Exception as exc:
            warnings.append(f"{product}：消费者反馈统计读取失败（{exc}）")

    layout = configs.get("html_layout", {})
    page = layout.get("page", {})
    fonts = layout.get("fonts", {})
    colors = layout.get("colors", {})
    tables = layout.get("tables", {})
    product_image = layout.get("productImage", {})
    font_css, font_family = _font_css(configs, warnings)
    logo_path = _safe_config_path(_text(layout.get("logo", {}).get("path")))
    logo_uri = _file_data_uri(logo_path) if logo_path else ""
    if not logo_uri:
        warnings.append("HTML 报告 Logo 不可访问，已使用纯文字标题")

    product_sections: list[str] = []
    for product in products:
        info = product_infos.get(product)
        if not info:
            warnings.append(f"{product}：周报产品信息暂无法获取")
        social = social_reports.get(product)
        if not social:
            warnings.append(f"{product}：消费者反馈统计暂无法获取")
        image_uri = ""
        if info and info.image_urls:
            try:
                image_uri = _remote_image_data_uri(info.image_urls[0])
            except Exception as exc:
                warnings.append(f"{product}：产品外观图片下载失败（{exc}）")
        price_suffix = f"（{_price_html(info.price)}）" if info and info.price else ""
        product_sections.append(
            f'<section class="product-section"><div class="pdf-keep-together product-info-module">'
            f'<h2 class="product-title"><span aria-hidden="true">●</span>'
            f'<span class="product-title-text">{_escape(product)}{price_suffix}</span></h2>'
            f'<h3>1. 产品信息</h3>{_product_info_html(product, info, image_uri)}</div>'
            f'<div class="pdf-keep-together feedback-summary-module">'
            f'<h3>2. 新品消费者反馈汇总</h3>{_social_summary_html(social)}</div>'
            f'<h3 class="feedback-detail-heading">3. 消费者反馈详情</h3>'
            f'{_social_detail_html(social)}</section>'
        )

    launch_values = [value for value in launch_dates.values() if value]
    launch_text = ""
    if launch_values and len(set(launch_values)) == 1:
        launch_text = f"{launch_values[0].month}.{launch_values[0].day} "
    intro_rows: list[str] = []
    for product in products:
        delivery_status = delivery_statuses.get(product)
        if delivery_status:
            intro_rows.append(
                f'<li><strong>{_escape(product)}</strong>：{_escape(delivery_status)}，暂无月销数据。</li>'
            )
            continue
        sale = delivery.tracked_rows.get(product)
        if sale:
            intro_rows.append(
                f'<li><strong>{_escape(product)}</strong>：销量排名第 {sale.rank}，线上外卖日店均 '
                f'{_format_daily(sale.combined_daily)} 杯，销量占比 {_format_percent(sale.share)}</li>'
            )
        else:
            warnings.append(f"{product}：未在美团和饿了么外卖数表中匹配到对应商品")
            intro_rows.append(f'<li><strong>{_escape(product)}</strong>：暂未匹配到外卖销售数据；</li>')

    logo_html = f'<img class="logo" src="{logo_uri}" alt="报告 Logo">' if logo_uri else ""
    title_text = f"{brand}：{'、'.join(products)}"
    document_title = title_text or title
    page_width = int(page.get("widthPx", 794))
    border_width = tables.get("borderWidthPx", 0.5)
    detail_line_height = tables.get("detailLineHeightPx", tables.get("lineHeightPx", 20))
    mobile_zoom_script = _mobile_zoom_script(page_width)
    css = f"""
{font_css}
*{{box-sizing:border-box}}
html,body{{margin:0;padding:0;background:#f2f2f0;-webkit-text-size-adjust:100%;text-size-adjust:100%}}
body{{width:100%;min-width:{page_width}px;color:#{colors.get('black','000000')};font-family:{font_family};font-size:{fonts.get('defaultSizePx',16)}px;line-height:{fonts.get('bodyLineHeightPx',26.67)}px;letter-spacing:{fonts.get('letterSpacingCm',0.02)}cm}}
.report-viewport{{width:{page_width}px;margin:24px auto;overflow:visible}}
.report-scale{{width:{page_width}px;transform-origin:top left}}
.report{{width:{page_width}px;min-height:1123px;margin:0;background:#{page.get('background','FFFFFF')};padding:{page.get('paddingTopPx',34)}px {page.get('paddingRightPx',44)}px {page.get('paddingBottomPx',52)}px {page.get('paddingLeftPx',44)}px;box-shadow:0 2px 18px rgba(0,0,0,.08)}}
.logo{{display:block;height:{layout.get('logo',{}).get('heightPx',72)}px;max-width:70%;object-fit:contain;margin:0 auto {layout.get('logo',{}).get('marginBottomPx',8)}px}}
h1{{font-size:{fonts.get('titleSizePx',21.33)}px;line-height:{fonts.get('bodyLineHeightPx',26.67)}px;text-align:center;margin:0 0 24px;font-weight:700}}
h2{{font-size:{fonts.get('productTitleSizePx',16)}px;line-height:{fonts.get('bodyLineHeightPx',26.67)}px;margin:0}}
h3{{font-size:{fonts.get('sectionSizePx',16)}px;line-height:{fonts.get('bodyLineHeightPx',26.67)}px;margin:20px 0 0}}
p{{margin:0}}.lead{{margin-bottom:10px}}.social-summary{{margin:0}}.sales-summary{{padding-left:20px;margin:4px 0 14px}}
.sales-summary li{{margin:2px 0}}.table-scroll{{width:100%;overflow:visible}}
table{{border-collapse:collapse;table-layout:fixed;width:100%;margin:8px auto 12px;font-size:{tables.get('fontSizePx',13.33)}px;line-height:{tables.get('lineHeightPx',20)}px}}
caption{{font-weight:700;font-size:{tables.get('fontSizePx',13.33)}px;line-height:{tables.get('lineHeightPx',20)}px;text-align:center;margin:8px 0 5px}}
th,td{{border:{border_width}px solid #{colors.get('black','000000')};padding:{tables.get('cellPaddingVerticalPx',0)}px {tables.get('cellPaddingHorizontalPx',6)}px;text-align:center;vertical-align:middle;word-break:normal;overflow-wrap:normal}}
thead th,.platform-header .detail-header,.feedback-title-cell{{background:#{colors.get('headerFill','D9D9D9')};font-weight:700}}
.tracked-row td{{background:#{colors.get('trackedFill','FCE4D6')}}}.total-row>*{{font-weight:700;background:#{colors.get('headerFill','D9D9D9')}}}
.sales-table .product-column{{width:36%}}.sales-table .platform-column,.sales-table .combined-column,.sales-table .share-column{{width:14.5%}}.sales-table .rank-column{{width:6%}}
.sales-table tbody td:first-child{{white-space:nowrap;word-break:keep-all}}
.jd-table .rank-column{{width:7%}}.jd-table .product-column{{width:59%}}.jd-table .sales-column{{width:20%}}.jd-table .share-column{{width:14%}}.jd-table tbody td:nth-child(2){{white-space:nowrap;word-break:keep-all}}
.rank-header,.rank-cell{{white-space:nowrap;word-break:keep-all;overflow-wrap:normal;padding-left:2px;padding-right:2px}}
.sales-table,.jd-table{{margin-bottom:0}}.source-note,.jd-note,.jd-missing-note{{line-height:18px;color:#{colors.get('sourceNote','999999')};text-align:justify;text-align-last:left;margin:1px 0 20px}}.source-note,.jd-note{{font-size:{fonts.get('sourceNoteSizePx',13.33)}px}}.jd-missing-note{{font-size:{fonts.get('defaultSizePx',16)}px}}
.missing-note{{border-left:3px solid #{colors.get('headerFill','D9D9D9')};padding:6px 10px;color:#{colors.get('muted','666666')}}}
.product-section{{margin-top:28px}}.product-title{{display:flex;align-items:baseline;gap:10px}}.product-title>span[aria-hidden="true"]{{font-size:16px;line-height:1}}.product-title+h3{{margin-top:0}}
.price-strike{{text-decoration:line-through}}
.product-info th{{width:18%;background:#{colors.get('white','FFFFFF')};font-weight:400}}.product-info td{{text-align:center}}
.product-info .name-row>*,.product-info .series-row>*{{background:#{colors.get('labelFill','E7EFFA')}}}.product-info .name-row>*{{font-weight:700}}
.product-info .long td{{text-align:justify}}.product-info .ingredients-row td{{text-align:center}}.image-row td{{height:calc({product_image.get('heightCm',4)}cm + 12px)}}
.product-image{{display:block;width:auto;height:{product_image.get('heightCm',4)}cm;max-width:{product_image.get('maxWidthPx',300)}px;max-height:{product_image.get('heightCm',4)}cm;margin:0 auto;object-fit:contain}}
.missing-inline,.empty-platform{{color:#{colors.get('muted','666666')}}}
.feedback-table .positive-text{{width:32%}}.feedback-table .positive-count{{width:8%}}.feedback-table .negative-text{{width:32%}}.feedback-table .negative-count{{width:8%}}.feedback-table .kpi-label-column{{width:14%}}.feedback-table .kpi-value-column{{width:6%}}
.feedback-table .count-header{{white-space:nowrap;word-break:keep-all;overflow-wrap:normal;padding-left:2px;padding-right:2px}}
.feedback-table .kpi-label,.feedback-table .kpi-value{{background:#{colors.get('white','FFFFFF')}}}.feedback-table .kpi-spacer{{border:0;background:transparent}}
.platform-total th,.platform-total td{{font-weight:700}} .product-info,.feedback-title-row{{break-inside:avoid}}
.feedback-detail-table-wrap{{width:100%;margin-top:8px}}.feedback-detail-table{{margin:0 auto;line-height:{detail_line_height}px}}
@media(max-width:767px){{html,body{{width:100%;min-width:0;overflow-x:hidden;background:#{page.get('background','FFFFFF')}}}body{{min-height:100%;-webkit-overflow-scrolling:touch}}.report-viewport{{width:100vw;margin:0;overflow:hidden;touch-action:pan-y pinch-zoom}}.report-scale{{margin:0;will-change:transform}}.report{{margin:0;box-shadow:none}}th,td{{border-width:1.5px}}}}
@media print{{@page{{size:A4 portrait;margin:10mm}}html,body{{width:auto!important;min-width:0;background:#fff}}.report-viewport,.report-scale{{width:auto!important;max-width:none!important;height:auto!important;overflow:visible!important;margin:0;transform:none!important}}.report{{width:auto!important;max-width:none!important;min-height:0;margin:0;padding:0;box-shadow:none}}.pdf-keep-together{{break-inside:avoid-page!important;page-break-inside:avoid!important}}.feedback-detail-heading{{break-after:avoid-page;page-break-after:avoid}}.feedback-detail-table-wrap,.feedback-detail-table{{break-inside:auto!important;page-break-inside:auto!important}}.feedback-table .feedback-title-row{{break-after:avoid-page;page-break-after:avoid}}.table-scroll{{width:100%;overflow:visible}}table{{width:100%!important;max-width:100%}}.table-scroll>table,.product-info{{width:calc(100% - 1px)!important;max-width:calc(100% - 1px)!important;margin-left:0;margin-right:1px}}thead{{display:table-header-group}}tr{{break-inside:avoid}}}}
"""
    document = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=yes">
<title>{_escape(document_title)}</title><style>{css}</style></head>
<body><div class="report-viewport"><div class="report-scale"><main class="report">{logo_html}<h1>{_escape(title_text)}</h1>
<div class="pdf-keep-together report-intro-module"><p class="lead"><strong>以下是{_escape(brand)} {launch_text}新品30日销量表现及消费者评论情况：</strong></p>
<ul class="sales-summary">{''.join(intro_rows)}</ul></div>
<div class="pdf-keep-together sales-module">{_delivery_html(brand, delivery)}</div>
<div class="pdf-keep-together jd-module">{_jd_html(brand, jd_rows, jd_total, products)}</div>{''.join(product_sections)}
</main></div></div>{mobile_zoom_script}</body></html>"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(document, encoding="utf-8")
    return ReportBuildResult(output_path, tuple(dict.fromkeys(warnings)))
