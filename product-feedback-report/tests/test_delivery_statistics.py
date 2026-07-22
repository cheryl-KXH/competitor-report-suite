from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.delivery.processing import (
    AnnotationRow,
    RawSaleRow,
    _annotation_lookup,
    _store_sale_days,
    generate_jd_summary,
    generate_platform_delivery_summary,
    read_annotation,
    tracked_product_statuses,
    write_product_menu,
)
from scripts.delivery.generate_tables import generate_delivery_tables


def sale(
    platform: str,
    store: str,
    product: str,
    monthly_sales: float,
    crawl_date: date,
    brand: str = "品牌A",
) -> RawSaleRow:
    return RawSaleRow(platform, brand, store, "", product, monthly_sales, crawl_date)


def annotation(
    platform: str,
    product: str,
    launch_date: date | None,
    brand: str = "品牌A",
) -> AnnotationRow:
    return AnnotationRow(platform, brand, product, launch_date)


class DeliveryStatisticsTests(unittest.TestCase):
    def test_product_menu_offers_manual_statuses_and_reads_them(self) -> None:
        rows = [sale("美团", "门店1", "在售产品", 10, date(2026, 7, 10))]
        with tempfile.TemporaryDirectory() as tmp:
            path = write_product_menu(rows, Path(tmp) / "产品清单.xlsx")
            workbook = load_workbook(path)
            sheet = workbook.active
            validations = list(sheet.data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].formula1, '"提前下架,外卖无售"')
            self.assertEqual(
                validations[0].prompt,
                "本列可人工增加上新日期标注；关注产品无数据时，请在各平台新增一行填入产品名称并标注"
                "“提前下架”或“外卖无售”",
            )
            sheet.append(["美团", "品牌A", "缺失产品A", "提前下架"])
            sheet.append(["饿了么", "品牌A", "缺失产品B", "外卖无售"])
            workbook.save(path)

            annotations = read_annotation(path)

        self.assertEqual(
            tracked_product_statuses(
                annotations, "品牌A", ["缺失产品A", "缺失产品B"]
            ),
            {"缺失产品A": "提前下架", "缺失产品B": "外卖无售"},
        )

    def test_product_menu_rejects_unknown_manual_status(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "产品清单.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["平台", "品牌", "产品", "近30日上新日期"])
            sheet.append(["美团", "品牌A", "缺失产品", "已下架"])
            workbook.save(path)

            with self.assertRaisesRegex(RuntimeError, "日期标注.*无效"):
                read_annotation(path)

    def test_product_menu_rejects_old_32_day_header(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "旧产品清单.xlsx"
            workbook = Workbook()
            workbook.active.append(["平台", "品牌", "产品", "近32日上新日期"])
            workbook.save(path)

            with self.assertRaisesRegex(RuntimeError, "缺少字段.*近30日上新日期"):
                read_annotation(path)

    def test_meituan_uses_latest_rows_and_sums_store_sale_days(self) -> None:
        rows = [
            sale("美团", "门店1", "产品A", 999, date(2026, 7, 9)),
            sale("美团", "门店1", "产品A", 80, date(2026, 7, 10)),
            sale("美团", "门店1", "产品A", 90, date(2026, 7, 10)),
            sale("美团", "门店2", "产品A", 0, date(2026, 7, 12)),
            sale("美团", "门店1", "产品B", 45, date(2026, 7, 12)),
            sale("美团", "门店2", "产品B", 45, date(2026, 7, 12)),
        ]
        annotations = [
            annotation("美团", "产品A", date(2026, 7, 5)),
            annotation("饿了么", "产品A", date(2026, 7, 1)),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "美团.xlsx"
            generate_platform_delivery_summary(rows, annotations, "美团", output)
            values = list(load_workbook(output, data_only=True).active.values)

        self.assertEqual(
            values[0],
            ("排名", "商品名称", "日店均销量", "总销量", "总销量占比", "总在售天数", "在售门店数", "在售天数", "上新日期"),
        )
        self.assertEqual(values[1], (1, "产品A", 6.4286, 90, 0.5, 14, 2, 8, datetime(2026, 7, 5)))
        self.assertEqual(values[2], (1, "产品B", 1.5, 90, 0.5, 60, 2, 30, None))

    def test_sale_days_boundaries_and_exact_annotation_key(self) -> None:
        lookup = _annotation_lookup(
            [
                annotation("美团", "产品A", date(2026, 7, 10)),
                annotation("饿了么", "产品A", date(2026, 7, 1)),
                annotation("美团", "产品A", date(2026, 6, 1), brand="品牌B"),
            ]
        )
        self.assertEqual(_store_sale_days(sale("美团", "门店1", "产品A", 1, date(2026, 7, 10)), lookup), 1)
        self.assertEqual(_store_sale_days(sale("美团", "门店1", "产品A", 1, date(2026, 7, 15)), lookup), 6)
        self.assertEqual(_store_sale_days(sale("美团", "门店1", "产品A", 1, date(2026, 8, 20)), lookup), 30)
        self.assertEqual(_store_sale_days(sale("美团", "门店1", "未标注", 1, date(2026, 7, 10)), lookup), 30)

    def test_delivery_output_uses_four_decimal_average_and_two_decimal_share(self) -> None:
        rows = [sale("饿了么", "门店1", "产品A", 1, date(2026, 7, 10))]
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "饿了么.xlsx"
            generate_platform_delivery_summary(rows, [], "饿了么", output)
            ws = load_workbook(output, data_only=True).active
            average = ws["C2"].value
            average_format = ws["C2"].number_format
            share_format = ws["E2"].number_format

        self.assertEqual(average, 0.0333)
        self.assertEqual(average_format, "0.0000")
        self.assertEqual(share_format, "0.00%")

    def test_future_launch_date_within_two_days_is_kept(self) -> None:
        rows = [sale("美团", "门店1", "产品A", 0, date(2026, 7, 10))]
        annotations = [annotation("美团", "产品A", date(2026, 7, 11))]
        lookup = _annotation_lookup(annotations)

        self.assertEqual(_store_sale_days(rows[0], lookup), -1)
        self.assertEqual(
            _store_sale_days(
                rows[0],
                _annotation_lookup([annotation("美团", "产品A", date(2026, 7, 12))]),
            ),
            -2,
        )

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "美团.xlsx"
            generate_platform_delivery_summary(rows, annotations, "美团", output)
            values = list(load_workbook(output, data_only=True).active.values)

        self.assertEqual(
            values[1],
            (1, "产品A", 0, 0, 0, -1, 1, -1, datetime(2026, 7, 11)),
        )

    def test_future_launch_date_over_two_days_reports_context(self) -> None:
        rows = [sale("美团", "门店1", "产品A", 1, date(2026, 7, 10))]
        annotations = [annotation("美团", "产品A", date(2026, 7, 13))]
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(
                RuntimeError,
                "平台=美团，品牌=品牌A，产品=产品A，上架日期=2026-07-13，抓取日期=2026-07-10",
            ):
                generate_platform_delivery_summary(rows, annotations, "美团", Path(tmp) / "美团.xlsx")

    def test_missing_store_is_rejected(self) -> None:
        rows = [sale("饿了么", "", "产品A", 1, date(2026, 7, 10))]
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(RuntimeError, "饿了么外卖数据缺少门店名称"):
                generate_platform_delivery_summary(rows, [], "饿了么", Path(tmp) / "饿了么.xlsx")

    def test_jd_averages_latest_store_snapshots_and_keeps_zero_sales(self) -> None:
        rows = [
            sale("京东秒送", "门店1", "产品A", 100, date(2026, 7, 9)),
            sale("京东秒送", "门店1", "产品A", 120, date(2026, 7, 10)),
            sale("京东外卖", "门店2", "产品A", 80, date(2026, 7, 10)),
            sale("京东", "门店1", "产品B", 100, date(2026, 7, 10)),
            sale("京东", "门店1", "产品C", 0, date(2026, 7, 10)),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "京东.xlsx"
            generate_jd_summary(rows, [annotation("京东秒送", "产品A", date(2026, 7, 5))], output)
            ws = load_workbook(output, data_only=True).active
            values = list(ws.values)
            share_format = ws["F2"].number_format

        self.assertEqual(values[0], ("排名", "商品名称", "销量加总", "在售门店数", "总销量", "总销量占比", "在售天数", "上新日期"))
        self.assertEqual(values[1], (1, "产品A", 200, 2, 100, 0.5, 6, datetime(2026, 7, 5)))
        self.assertEqual(values[2], (1, "产品B", 100, 1, 100, 0.5, 30, None))
        self.assertEqual(values[3], (3, "产品C", 0, 1, 0, 0, 30, None))
        self.assertEqual(share_format, "0.00%")

    def test_empty_platform_creates_header_only_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            mt_output = Path(tmp) / "美团.xlsx"
            jd_output = Path(tmp) / "京东.xlsx"
            generate_platform_delivery_summary([], [], "美团", mt_output)
            generate_jd_summary([], [], jd_output)
            mt_values = list(load_workbook(mt_output, data_only=True).active.values)
            jd_values = list(load_workbook(jd_output, data_only=True).active.values)

        self.assertEqual(len(mt_values), 1)
        self.assertEqual(len(jd_values), 1)
        self.assertEqual(mt_values[0][1], "商品名称")
        self.assertEqual(jd_values[0][1], "商品名称")

    def test_end_to_end_generation_from_source_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "raw"
            output_dir = root / "results"
            input_dir.mkdir()

            raw_wb = Workbook()
            raw_ws = raw_wb.active
            raw_ws.append(["平台", "品牌", "店铺名称", "分类描述", "宝贝名称", "原价", "现价", "宝贝月销", "抓取日期"])
            raw_ws.append(["美团", "品牌A", "门店1", "饮品", "产品A", 20, 18, 100, "2026-07-10"])
            raw_ws.append(["饿了么", "品牌A", "门店1", "饮品", "产品A", 20, 18, 60, "2026-07-10"])
            raw_ws.append(["京东秒送", "品牌A", "门店1", "饮品", "产品A", 20, 18, 80, "2026-07-10"])
            raw_wb.save(input_dir / "外卖原始数据.xlsx")

            annotation_wb = Workbook()
            annotation_ws = annotation_wb.active
            annotation_ws.append(["平台", "品牌", "产品", "近30日上新日期"])
            annotation_ws.append(["美团", "品牌A", "产品A", "2026-07-05"])
            annotation_ws.append(["饿了么", "品牌A", "产品A", "2026-07-08"])
            annotation_ws.append(["京东秒送", "品牌A", "产品A", "2026-07-05"])
            annotation_path = root / "产品清单.xlsx"
            annotation_wb.save(annotation_path)

            outputs = generate_delivery_tables("test-record", input_dir, annotation_path, output_dir)
            output_names = {key: path.name for key, path in outputs.items()}
            mt_values = list(load_workbook(outputs["meituanData"], data_only=True).active.values)
            eleme_values = list(load_workbook(outputs["elemeData"], data_only=True).active.values)
            jd_values = list(load_workbook(outputs["jdData"], data_only=True).active.values)

        self.assertEqual(
            output_names,
            {
                "meituanData": "品牌A-20260710-美团.xlsx",
                "elemeData": "品牌A-20260710-饿了么.xlsx",
                "jdData": "品牌A-20260710-京东.xlsx",
            },
        )
        self.assertEqual(mt_values[1], (1, "产品A", 16.6667, 100, 1, 6, 1, 6, datetime(2026, 7, 5)))
        self.assertEqual(eleme_values[1], (1, "产品A", 20, 60, 1, 3, 1, 3, datetime(2026, 7, 8)))
        self.assertEqual(jd_values[1], (1, "产品A", 80, 1, 80, 1, 6, datetime(2026, 7, 5)))


if __name__ == "__main__":
    unittest.main()
