from __future__ import annotations

from datetime import date
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from scripts.social.processing import (
    PlatformSummary,
    build_social_feedback_workbook,
    normalize_social_product_key,
    social_cleaned_workbook_products,
    social_workbook_products,
    split_social_workbook,
    summarize_social_cleaned_workbook,
    summarize_social_rows,
)
from service import dingtalk_table
from service.jobs import (
    _collect_social_inputs,
    _consumer_feedback_configs,
    _download_social_input_source,
    _social_cleaned_input_filename,
    _wait_for_ai_table_attachment_url,
)


def _write_cleaned_social_workbook(
    path: Path,
    product: str,
    *,
    extra_product: str | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    screenshot = workbook.create_sheet("大众点评-截图")
    screenshot.append(["评价内容", "情感识别", "标签-1"])
    screenshot.append(["忽略这张表", "负向", "不应被统计"])

    dianping = workbook.create_sheet("大众点评")
    dianping.append([None, None, None, None, None])
    dianping.append([None, "大众点评好评", "评论数", "大众点评差评", "评论数"])
    dianping.append([None, "清爽不腻", 3, "太甜", 2])
    dianping.append([None, "好评用户数", 4, "差评用户数", 2])

    header = [
        "品牌",
        "主贴id",
        "产品名称",
        "情感识别",
        "评价1-对应标签",
        "评价2-对应标签",
    ]
    for platform in ("微博", "小红书", "抖音", "B站"):
        sheet = workbook.create_sheet(platform)
        sheet.append(header)
    workbook["微博"].append(["古茗", "w1", product, "正向", "好喝", "清爽"])
    workbook["微博"].append(["古茗", "w2", product, "负向", "味道怪", None])
    workbook["小红书"].append(["古茗", "x1", product, "正向", "适合夏天", None])
    if extra_product:
        workbook["抖音"].append(["古茗", "d1", extra_product, "正向", "好喝", None])
    workbook.save(path)
    return path


class SocialFeedbackStatisticsTests(unittest.TestCase):
    def test_report_collection_reads_archived_dingtalk_social_link(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = _write_cleaned_social_workbook(
                root / "古茗-20260718-咸乳酪泰奶-社媒清洗数据.xlsx",
                "咸乳酪泰奶",
            )
            record = dingtalk_table.TaskRecord(
                "social",
                {
                    "brand": {"name": "古茗"},
                    "productName": "咸乳酪泰奶",
                    "socialCleanedRawData": [
                        {
                            "name": source.name,
                            "url": "https://alidocs.dingtalk.com/i/nodes/weibo-node",
                        }
                    ],
                },
            )
            product_key = normalize_social_product_key("咸乳酪泰奶")
            matches = {
                product_key: (
                    "咸乳酪泰奶",
                    record,
                    date(2026, 6, 18),
                )
            }

            with patch(
                "service.jobs._download_linked_social_file",
                return_value=source,
            ) as download_linked:
                collected = _collect_social_inputs(
                    {}, matches, root / "workspace"
                )

            expected_target = (
                root
                / "workspace"
                / "social_sources"
                / "social"
                / "socialCleanedRawData"
            )
            download_linked.assert_called_once_with(
                {},
                record,
                "socialCleanedRawData",
                "社媒CleanedRawData",
                target_dir=expected_target,
            )
            self.assertEqual(collected[product_key], source)

    def test_unreadable_dingtalk_social_link_requests_local_xlsx(self) -> None:
        record = dingtalk_table.TaskRecord(
            "social",
            {
                "socialCleanedRawData": [
                    {
                        "name": "古茗-20260718-咸乳酪泰奶-社媒清洗数据.xlsx",
                        "url": "https://alidocs.dingtalk.com/i/nodes/weibo-node",
                    }
                ]
            },
        )
        with patch(
            "service.jobs._download_linked_social_file",
            side_effect=RuntimeError("下载超时"),
        ):
            with self.assertRaisesRegex(
                RuntimeError,
                r"\[古茗-20260718-咸乳酪泰奶-社媒清洗数据\.xlsx\]已归档为钉钉文档链接导致读取失败，"
                "请重新从本地上传原始文件，再进行报告生成",
            ):
                _download_social_input_source(
                    {}, record, "socialCleanedRawData", "社媒CleanedRawData"
                )

    def test_report_lists_unreadable_archived_cleaned_file(self) -> None:
        record = dingtalk_table.TaskRecord(
            "social",
            {
                "brand": {"name": "古茗"},
                "productName": "青橘芦荟冰冰茶",
                "socialCleanedRawData": [
                    {
                        "name": "古茗-20260718-青橘芦荟冰冰茶-社媒清洗数据.xlsx",
                        "url": "https://alidocs.dingtalk.com/i/nodes/weibo-node",
                    }
                ],
            },
        )
        product_key = normalize_social_product_key("青橘芦荟冰冰茶")
        matches = {
            product_key: (
                "青橘芦荟冰冰茶",
                record,
                date(2026, 6, 18),
            )
        }

        with tempfile.TemporaryDirectory() as temporary, patch(
            "service.jobs._download_linked_social_file",
            side_effect=RuntimeError("无可下载 URL"),
        ) as download_linked:
            with self.assertRaisesRegex(
                RuntimeError,
                r"\[古茗-20260718-青橘芦荟冰冰茶-社媒清洗数据\.xlsx\]"
                "已归档为钉钉文档链接导致读取失败",
            ):
                _collect_social_inputs(
                    {}, matches, Path(temporary) / "workspace"
                )

        self.assertEqual(download_linked.call_count, 1)

    def test_social_workbook_is_split_by_normalized_product_and_keeps_formatting(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "合并.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["品牌", "产 品 名 称", "内容"])
            sheet.append(["霸王茶姬", "糯青山 柠檬奶", "A1"])
            sheet.append(["霸王茶姬", "雾红尘-柠檬奶", "B1"])
            sheet.append(["霸王茶姬", "糯青山柠檬奶", "A2"])
            sheet["A1"].fill = PatternFill("solid", fgColor="FF0000")
            sheet.column_dimensions["C"].width = 42
            workbook.save(source)

            first_key = normalize_social_product_key("糯青山柠檬奶")
            second_key = normalize_social_product_key("雾红尘柠檬奶")
            self.assertEqual(
                set(social_workbook_products(source)), {first_key, second_key}
            )
            outputs = split_social_workbook(
                source,
                product_names={
                    first_key: "糯青山柠檬奶",
                    second_key: "雾红尘柠檬奶",
                },
                output_dir=root / "split",
                output_names={first_key: "产品A.xlsx", second_key: "产品B.xlsx"},
            )

            first = load_workbook(outputs[first_key])
            second = load_workbook(outputs[second_key])
            try:
                self.assertEqual(
                    [row[2] for row in first.active.iter_rows(min_row=2, values_only=True)],
                    ["A1", "A2"],
                )
                self.assertEqual(
                    [row[2] for row in second.active.iter_rows(min_row=2, values_only=True)],
                    ["B1"],
                )
                self.assertEqual(first.active["A1"].fill.fgColor.rgb, "00FF0000")
                self.assertEqual(first.active.column_dimensions["C"].width, 42)
            finally:
                first.close()
                second.close()

    def test_social_workbook_rejects_missing_or_blank_product_name(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            missing = root / "missing.xlsx"
            workbook = Workbook()
            workbook.active.append(["品牌", "内容"])
            workbook.active.append(["品牌", "评论"])
            workbook.save(missing)
            with self.assertRaisesRegex(RuntimeError, "缺少“产品名称”"):
                social_workbook_products(missing)

            blank = root / "blank.xlsx"
            workbook = Workbook()
            workbook.active.append(["品牌", "产品名称", "内容"])
            workbook.active.append(["品牌", "", "评论"])
            workbook.save(blank)
            with self.assertRaisesRegex(RuntimeError, "产品名称.*为空"):
                social_workbook_products(blank)

    def test_all_row_tags_and_users_follow_sentiment(self) -> None:
        rows = [
            ["主贴id", "情感识别", "评价1-对应标签", "评价2-对应标签"],
            ["p1", "正向", "好喝", "清爽"],
            ["p2", "负向", "太甜", "味道怪"],
            ["p3", "中性", "中性标签", ""],
        ]
        summary = summarize_social_rows("weibo", "微博", rows)
        self.assertEqual(summary.positive_tags, (("好喝", 1), ("清爽", 1)))
        self.assertEqual(summary.negative_tags, (("味道怪", 1), ("太甜", 1)))
        self.assertEqual(summary.positive_users, 1)
        self.assertEqual(summary.negative_users, 1)
        self.assertEqual(summary.positive_count, 2)
        self.assertEqual(summary.negative_count, 2)

    def test_missing_required_columns_is_explicit(self) -> None:
        with self.assertRaisesRegex(RuntimeError, "情感识别"):
            summarize_social_rows("weibo", "微博", [["主贴id"], ["p1"]])

    def test_real_export_shape_supports_five_evaluation_groups(self) -> None:
        header = ["品牌", "主贴id", "产品名称", "搜索关键词", "用户名", "内容", "发布时间", "链接", "情感识别"]
        for index in range(1, 6):
            header.append(f"评价{index}-对应标签")
        row = ["茉莉奶白", "p1", "青芒黄皮冰茶", "关键词", "用户", "内容", "2026-07-09", "链接", "正向"]
        for index in range(1, 6):
            row.append(f"标签{index}")
        summary = summarize_social_rows("douyin", "抖音", [header, row])
        self.assertEqual(sum(count for _, count in summary.positive_tags), 5)
        self.assertEqual(summary.positive_users, 1)

    def test_bilibili_export_shape_counts_three_evaluation_groups(self) -> None:
        rows = [
            [
                "品牌",
                "主贴id",
                "产品名称",
                "搜索关键词",
                "用户名",
                "内容",
                "发布时间",
                "链接",
                "情感识别",
                "评价1-对应标签",
                "评价2-对应标签",
                "评价3-对应标签",
            ],
            [
                "霸王茶姬",
                "BV1FWjn6KEmh",
                "糯青山柠檬奶",
                "关键词",
                "用户",
                "内容",
                "2026-06-20 16:29:13",
                "链接",
                "负向",
                "难喝，不喜欢，不推荐",
                "奶盖拉跨",
                "性价比低/贵",
            ],
        ]
        summary = summarize_social_rows("bilibili", "B站", rows)
        self.assertEqual(
            summary.negative_tags,
            (("奶盖拉跨", 1), ("性价比低/贵", 1), ("难喝，不喜欢，不推荐", 1)),
        )
        self.assertEqual(summary.negative_users, 1)

    def test_cleaned_workbook_uses_prepared_dianping_and_raw_platform_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = _write_cleaned_social_workbook(
                Path(temporary) / "古茗-20260718-青橘芦荟冰冰茶-社媒清洗数据.xlsx",
                "青橘芦荟冰冰茶",
            )
            summaries = summarize_social_cleaned_workbook(source)

        by_key = {summary.key: summary for summary in summaries}
        self.assertEqual([summary.key for summary in summaries], [
            "dianping",
            "weibo",
            "xiaohongshu",
            "douyin",
            "bilibili",
        ])
        self.assertEqual(by_key["dianping"].positive_tags, (("清爽不腻", 3),))
        self.assertEqual(by_key["dianping"].negative_tags, (("太甜", 2),))
        self.assertEqual(
            (by_key["dianping"].positive_users, by_key["dianping"].negative_users),
            (4, 2),
        )
        self.assertEqual(by_key["weibo"].positive_tags, (("好喝", 1), ("清爽", 1)))
        self.assertEqual(by_key["weibo"].negative_tags, (("味道怪", 1),))
        self.assertEqual((by_key["bilibili"].positive_users, by_key["bilibili"].negative_users), (0, 0))

    def test_cleaned_workbook_product_and_archive_filename_follow_new_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = _write_cleaned_social_workbook(
                Path(temporary) / "input.xlsx", "青橘芦荟冰冰茶"
            )
            products = social_cleaned_workbook_products(source)
        self.assertEqual(
            products,
            {normalize_social_product_key("青橘芦荟冰冰茶"): "青橘芦荟冰冰茶"},
        )
        self.assertEqual(
            _social_cleaned_input_filename(
                "古茗", "青橘芦荟冰冰茶", date(2026, 7, 18)
            ),
            "古茗-20260718-青橘芦荟冰冰茶-社媒清洗数据.xlsx",
        )

    def test_workbook_matches_single_product_summary_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = build_social_feedback_workbook(
                brand="测试品牌",
                product="测试新品",
                start_date="6.10",
                end_date="7.9",
                summaries=[
                    PlatformSummary("weibo", "微博", (("好喝", 2),), (("太甜", 1),), 3, 1),
                    PlatformSummary("xiaohongshu", "小红书", (), (), 0, 0),
                    PlatformSummary("douyin", "抖音", (("清爽", 1),), (), 1, 0),
                    PlatformSummary("bilibili", "B站", (), (("奶盖拉跨", 1),), 0, 1),
                ],
                output_dir=Path(temporary),
            )
            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook["测试新品"]
                self.assertEqual(sheet["A1"].value, "测试新品 6.10-7.9 第三方平台评价反馈")
                self.assertEqual(sheet["F1"].value, 4 / 6)
                self.assertEqual(
                    [sheet[f"E{row}"].value for row in range(1, 5)],
                    ["好评率", "总计", "好评用户数", "差评用户数"],
                )
                self.assertEqual(
                    [sheet[f"F{row}"].value for row in range(2, 5)],
                    [6, 4, 2],
                )
                self.assertIsNone(sheet["E5"].value)
                self.assertIsNone(sheet["F5"].value)
                self.assertNotEqual(sheet["F1"].data_type, "f")
                self.assertTrue(
                    all(sheet[f"E{row}"].font.color.rgb == "00000000" for row in range(1, 5))
                )
                self.assertIn("小红书好评", [cell.value for cell in sheet["A"]])
                self.assertIn("B站好评", [cell.value for cell in sheet["A"]])
                xiaohongshu_header = next(
                    cell.row for cell in sheet["A"] if cell.value == "小红书好评"
                )
                self.assertEqual(sheet.cell(xiaohongshu_header + 1, 1).value, "好评用户数")
                bilibili_header = next(
                    cell.row for cell in sheet["A"] if cell.value == "B站好评"
                )
                self.assertIsNone(sheet.cell(bilibili_header + 1, 1).value)
                self.assertIsNone(sheet.cell(bilibili_header + 1, 2).value)
                self.assertEqual(sheet.cell(bilibili_header + 1, 3).value, "奶盖拉跨")
                self.assertEqual(sheet.cell(bilibili_header + 1, 4).value, 1)
                self.assertIn("A1:D1", [str(item) for item in sheet.merged_cells.ranges])
                self.assertIn("$A$1:$F$", str(sheet.print_area))
                self.assertEqual(sheet.page_setup.fitToWidth, 1)
                self.assertEqual(sheet.page_setup.fitToHeight, 0)
                self.assertEqual(sheet.column_dimensions["A"].width, 27)
            finally:
                workbook.close()


class SocialFeedbackJobTests(unittest.TestCase):
    def test_consumer_table_mapping_uses_current_social_fields(self) -> None:
        configs = _consumer_feedback_configs({"dingtalk": {"tableId": "old"}})
        fields = configs["field_mapping"]["fields"]
        self.assertEqual(configs["dingtalk"]["tableId"], "hvcu7Bw")
        self.assertEqual(fields["launchDate"]["fieldId"], "mKUEya0")
        self.assertEqual(fields["day30"]["fieldId"], "SSrz1N1")
        self.assertEqual(fields["report"]["fieldId"], "amWTton")
        self.assertEqual(fields["report"]["cellType"], "link")
        self.assertEqual(fields["feedback"]["fieldId"], "j8IgB7P")
        self.assertEqual(fields["socialCleanedRawData"]["fieldId"], "Q4mbLWx")
        for old_platform in ("dianping", "weibo", "xiaohongshu", "douyin", "bilibili"):
            self.assertNotIn(old_platform, fields)
        self.assertNotIn("allData", fields)

    def test_local_attachment_url_is_polled_from_ai_table(self) -> None:
        initial = dingtalk_table.TaskRecord(
            "record",
            {
                "socialCleanedRawData": [
                    {"filename": "社媒清洗数据.xlsx", "resourceId": "r1"}
                ]
            },
        )
        refreshed = dingtalk_table.TaskRecord(
            "record",
            {
                "socialCleanedRawData": [
                    {
                        "filename": "社媒清洗数据.xlsx",
                        "resourceId": "r1",
                        "resourceUrl": "https://example.test/social-cleaned.xlsx",
                    }
                ]
            },
        )
        with patch("service.jobs.dingtalk_table.fetch_record", return_value=refreshed) as fetch:
            result = _wait_for_ai_table_attachment_url(
                {},
                initial,
                "socialCleanedRawData",
                "社媒CleanedRawData",
                attempts=2,
                delay_seconds=0,
            )
        self.assertEqual(
            result,
            ("社媒清洗数据.xlsx", "https://example.test/social-cleaned.xlsx"),
        )
        fetch.assert_called_once_with({}, "record")

    def test_fetch_records_reads_all_ai_table_pages(self) -> None:
        configs = _consumer_feedback_configs(
            {"dingtalk": {"baseId": "base", "tableId": "ignored"}}
        )
        brand_field = configs["field_mapping"]["fields"]["brand"]["fieldId"]
        with patch(
            "service.dingtalk_table.call_table_tool",
            side_effect=[
                {
                    "data": {
                        "records": [
                            {"recordId": "r1", "cells": {brand_field: {"name": "茉莉奶白"}}}
                        ],
                        "nextCursor": "next",
                    }
                },
                {
                    "data": {
                        "records": [
                            {"recordId": "r2", "cells": {brand_field: {"name": "瑞幸"}}}
                        ]
                    }
                },
            ],
        ) as query:
            records = dingtalk_table.fetch_records(configs)

        self.assertEqual([record.record_id for record in records], ["r1", "r2"])
        self.assertEqual(records[0].cells["brand"], {"name": "茉莉奶白"})
        self.assertEqual(query.call_args_list[0].args[2]["limit"], 100)
        self.assertNotIn("cursor", query.call_args_list[0].args[2])
        self.assertEqual(query.call_args_list[1].args[2]["cursor"], "next")

if __name__ == "__main__":
    unittest.main()
