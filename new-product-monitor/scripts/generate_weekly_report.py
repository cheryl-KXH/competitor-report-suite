#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import concurrent.futures
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    from PIL import Image as PILImage
except ModuleNotFoundError as exc:
    missing = exc.name or "dependency"
    raise SystemExit(
        f"缺少 Python 依赖：{missing}。\n"
        "请先运行：pip install -r requirements.txt\n"
        "或使用 Codex bundled Python 运行本脚本。"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"


@dataclass
class ReportRecord:
    record_id: str
    brand: str
    product_name: str
    launch_date: date | None
    price: str
    category: str
    series: str
    selling_point: str
    ingredients: str
    image_urls: list[str]
    remark: str
    source_index: int = 0


@dataclass
class DataQualityReport:
    missing_fields: list[str]
    missing_images: list[str]
    image_download_failures: list[str]

    @property
    def missing_field_count(self) -> int:
        return len(self.missing_fields)

    @property
    def image_issue_count(self) -> int:
        return len(self.missing_images) + len(self.image_download_failures)


@dataclass
class CachedImage:
    path: Path
    content_type: str


class ImageCache:
    def __init__(self, timeout_seconds: int = 30, retry_count: int = 3) -> None:
        self.timeout_seconds = timeout_seconds
        self.retry_count = retry_count
        self._manager = tempfile.TemporaryDirectory(prefix="weekly_report_images_")
        self.root = Path(self._manager.name)
        self._cache: dict[str, CachedImage] = {}
        self._lock = threading.Lock()

    def cleanup(self) -> None:
        self._manager.cleanup()

    def prefetch(self, records: list[ReportRecord], max_workers: int = 1) -> None:
        jobs: dict[str, tuple[str, str]] = {}
        for record in records:
            if record.image_urls:
                jobs.setdefault(record.image_urls[0], (record.record_id, record.image_urls[0]))
        if not jobs:
            return
        workers = max(1, min(max_workers, 2, len(jobs)))
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            list(executor.map(lambda item: self.fetch(item[1], item[0], log_failure=False), jobs.values()))

    def fetch(self, url: str, record_id: str, log_failure: bool = True) -> CachedImage | None:
        if not url:
            return None
        with self._lock:
            if url in self._cache:
                return self._cache[url]

        clean_id = re.sub(r"[^A-Za-z0-9_-]+", "_", record_id or "image")
        digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]
        target = self.root / f"{clean_id}_{digest}.img"
        last_exc: Exception | None = None
        for attempt in range(1, self.retry_count + 1):
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=self.timeout_seconds) as resp:
                    content_type = resp.headers.get("Content-Type", "image/png").split(";")[0].strip() or "image/png"
                    target.write_bytes(resp.read())
                cached = CachedImage(target, content_type)
                with self._lock:
                    self._cache[url] = cached
                return cached
            except Exception as exc:
                last_exc = exc
                if attempt < self.retry_count:
                    time.sleep(float(attempt))

        if log_failure:
            print(f"WARNING: 图片下载失败：{record_id} {last_exc}", file=sys.stderr)
        else:
            print(f"INFO: 图片预下载未成功，生成时将重试：{record_id} {last_exc}", file=sys.stderr)
        return None

    def data_uri(self, url: str, record_id: str) -> str:
        cached = self.fetch(url, record_id)
        if not cached:
            return ""
        data = base64.b64encode(cached.path.read_bytes()).decode("ascii")
        return f"data:{cached.content_type};base64,{data}"


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_configs() -> dict[str, dict[str, Any]]:
    configs = {
        "dingtalk": load_json(CONFIG_DIR / "dingtalk.json"),
        "field_mapping": load_json(CONFIG_DIR / "field_mapping.json"),
        "report_rules": load_json(CONFIG_DIR / "report_rules.json"),
        "excel_layout": load_json(CONFIG_DIR / "excel_layout.json"),
        "font_files": load_json(CONFIG_DIR / "font_files.json"),
    }
    return configs


def configured_font_path(configs: dict[str, dict[str, Any]], font_key: str) -> Path | None:
    font_files = configs.get("font_files", {})
    font_cfg = font_files.get(font_key, {})
    file_name = str(font_cfg.get("fileName") or "").strip()
    if not file_name or Path(file_name).name != file_name:
        return None
    font_dir_value = Path(str(font_files.get("fontDirectory") or "assets/fonts"))
    if font_dir_value.is_absolute():
        return None
    font_dir = ROOT / font_dir_value
    path = font_dir / file_name
    return path if path.exists() else None


def configured_excel_font_name(configs: dict[str, dict[str, Any]], font_key: str) -> str:
    font_files = configs.get("font_files", {})
    font_cfg = font_files.get(font_key, {})
    default_fonts = font_files.get("defaultFonts", {})
    default_key = "latinExcelName" if font_key == "latinFont" else "chineseExcelName"
    default_name = font_cfg.get("defaultExcelName") or default_fonts.get(default_key) or "Arial"
    if configured_font_path(configs, font_key):
        return str(font_cfg.get("excelName") or default_name)
    return str(default_name)


def validate_config(configs: dict[str, dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    dingtalk = configs["dingtalk"]
    if not dingtalk.get("baseId"):
        warnings.append("config/dingtalk.json 缺少 baseId")
    if not dingtalk.get("tableId"):
        warnings.append("config/dingtalk.json 缺少 tableId")
    if not dingtalk.get("streamableHttpUrl") and not dingtalk.get("serverName"):
        warnings.append("config/dingtalk.json 需要 streamableHttpUrl 或 serverName")

    fields = configs["field_mapping"].get("standardFields", {})
    for key in ("brand", "productName", "launchDate"):
        if key not in fields:
            warnings.append(f"field_mapping.json 缺少必需标准字段：{key}")

    rules = configs["report_rules"]
    business_values = rules.get("businessFilter", {}).get("allowedValues", [])
    if not business_values:
        warnings.append("report_rules.json 缺少 businessFilter.allowedValues")
    tracked_table = rules.get("trackedBrandsTable", {})
    for key in ("tableId", "brandFieldId", "businessFieldId", "statusFieldId", "activeStatusOptionId"):
        if not tracked_table.get(key):
            warnings.append(f"report_rules.json 缺少 trackedBrandsTable.{key}")
    weekly_table = rules.get("weeklyReportTable", {})
    if not weekly_table.get("tableId"):
        warnings.append("report_rules.json 缺少 weeklyReportTable.tableId")
    weekly_fields = weekly_table.get("fields", {})
    for key in ("reportUrl", "business", "startDate", "endDate", "year", "week", "status", "generatedAt"):
        if not weekly_fields.get(key):
            warnings.append(f"report_rules.json 缺少 weeklyReportTable.fields.{key}")
    weekly_statuses = weekly_table.get("statusOptionIds", {})
    for status_name in ("未生成", "等待1-2分钟", "已生成", "生成异常", "生成失败"):
        if not weekly_statuses.get(status_name):
            warnings.append(f"report_rules.json 缺少 weeklyReportTable.statusOptionIds.{status_name}")
    if not rules.get("categoryRule", {}).get("priority"):
        warnings.append("report_rules.json 缺少 categoryRule.priority")

    layout = configs["excel_layout"]
    if not layout.get("worksheetName"):
        warnings.append("excel_layout.json 缺少 worksheetName")
    image_cm = layout.get("image", {}).get("heightCm")
    if not isinstance(image_cm, (int, float)) or image_cm <= 0:
        warnings.append("excel_layout.json 的 image.heightCm 必须为正数")

    font_files = configs["font_files"]
    for font_key in ("chineseFont", "latinFont"):
        font_cfg = font_files.get(font_key, {})
        if not configured_font_path(configs, font_key):
            fallback = configured_excel_font_name(configs, font_key)
            file_name = font_cfg.get("fileName") or "未配置"
            warnings.append(f"font_files.json 中 {font_key} 的字体文件不可访问：{file_name}；将使用默认字体 {fallback}")
    return warnings


def explain_config(configs: dict[str, dict[str, Any]]) -> str:
    dingtalk = configs["dingtalk"]
    fields = configs["field_mapping"]["standardFields"]
    rules = configs["report_rules"]
    layout = configs["excel_layout"]
    font_files = configs["font_files"]
    lines = [
        "当前配置摘要：",
        f"- 钉钉连接：baseId={dingtalk.get('baseId')}，tableId={dingtalk.get('tableId')}，"
        f"连接方式={'streamableHttpUrl' if dingtalk.get('streamableHttpUrl') else dingtalk.get('serverName')}",
        f"- 默认周期：周五导出上周六到本周五，周一导出上一个完整周六到周五",
        f"- 默认品牌：不传 --brands 或 --business 时，按本次钉钉表取数结果里的品牌输出，顺序按钉钉品牌字段标签列表",
        f"- 业务筛选：{ '、'.join(rules.get('businessFilter', {}).get('allowedValues', [])) }",
        f"- 关注品牌小字：传 --business 时读取 {rules.get('trackedBrandsTable', {}).get('tableName', '关注竞品品牌')} 表中对应业务且进行中的品牌",
        f"- 按钮状态机表：{rules.get('weeklyReportTable', {}).get('tableName', '竞品新品周报')}",
        f"- 品类规则：{' > '.join(rules.get('categoryRule', {}).get('priority', []))} > 主品类",
        f"- 备注规则：{'，'.join(rules.get('remarkRule', {}).get('order', []))}",
        f"- 图片高度：{layout.get('image', {}).get('heightCm')}cm",
        f"- 字体目录：{font_files.get('fontDirectory', 'assets/fonts')}",
        f"- 中文字体：{configured_excel_font_name(configs, 'chineseFont')}",
        f"- 英文字体：{configured_excel_font_name(configs, 'latinFont')}",
        "- 标准字段：",
    ]
    for key, cfg in fields.items():
        names = " / ".join(cfg.get("fieldNames", []))
        lines.append(f"  - {key}: {cfg.get('label')} ({names})")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从钉钉 AI 表生成竞品新品周报 Excel")
    parser.add_argument("--startDate", help="开始日期，格式 YYYYMMDD，例如 20260502")
    parser.add_argument("--endDate", help="结束日期，格式 YYYYMMDD，例如 20260508")
    parser.add_argument("--brands", help="品牌范围，使用英文逗号分隔，例如：霸王茶姬,古茗")
    parser.add_argument("--business", help="业务范围：喜茶、野萃山、茶坊；--brands 优先于本参数")
    parser.add_argument("--output", help="输出 xlsx 路径")
    parser.add_argument("--validate-config", action="store_true", help="校验配置后退出")
    parser.add_argument("--explain-config", action="store_true", help="解释当前配置后退出")
    return parser.parse_args()


def parse_compact_date(value: str, option_name: str) -> date:
    if not re.fullmatch(r"\d{8}", value or ""):
        raise SystemExit(f"{option_name} 格式必须为 YYYYMMDD，例如 20260502")
    try:
        return datetime.strptime(value, "%Y%m%d").date()
    except ValueError as exc:
        raise SystemExit(f"{option_name} 不是有效日期：{value}") from exc


def resolve_date_window(args: argparse.Namespace) -> tuple[date, date]:
    if args.startDate or args.endDate:
        if not args.startDate or not args.endDate:
            raise SystemExit("--startDate 和 --endDate 必须同时提供")
        start = parse_compact_date(args.startDate, "--startDate")
        end = parse_compact_date(args.endDate, "--endDate")
        if start > end:
            raise SystemExit("--startDate 不能晚于 --endDate")
        return start, end

    today = date.today()
    friday_weekday = 4
    # Friday uses the current day as the window end; Monday rolls back to the previous Friday.
    days_since_friday = (today.weekday() - friday_weekday) % 7
    end = today - timedelta(days=days_since_friday)
    start = end - timedelta(days=6)
    return start, end


def parse_brands(arg: str | None) -> list[str]:
    if not arg:
        return []
    brands = [part.strip() for part in arg.split(",") if part.strip()]
    if not brands:
        raise SystemExit("--brands 不能为空")
    return brands


def parse_business(arg: str | None, rules: dict[str, Any]) -> str | None:
    if not arg:
        return None
    business = arg.strip()
    allowed = rules.get("businessFilter", {}).get("allowedValues", [])
    if business not in allowed:
        available = "、".join(allowed) or "无"
        raise SystemExit(f"--business 不存在：{business}。可用值：{available}")
    return business


def mcporter_selector(dingtalk_cfg: dict[str, Any], tool_name: str) -> list[str]:
    url = dingtalk_cfg.get("streamableHttpUrl", "").strip()
    if url:
        return ["mcporter", "call", url, f".{tool_name}"]
    server = dingtalk_cfg.get("serverName", "dingtalk-ai-table")
    return ["mcporter", "call", f"{server}.{tool_name}"]


def call_dingtalk_tool(dingtalk_cfg: dict[str, Any], tool_name: str, payload: dict[str, Any]) -> dict[str, Any]:
    cmd = mcporter_selector(dingtalk_cfg, tool_name) + ["--args", json.dumps(payload, ensure_ascii=False)]
    try:
        result = subprocess.run(cmd, check=False, capture_output=True, text=True)
    except FileNotFoundError as exc:
        raise SystemExit("找不到 mcporter，请先安装并配置钉钉 AI 表 MCP。") from exc

    if result.returncode != 0:
        raise SystemExit(f"mcporter 调用失败：{result.stderr.strip() or result.stdout.strip()}")
    try:
        data = parse_mcporter_json(result.stdout)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"mcporter 返回不是 JSON：{result.stdout[:1000]}") from exc
    if data.get("status") not in (None, "success", "ok"):
        raise SystemExit(f"钉钉工具返回失败：{json.dumps(data.get('error', data), ensure_ascii=False)}")
    return data


def parse_mcporter_json(stdout: str) -> dict[str, Any]:
    text = stdout.strip()
    if not text:
        raise json.JSONDecodeError("empty stdout", stdout, 0)
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        if start < 0:
            raise
        decoder = json.JSONDecoder()
        data, _ = decoder.raw_decode(text[start:])
    if not isinstance(data, dict):
        raise json.JSONDecodeError("mcporter JSON root is not an object", stdout, 0)
    return data


def fetch_table_fields(configs: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    dingtalk = configs["dingtalk"]
    payload = {"baseId": dingtalk["baseId"], "tableIds": [dingtalk["tableId"]]}
    data = call_dingtalk_tool(dingtalk, "get_tables", payload)
    tables = data.get("data", {}).get("tables", [])
    if not tables:
        raise SystemExit("没有读取到钉钉表结构，请检查 baseId/tableId。")
    return tables[0].get("fields", [])


def resolve_field_ids(field_mapping: dict[str, Any], table_fields: list[dict[str, Any]]) -> dict[str, str]:
    by_id = {f.get("fieldId"): f for f in table_fields if f.get("fieldId")}
    by_name = {f.get("fieldName"): f for f in table_fields if f.get("fieldName")}
    resolved: dict[str, str] = {}
    missing_required: list[str] = []

    for key, cfg in field_mapping.get("standardFields", {}).items():
        field_id = cfg.get("fieldId")
        if field_id and field_id in by_id:
            resolved[key] = field_id
            continue
        for name in cfg.get("fieldNames", []):
            if name in by_name:
                resolved[key] = by_name[name]["fieldId"]
                break
        if key not in resolved and cfg.get("required"):
            missing_required.append(f"{key} ({cfg.get('label')})")

    if missing_required:
        raise SystemExit("缺少必需钉钉字段：" + "、".join(missing_required))
    return resolved


def extract_option_name(option: Any) -> str:
    if isinstance(option, str):
        return option.strip()
    if isinstance(option, dict):
        for key in ("name", "label", "text", "value"):
            value = option.get(key)
            if value:
                return str(value).strip()
    return ""


def extract_options_from_container(container: Any) -> list[str]:
    if not isinstance(container, dict):
        return []
    for key in ("options", "choices", "items"):
        raw_options = container.get(key)
        if isinstance(raw_options, list):
            names = [extract_option_name(option) for option in raw_options]
            return [name for name in names if name]
    return []


def brand_option_order(table_fields: list[dict[str, Any]], field_ids: dict[str, str]) -> list[str]:
    brand_field_id = field_ids.get("brand")
    if not brand_field_id:
        return []
    brand_field = next((field for field in table_fields if field.get("fieldId") == brand_field_id), None)
    if not brand_field:
        return []
    candidates = [
        brand_field,
        brand_field.get("property"),
        brand_field.get("fieldProperty"),
        brand_field.get("typeOptions"),
        brand_field.get("config"),
    ]
    for candidate in candidates:
        options = extract_options_from_container(candidate)
        if options:
            return list(dict.fromkeys(options))
    return []


def field_option_order(configs: dict[str, dict[str, Any]], table_id: str, field_id: str) -> list[str]:
    dingtalk = configs["dingtalk"]
    payload = {"baseId": dingtalk["baseId"], "tableId": table_id, "fieldIds": [field_id]}
    data = call_dingtalk_tool(dingtalk, "get_fields", payload)
    fields = data.get("data", {}).get("fields", []) or []
    if not fields:
        return []
    field = fields[0]
    for candidate in (field, field.get("config"), field.get("property"), field.get("fieldProperty")):
        options = extract_options_from_container(candidate)
        if options:
            return list(dict.fromkeys(options))
    return []


def default_brand_option_order(configs: dict[str, dict[str, Any]], field_ids: dict[str, str]) -> list[str]:
    brand_field_id = field_ids.get("brand")
    if not brand_field_id:
        return []
    table_id = configs["dingtalk"].get("tableId")
    if not table_id:
        return []
    return field_option_order(configs, table_id, brand_field_id)


def effective_output_brands(records: list[ReportRecord], preferred_order: list[str]) -> list[str]:
    grouped = set(record.brand for record in records)
    ordered = [brand for brand in preferred_order if brand in grouped]
    ordered_set = set(ordered)
    ordered.extend(record.brand for record in records if record.brand not in ordered_set and not ordered_set.add(record.brand))
    return ordered


def fetch_tracked_brands(configs: dict[str, dict[str, Any]], business: str | None) -> list[str]:
    if not business:
        return []
    rules = configs["report_rules"]
    tracked_cfg = rules.get("trackedBrandsTable", {})
    table_id = tracked_cfg.get("tableId")
    brand_field = tracked_cfg.get("brandFieldId")
    business_field = tracked_cfg.get("businessFieldId")
    status_field = tracked_cfg.get("statusFieldId")
    active_status_id = tracked_cfg.get("activeStatusOptionId")
    business_id = tracked_cfg.get("businessOptionIds", {}).get(business)
    if not all((table_id, brand_field, business_field, status_field, active_status_id, business_id)):
        raise SystemExit("report_rules.json 的 trackedBrandsTable 配置不完整，无法读取关注品牌小字。")

    dingtalk = configs["dingtalk"]
    records: list[dict[str, Any]] = []
    cursor = None
    seen_cursors: set[str] = set()
    while True:
        payload: dict[str, Any] = {
            "baseId": dingtalk["baseId"],
            "tableId": table_id,
            "limit": 100,
            "fieldIds": [brand_field, business_field, status_field],
            "filters": {
                "operator": "and",
                "operands": [
                    {"operator": "any_of", "operands": [business_field, business_id]},
                    {"operator": "eq", "operands": [status_field, active_status_id]},
                ],
            },
        }
        if cursor:
            payload["cursor"] = cursor
        data = call_dingtalk_tool(dingtalk, "query_records", payload)
        batch = data.get("data", {}).get("records", []) or []
        records.extend(batch)
        next_cursor = data.get("data", {}).get("nextCursor")
        if not next_cursor or next_cursor in seen_cursors or not batch:
            break
        seen_cursors.add(next_cursor)
        cursor = next_cursor

    brands: list[str] = []
    for record in records:
        brand = str(extract_cell(record.get("cells", {}).get(brand_field), "select_name") or "").strip()
        if brand:
            brands.append(brand)
    brands = list(dict.fromkeys(brands))
    preferred_order = field_option_order(configs, table_id, brand_field)
    ordered = [brand for brand in preferred_order if brand in brands]
    ordered_set = set(ordered)
    ordered.extend(brand for brand in brands if brand not in ordered_set and not ordered_set.add(brand))
    return ordered


def query_records(
    configs: dict[str, dict[str, Any]],
    field_ids: dict[str, str],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    dingtalk = configs["dingtalk"]
    date_field = field_ids.get("launchDate")
    if not date_field:
        raise SystemExit("字段映射中缺少 launchDate")

    records: list[dict[str, Any]] = []
    cursor = None
    seen_cursors: set[str] = set()
    seen_record_ids: set[str] = set()
    requested_field_ids = list(dict.fromkeys(field_ids.values()))

    while True:
        payload: dict[str, Any] = {
            "baseId": dingtalk["baseId"],
            "tableId": dingtalk["tableId"],
            "limit": 20,
            "fieldIds": requested_field_ids,
            "filters": {
                "operator": "and",
                "operands": [
                    {"operator": "not_before", "operands": [date_field, start.isoformat()]},
                    {"operator": "not_after", "operands": [date_field, end.isoformat()]},
                ],
            },
        }
        if cursor:
            payload["cursor"] = cursor
        data = call_dingtalk_tool(dingtalk, "query_records", payload)
        batch = data.get("data", {}).get("records", []) or []
        new_records = 0
        for raw_record in batch:
            record_id = str(raw_record.get("recordId") or "").strip()
            if record_id:
                if record_id in seen_record_ids:
                    continue
                seen_record_ids.add(record_id)
            records.append(raw_record)
            new_records += 1
        next_cursor = data.get("data", {}).get("nextCursor")
        if not next_cursor or next_cursor in seen_cursors or not batch or new_records == 0:
            break
        seen_cursors.add(next_cursor)
        cursor = next_cursor
    return records


def extract_cell(raw: Any, extract: str) -> Any:
    if raw is None:
        return "" if extract not in ("multi_select_names", "attachment_image_urls", "filter_up_names") else []
    if extract == "select_name":
        if isinstance(raw, dict):
            return str(raw.get("name", "")).strip()
        return str(raw).strip()
    if extract == "multi_select_names":
        if isinstance(raw, list):
            names = []
            for item in raw:
                if isinstance(item, dict):
                    name = str(item.get("name", "")).strip()
                    if name:
                        names.append(name)
                elif item:
                    names.append(str(item).strip())
            return names
        if isinstance(raw, dict):
            return [str(raw.get("name", "")).strip()] if raw.get("name") else []
        return [str(raw).strip()] if raw else []
    if extract == "filter_up_names":
        return extract_filter_up_names(raw)
    if extract == "rich_text_markdown":
        if isinstance(raw, dict):
            text = raw.get("markdown") or raw.get("text") or ""
        else:
            text = str(raw)
        return clean_text(text)
    if extract == "attachment_image_urls":
        return extract_attachment_image_urls(raw)
    if extract == "date":
        return parse_date_value(raw)
    if extract == "generic":
        if isinstance(raw, dict):
            return raw.get("name") or raw.get("text") or raw.get("value") or ""
        if isinstance(raw, list):
            return [extract_cell(item, "generic") for item in raw]
        return raw
    return clean_text(str(raw))


def extract_filter_up_names(raw: Any) -> list[str]:
    names: list[str] = []

    def visit(value: Any) -> None:
        if value is None:
            return
        if isinstance(value, dict):
            name = str(value.get("name") or "").strip()
            if name:
                names.append(name)
            text = str(value.get("text") or value.get("label") or "").strip()
            if text:
                names.append(text)
            if "value" in value:
                visit(value.get("value"))
            if "values" in value:
                visit(value.get("values"))
            if "records" in value:
                visit(value.get("records"))
            if "items" in value:
                visit(value.get("items"))
            return
        if isinstance(value, list):
            for item in value:
                visit(item)
            return
        text = str(value).strip()
        if text:
            names.append(text)

    visit(raw)
    return list(dict.fromkeys(names))


def extract_attachment_image_urls(raw: Any) -> list[str]:
    if not raw:
        return []
    items = raw if isinstance(raw, list) else [raw]
    urls: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        nested = item.get("files") or item.get("attachments")
        if isinstance(nested, list):
            urls.extend(extract_attachment_image_urls(nested))
            continue
        file_type = str(item.get("type") or item.get("mimeType") or item.get("contentType") or "").lower()
        filename = str(item.get("filename") or item.get("fileName") or item.get("name") or "").lower()
        is_image = (
            file_type == "image"
            or file_type.startswith("image/")
            or filename.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"))
        )
        if not is_image:
            continue
        for key in ("url", "resourceUrl", "downloadUrl", "previewUrl", "thumbnailUrl"):
            url = item.get(key)
            if url:
                urls.append(str(url))
                break
    return urls


def parse_date_value(raw: Any) -> date | None:
    if not raw:
        return None
    if isinstance(raw, (int, float)):
        # Dingtalk date fields normally return ISO strings; keep a safe Excel-like fallback.
        return datetime.fromtimestamp(raw / 1000).date()
    text = str(raw).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None


def clean_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    lines = [line for line in lines if line]
    return "\n".join(lines).strip()


def normalize_paragraph_text(text: str) -> str:
    parts = [line.strip() for line in str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    parts = [part for part in parts if part]
    if not parts:
        return ""
    end_punctuation = set("，。；：！？、,.!?;:")
    normalized: list[str] = []
    for index, part in enumerate(parts):
        if index < len(parts) - 1 and part[-1] not in end_punctuation:
            part += "。"
        normalized.append(part)
    return "".join(normalized)


def normalize_records(
    raw_records: list[dict[str, Any]],
    configs: dict[str, dict[str, Any]],
    field_ids: dict[str, str],
    brands: list[str],
    business: str | None = None,
    sort_brands: list[str] | None = None,
) -> list[ReportRecord]:
    fields_cfg = configs["field_mapping"]["standardFields"]
    rules = configs["report_rules"]
    brand_set = set(brands)
    brand_order = {brand: i for i, brand in enumerate(sort_brands or brands)}
    normalized: list[ReportRecord] = []
    seen_record_ids: set[str] = set()
    seen_business_keys: set[tuple[str, str, str, str, str]] = set()
    skipped_duplicates: list[str] = []

    for source_index, raw_record in enumerate(raw_records):
        cells = raw_record.get("cells", {})
        values: dict[str, Any] = {}
        for standard_key, field_id in field_ids.items():
            cfg = fields_cfg.get(standard_key, {})
            values[standard_key] = extract_cell(cells.get(field_id), cfg.get("extract", "text"))

        brand = str(values.get("brand") or "").strip()
        product_name = str(values.get("productName") or "").strip()
        business_values = values.get("business") or []
        if isinstance(business_values, str):
            business_values = [business_values]
        if not brand or not product_name or (brand_set and brand not in brand_set):
            continue
        if not brand_set and business and business not in business_values:
            continue
        if brand not in brand_order:
            brand_order[brand] = len(brand_order)

        category = compute_category(values, rules)
        remark = compute_remark(values, rules)
        record = ReportRecord(
            record_id=str(raw_record.get("recordId") or "").strip(),
            brand=brand,
            product_name=product_name,
            launch_date=values.get("launchDate"),
            price=str(values.get("price") or ""),
            category=category,
            series=str(values.get("series") or ""),
            selling_point=normalize_paragraph_text(str(values.get("sellingPoint") or "")),
            ingredients=normalize_paragraph_text(str(values.get("ingredients") or "")),
            image_urls=values.get("appearanceImages") or [],
            remark=remark,
            source_index=source_index,
        )
        if record.record_id and record.record_id in seen_record_ids:
            skipped_duplicates.append(describe_report_record(record))
            continue
        business_key = normalized_record_business_key(record)
        if business_key in seen_business_keys:
            skipped_duplicates.append(describe_report_record(record))
            continue
        if record.record_id:
            seen_record_ids.add(record.record_id)
        seen_business_keys.add(business_key)
        normalized.append(record)

    series_order: dict[tuple[str, date | None, str], int] = {}
    for record in sorted(normalized, key=lambda r: r.source_index):
        series_key = (record.brand, record.launch_date, record.series)
        if series_key not in series_order:
            series_order[series_key] = len(series_order)

    normalized.sort(
        key=lambda r: (
            brand_order.get(r.brand, 9999),
            r.launch_date or date.min,
            series_order.get((r.brand, r.launch_date, r.series), 9999),
            r.source_index,
        )
    )
    if skipped_duplicates:
        print_duplicate_warning("标准化重复记录", skipped_duplicates)
    return normalized


def ordered_record_brands(records: list[ReportRecord]) -> list[str]:
    return list(dict.fromkeys(record.brand for record in records))


def compute_category(values: dict[str, Any], rules: dict[str, Any]) -> str:
    category_rule = rules.get("categoryRule", {})
    additional = values.get(category_rule.get("additionalField", "additionalCategory")) or []
    if isinstance(additional, str):
        additional = [additional]
    for candidate in category_rule.get("priority", []):
        if candidate in additional:
            return candidate
    main = values.get(category_rule.get("mainField", "mainCategory"))
    if isinstance(main, list):
        return str(main[0]) if main else ""
    return str(main or "")


def truthy_return_status(value: Any, rule: dict[str, Any]) -> bool:
    if isinstance(value, list):
        return any(truthy_return_status(item, rule) for item in value)
    if isinstance(value, bool):
        return value
    text = str(value or "").strip()
    if not text:
        return False
    ignored = set(rule.get("ignoredValues", []))
    truthy = set(rule.get("truthyValues", []))
    if text in ignored:
        return False
    return text in truthy or rule.get("showValue", "回归") in text


def compute_remark(values: dict[str, Any], rules: dict[str, Any]) -> str:
    remark_rule = rules.get("remarkRule", {})
    parts: list[str] = []
    for key in remark_rule.get("order", []):
        if key == "returnStatus":
            if truthy_return_status(values.get(key), remark_rule.get("returnStatus", {})):
                parts.append(remark_rule.get("returnStatus", {}).get("showValue", "回归"))
        elif key == "collaboration":
            text = clean_text(str(values.get(key) or ""))
            if text:
                suffix = remark_rule.get("collaboration", {}).get("suffix", "联名")
                parts.append(text if text.endswith(suffix) else text + suffix)
        else:
            text = clean_text(str(values.get(key) or ""))
            if text:
                parts.append(text)
    return remark_rule.get("separator", "，").join(parts) if parts else remark_rule.get("emptyText", "/")


def format_title_date(start: date, end: date, template: str) -> str:
    return template.format(start_m=start.month, start_d=start.day, end_m=end.month, end_d=end.day)


def cm_to_points(cm: float) -> float:
    return cm / 2.54 * 72


def cm_to_pixels(cm: float) -> int:
    return int(round(cm / 2.54 * 96))


def pixels_to_emu(px: int) -> int:
    return int(px * 9525)


def points_to_pixels(points: float) -> int:
    return int(round(points * 96 / 72))


def column_width_to_pixels(width: float | None) -> int:
    if not width:
        return 64
    return int(round(width * 7.0 + 5))


def column_width_to_pixels_with_layout(width: float | None, layout: dict[str, Any]) -> int:
    auto_fit = layout.get("autoFit", {})
    scale = float(auto_fit.get("columnWidthPixelScale", 7.0))
    padding = float(auto_fit.get("columnWidthPixelPadding", 5.0))
    if not width:
        return int(round(8.43 * scale + padding))
    return int(round(width * scale + padding))


def pixels_to_column_width(pixels: float, layout: dict[str, Any]) -> float:
    auto_fit = layout.get("autoFit", {})
    scale = float(auto_fit.get("columnWidthPixelScale", 7.0))
    padding = float(auto_fit.get("columnWidthPixelPadding", 5.0))
    return max(1.0, round((pixels - padding) / scale, 2))


def text_to_column_width_chars(text: str, padding_chars: float = 0.0) -> float:
    return max(1.0, visual_len(text) + padding_chars)


def price_blocks(text: str) -> list[str]:
    value = clean_price_text(text)
    return split_outside_parentheses(value) or ([value] if value else [])


def split_outside_parentheses(text: str, separator: str = "/") -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    for char in text:
        if char in "(（":
            depth += 1
        elif char in ")）" and depth > 0:
            depth -= 1
        if char == separator and depth == 0:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)
    tail = "".join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def compute_column_widths(records: list[ReportRecord], layout: dict[str, Any]) -> dict[str, float]:
    rules = layout.get("columnWidthRules", {})
    widths: dict[str, float] = {}
    widths.update({col: float(width) for col, width in rules.get("sideColumnsChars", {}).items()})
    widths.update({col: float(width) for col, width in rules.get("fixedColumnsChars", {}).items()})

    padding = rules.get("paddingChars", {})
    mins = rules.get("minColumnsChars", {})
    maxes = rules.get("maxColumnsChars", {})

    longest_product = max((record.product_name for record in records), key=visual_len, default="")
    e_width = text_to_column_width_chars(longest_product, float(padding.get("E", 2.0)))
    e_width = max(float(mins.get("E", 1.0)), e_width)
    if "E" in maxes:
        e_width = min(float(maxes["E"]), e_width)
    widths["E"] = round(e_width, 2)

    longest_price_block = ""
    for record in records:
        for block in price_blocks(record.price):
            if visual_len(block) > visual_len(longest_price_block):
                longest_price_block = block
    g_width = text_to_column_width_chars(longest_price_block, float(padding.get("G", 2.0)))
    g_width = max(float(mins.get("G", 1.0)), g_width)
    if "G" in maxes:
        g_width = min(float(maxes["G"]), g_width)
    widths["G"] = round(g_width, 2)

    table_range = rules.get("tableRange", "B:H")
    start_col, end_col = table_range.split(":")
    table_cols = [get_column_letter(idx) for idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1)]
    fill_col = rules.get("fillRemainingColumn", "H")
    total_px = cm_to_pixels(float(rules.get("totalTableWidthCm", 18.0)))
    used_px = sum(column_width_to_pixels_with_layout(widths.get(col), layout) for col in table_cols if col != fill_col)
    min_fill = float(mins.get(fill_col, 1.0))
    remaining_width = pixels_to_column_width(total_px - used_px, layout)
    if remaining_width < min_fill:
        if rules.get("allowOverflowTableWidth", True):
            print(
                f"WARNING: E/G 自动列宽较大，{fill_col} 列按最小宽度 {min_fill} 设置，B:H 总宽会超过 {rules.get('totalTableWidthCm', 18.0)}cm。",
                file=sys.stderr,
            )
            remaining_width = min_fill
        else:
            remaining_width = max(1.0, remaining_width)
    widths[fill_col] = round(remaining_width, 2)
    return widths


def apply_column_widths(ws, layout: dict[str, Any], records: list[ReportRecord]) -> dict[str, float]:
    columns = compute_column_widths(records, layout)
    for col, width in columns.items():
        ws.column_dimensions[col].width = float(width)
    return columns


def visual_len(text: str) -> int:
    return sum(2 if ord(ch) > 127 else 1 for ch in str(text))


def estimate_line_count(text: str, col_chars: float, chars_per_line_factor: float = 1.8) -> int:
    if not text:
        return 1
    lines = 0
    effective_chars = max(col_chars * chars_per_line_factor, 1)
    for part in str(text).split("\n"):
        length = visual_len(part)
        lines += max(1, math.ceil(length / effective_chars))
    return max(1, lines)


def row_height_for_line_count(line_count: int, layout: dict[str, Any], max_height: float | None = None) -> float:
    auto_fit = layout.get("autoFit", layout.get("rowHeightRule", {}))
    single = float(auto_fit.get("singleLineHeightPt", auto_fit.get("singleLinePt", 16.8)))
    additional = float(auto_fit.get("additionalLineHeightPt", auto_fit.get("additionalLinePt", 15.0)))
    height = single if line_count <= 1 else single + (line_count - 1) * additional
    if max_height is not None:
        height = min(max_height, height)
    return height


def estimate_auto_row_height(
    cells: list[tuple[str, float, bool]],
    layout: dict[str, Any],
    min_height: float | None = None,
    max_height: float | None = None,
    extra_padding: float = 0.0,
) -> float:
    auto_fit = layout.get("autoFit", {})
    max_lines = 1
    for text, width_chars, wrap in cells:
        if wrap:
            max_lines = max(max_lines, estimate_line_count(text, width_chars))
        else:
            max_lines = max(max_lines, max(1, str(text or "").count("\n") + 1))
    height = row_height_for_line_count(max_lines, layout, max_height) + extra_padding
    if min_height is None:
        min_height = float(auto_fit.get("minTextRowHeightPt", 16.8))
    return max(min_height, height)


def column_width_sum(column_widths: dict[str, float], start_col: str, end_col: str) -> float:
    return sum(
        column_widths.get(get_column_letter(idx), 8.43)
        for idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1)
    )


def row_height_for_text(
    text: str,
    width_chars: float,
    layout: dict[str, Any],
    chars_per_line_factor: float,
    padding_pt: float = 0.0,
    max_height: float | None = None,
) -> float:
    line_count = estimate_line_count(text, width_chars, chars_per_line_factor)
    padding = padding_pt if line_count > 1 else 0.0
    return estimate_height_for_line_count(line_count, layout, padding, max_height)


def estimate_height_for_line_count(
    line_count: int,
    layout: dict[str, Any],
    padding_pt: float = 0.0,
    max_height: float | None = None,
) -> float:
    auto_fit = layout.get("autoFit", {})
    min_height = float(auto_fit.get("minTextRowHeightPt", 16.8))
    if max_height is None:
        max_height = float(auto_fit.get("maxTextRowHeightPt", 120.0))
    return max(min_height, row_height_for_line_count(line_count, layout, max_height) + padding_pt)


def estimate_summary_row_height(
    record: ReportRecord,
    price_display: str,
    column_widths: dict[str, float],
    layout: dict[str, Any],
) -> float:
    auto_fit = layout.get("autoFit", {})
    factor = float(auto_fit.get("summaryCharsPerLineFactor", 1.45))
    padding = float(auto_fit.get("summaryVerticalPaddingPt", 3.0))
    wrap_columns = set(auto_fit.get("summaryWrapColumns", ["G", "H"]))
    line_counts = [1]
    if "G" in wrap_columns:
        line_counts.append(estimate_line_count(price_display, column_widths.get("G", 19.0), factor))
    if "H" in wrap_columns:
        line_counts.append(estimate_line_count(record.remark, column_widths.get("H", 12.0), factor))
    max_lines = max(line_counts)
    return estimate_height_for_line_count(max_lines, layout, padding if max_lines > 1 else 0.0)


def estimate_detail_row_height(key: str, text: str, value_width: float, layout: dict[str, Any]) -> float:
    auto_fit = layout.get("autoFit", {})
    max_height = float(auto_fit.get("maxTextRowHeightPt", 120.0))
    if key == "sellingPoint":
        return row_height_for_text(
            text,
            value_width,
            layout,
            float(auto_fit.get("detailLongTextCharsPerLineFactor", 1.5)),
            float(auto_fit.get("detailLongTextPaddingPt", 4.0)),
            max_height=max_height,
        )
    if key == "ingredients":
        return row_height_for_text(
            text,
            value_width,
            layout,
            float(auto_fit.get("detailShortTextCharsPerLineFactor", 1.55)),
            float(auto_fit.get("detailShortTextPaddingPt", 2.0)),
            max_height=max_height,
        )
    return estimate_height_for_line_count(1, layout)


def clean_price_text(text: str) -> str:
    text = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    parts = [line.strip() for line in text.split("\n") if line.strip()]
    text = "".join(parts)
    return re.sub(r"\s*/\s*", "/", text).strip()


def normalized_record_business_key(record: ReportRecord) -> tuple[str, str, str, str, str]:
    launch_date = record.launch_date.isoformat() if record.launch_date else ""
    return (
        record.brand.strip(),
        record.product_name.strip(),
        launch_date,
        clean_price_text(record.price),
        clean_text(record.series),
    )


def describe_raw_record(
    raw_record: dict[str, Any],
    configs: dict[str, dict[str, Any]],
    field_ids: dict[str, str],
) -> str:
    fields_cfg = configs["field_mapping"]["standardFields"]
    cells = raw_record.get("cells", {})
    product_field = field_ids.get("productName")
    brand_field = field_ids.get("brand")
    product_name = ""
    brand = ""
    if product_field:
        product_cfg = fields_cfg.get("productName", {})
        product_name = str(extract_cell(cells.get(product_field), product_cfg.get("extract", "text")) or "").strip()
    if brand_field:
        brand_cfg = fields_cfg.get("brand", {})
        brand = str(extract_cell(cells.get(brand_field), brand_cfg.get("extract", "text")) or "").strip()
    record_id = str(raw_record.get("recordId") or "").strip() or "无recordId"
    label = " / ".join(part for part in (brand, product_name) if part)
    return f"{label or '未命名记录'} ({record_id})"


def describe_report_record(record: ReportRecord) -> str:
    launch_date = record.launch_date.isoformat() if record.launch_date else "无日期"
    record_id = record.record_id or "无recordId"
    return f"{record.brand} / {record.product_name} / {launch_date} / {clean_price_text(record.price)} ({record_id})"


def quality_record_label(record: ReportRecord) -> str:
    launch_date = record.launch_date.isoformat() if record.launch_date else "无日期"
    record_id = record.record_id or "无recordId"
    product_name = record.product_name or "未命名新品"
    return f"{record.brand} / {product_name} / {launch_date} / {record_id}"


def is_missing_text(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, date):
        return False
    return not clean_text(str(value))


def collect_data_quality_report(records: list[ReportRecord], rules: dict[str, Any]) -> DataQualityReport:
    quality_rule = rules.get("dataQualityRule", {})
    if not quality_rule.get("enabled", True):
        return DataQualityReport([], [], [])

    default_fields = {
        "category": "品类",
        "productName": "新品名称",
        "launchDate": "上市日期",
        "price": "产品价格",
        "sellingPoint": "产品卖点介绍",
        "ingredients": "原料构成",
    }
    configured_fields = quality_rule.get("requiredTextFields") or default_fields
    missing_fields: list[str] = []
    for record in records:
        values = {
            "category": record.category,
            "productName": record.product_name,
            "launchDate": record.launch_date,
            "price": clean_price_text(record.price),
            "sellingPoint": record.selling_point,
            "ingredients": record.ingredients,
        }
        for field_key, label in configured_fields.items():
            if is_missing_text(values.get(field_key)):
                missing_fields.append(f"{quality_record_label(record)}：缺少{label}")

    missing_images: list[str] = []
    if quality_rule.get("warnMissingImages", True):
        for record in records:
            if not record.image_urls:
                missing_images.append(f"{quality_record_label(record)}：缺少产品外观图片")

    return DataQualityReport(missing_fields, missing_images, [])


def print_limited_warning(title: str, items: list[str], limit: int = 20) -> None:
    if not items:
        return
    print(title, file=sys.stderr)
    for item in items[:limit]:
        print(f"  - {item}", file=sys.stderr)
    if len(items) > limit:
        print(f"  - 还有 {len(items) - limit} 条未列出", file=sys.stderr)


def print_data_quality_warnings(report: DataQualityReport) -> None:
    print_limited_warning(
        f"WARNING: 发现 {len(report.missing_fields)} 处底表字段缺失，请检查钉钉 AI 表：",
        report.missing_fields,
    )
    print_limited_warning(
        f"WARNING: 发现 {len(report.missing_images)} 条产品外观图片缺失，请检查钉钉 AI 表：",
        report.missing_images,
    )
    print_limited_warning(
        f"WARNING: 发现 {len(report.image_download_failures)} 条产品外观图片下载失败，请检查钉钉 AI 表图片附件：",
        report.image_download_failures,
    )


def data_quality_warning_lines(report: DataQualityReport, limit: int = 20) -> list[str]:
    lines: list[str] = []

    def append_group(title: str, items: list[str]) -> None:
        if not items:
            return
        lines.append(title)
        for item in items[:limit]:
            lines.append(f"- {item}")
        if len(items) > limit:
            lines.append(f"- 还有 {len(items) - limit} 条未列出")

    append_group(f"WARNING: 发现 {len(report.missing_fields)} 处底表字段缺失，请检查钉钉 AI 表：", report.missing_fields)
    append_group(f"WARNING: 发现 {len(report.missing_images)} 条产品外观图片缺失，请检查钉钉 AI 表：", report.missing_images)
    append_group(f"WARNING: 发现 {len(report.image_download_failures)} 条产品外观图片下载失败，请检查钉钉 AI 表图片附件：", report.image_download_failures)
    return lines


def print_duplicate_warning(label: str, items: list[str]) -> None:
    print(f"WARNING: 已跳过 {len(items)} 条{label}：", file=sys.stderr)
    for item in items[:10]:
        print(f"  - {item}", file=sys.stderr)
    if len(items) > 10:
        print(f"  - 还有 {len(items) - 10} 条未列出", file=sys.stderr)


def price_with_slash_wrap(text: str, threshold_chars: int) -> str:
    text = clean_price_text(text)
    if not text or "/" not in text:
        return text
    return "/\n".join(split_outside_parentheses(text))


def build_price_rich_text(text: str, font_name: str, size: int, strike_pattern: str) -> CellRichText | str:
    if not text:
        return ""
    normal = InlineFont(rFont=font_name, sz=size)
    strike = InlineFont(rFont=font_name, sz=size, strike=True)
    price_pattern = re.compile(strike_pattern)
    rich = CellRichText()
    found = False

    pos = 0
    for match in price_pattern.finditer(text):
        if not is_inside_parentheses(text, match.start()):
            continue
        if match.start() > pos:
            rich.append(TextBlock(normal, text[pos : match.start()]))
        rich.append(TextBlock(strike, match.group(0)))
        found = True
        pos = match.end()
    if pos < len(text):
        rich.append(TextBlock(normal, text[pos:]))
    return rich if found else text


def is_inside_parentheses(text: str, index: int) -> bool:
    depth = 0
    for ch in text[:index]:
        if ch in "(（":
            depth += 1
        elif ch in ")）" and depth > 0:
            depth -= 1
    return depth > 0


def format_price_for_output(text: str, configs: dict[str, dict[str, Any]], wrap_after_slash: bool = False) -> CellRichText | str:
    rules = configs["report_rules"].get("priceRule", {})
    layout = configs["excel_layout"]
    font_name = configured_excel_font_name(configs, "chineseFont")
    size = layout.get("fonts", {}).get("defaultSize", 10)
    threshold = layout.get("_computedPriceWrapThresholdChars", layout.get("summary", {}).get("priceWrapThresholdChars", 19))
    value = price_with_slash_wrap(text, threshold) if wrap_after_slash else clean_price_text(text)
    if rules.get("strikeParenthesizedPrice", True):
        return build_price_rich_text(value, font_name, size, rules.get("parenthesizedPricePattern", r"\d+(?:\.\d+)?\s*元"))
    return value


def set_cell_price(cell, text: str, configs: dict[str, dict[str, Any]], wrap_after_slash: bool = False) -> None:
    layout = configs["excel_layout"]
    font_name = configured_excel_font_name(configs, "chineseFont")
    size = layout.get("fonts", {}).get("defaultSize", 10)
    cell.font = Font(name=font_name, size=size, color="000000")
    cell.value = format_price_for_output(text, configs, wrap_after_slash)


def apply_base_style(
    cell,
    font_name: str,
    size: int = 10,
    bold: bool = False,
    fill: str | None = None,
    border: Border | None = None,
    horizontal: str = "center",
    vertical: str = "center",
    wrap_text: bool = True,
):
    cell.font = Font(name=font_name, size=size, bold=bold, color="000000")
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = border


def style_range(
    ws,
    range_ref: str,
    font_name: str,
    size: int,
    bold: bool,
    fill: str | None,
    border: Border | None,
    horizontal: str = "center",
    vertical: str = "center",
    wrap_text: bool = True,
    outer_border_only: bool = False,
) -> None:
    range_cells = ws[range_ref]
    min_row = range_cells[0][0].row
    max_row = range_cells[-1][0].row
    min_col = range_cells[0][0].column
    max_col = range_cells[0][-1].column
    for row in ws[range_ref]:
        for c in row:
            c.font = Font(name=font_name, size=size, bold=bold, color="000000")
            c.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if border:
                if outer_border_only:
                    c.border = Border(
                        left=border.left if c.column == min_col else Side(style=None),
                        right=border.right if c.column == max_col else Side(style=None),
                        top=border.top if c.row == min_row else Side(style=None),
                        bottom=border.bottom if c.row == max_row else Side(style=None),
                    )
                else:
                    c.border = border


def write_across_range(
    ws,
    range_ref: str,
    value: Any,
    font_name: str,
    size: int,
    bold: bool,
    fill: str | None,
    border: Border | None,
    horizontal: str = "centerContinuous",
    vertical: str = "center",
    wrap_text: bool = True,
    outer_border_only: bool = True,
) -> None:
    start_ref = range_ref.split(":")[0]
    cell = ws[start_ref]
    cell.value = value
    style_range(
        ws,
        range_ref,
        font_name,
        size,
        bold,
        fill,
        border,
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text,
        outer_border_only=outer_border_only,
    )
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def download_image(url: str, record_id: str, cache_dir: Path, image_cache: ImageCache | None = None) -> Path | None:
    if not url:
        return None
    if image_cache:
        cached = image_cache.fetch(url, record_id)
        return cached.path if cached else None
    cache_dir.mkdir(parents=True, exist_ok=True)
    suffix = ".png"
    clean_id = re.sub(r"[^A-Za-z0-9_-]+", "_", record_id or "image")
    target = cache_dir / f"{clean_id}{suffix}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            target.write_bytes(resp.read())
        return target
    except Exception as exc:
        print(f"WARNING: 图片下载失败：{record_id} {exc}", file=sys.stderr)
        return None


def add_logo(ws, layout: dict[str, Any]) -> None:
    logo = layout.get("logo", {})
    if not logo.get("path"):
        return
    logo_path = ROOT / logo["path"]
    if not logo_path.exists():
        print(f"WARNING: logo 不存在：{logo_path}", file=sys.stderr)
        return
    row_number = int(logo.get("row", 2))
    logo_range = logo.get("range", logo.get("mergedRange", "B:H"))
    start_col, end_col = logo_range.split(":")
    image = XLImage(str(logo_path))
    target_height = cm_to_pixels(float(logo.get("heightCm", 2.0)))
    ratio = image.width / image.height if image.height else 1
    target_width = int(target_height * ratio)
    image.width = target_width
    image.height = target_height
    row_height_px = points_to_pixels(ws.row_dimensions[row_number].height or cm_to_points(float(logo.get("heightCm", 2.0))))
    marker = range_center_anchor_marker(
        ws,
        start_col,
        end_col,
        row_number,
        target_width,
        target_height,
        row_height_px,
        layout,
    )
    image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_emu(target_width), pixels_to_emu(target_height)),
    )
    ws.add_image(image)


def column_range_width_px(ws, start_col: str, end_col: str, layout: dict[str, Any]) -> int:
    start = column_index_from_string(start_col)
    end = column_index_from_string(end_col)
    total = 0
    for idx in range(start, end + 1):
        letter = get_column_letter(idx)
        total += column_width_to_pixels_with_layout(ws.column_dimensions[letter].width, layout)
    return total


def range_center_anchor_marker(
    ws,
    start_col: str,
    end_col: str,
    row_number: int,
    target_width_px: int,
    target_height_px: int,
    row_height_px: int,
    layout: dict[str, Any],
    padding_px: int = 0,
) -> AnchorMarker:
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    range_width_px = column_range_width_px(ws, start_col, end_col, layout)
    target_left_px = max(padding_px, (range_width_px - target_width_px) // 2)
    target_top_px = max(padding_px, (row_height_px - target_height_px) // 2)

    current_col_idx = start_idx
    remaining_px = target_left_px
    while current_col_idx < end_idx:
        col_letter = get_column_letter(current_col_idx)
        col_width_px = column_width_to_pixels_with_layout(ws.column_dimensions[col_letter].width, layout)
        if remaining_px < col_width_px:
            break
        remaining_px -= col_width_px
        current_col_idx += 1

    return AnchorMarker(
        col=current_col_idx - 1,
        colOff=pixels_to_emu(remaining_px),
        row=row_number - 1,
        rowOff=pixels_to_emu(target_top_px),
    )


def add_image(
    ws,
    image_path: Path,
    row_number: int,
    layout: dict[str, Any],
) -> None:
    try:
        image_cfg = layout["image"]
        with PILImage.open(image_path) as pil:
            width, height = pil.size
        target_height = cm_to_pixels(image_cfg["heightCm"])
        target_width = int(width * target_height / height) if height else target_height
        max_width_px = image_cfg.get("maxWidthPx", 520)
        if target_width > max_width_px:
            target_width = max_width_px
            target_height = int(height * target_width / width) if width else cm_to_pixels(image_cfg["heightCm"])
        image = XLImage(str(image_path))
        image.height = target_height
        image.width = target_width
        anchor_col = image_cfg.get("anchorColumn", "C")
        if image_cfg.get("centerInRange", image_cfg.get("centerInMergedCell", False)):
            image_range = image_cfg.get("range", image_cfg.get("mergedRange", "C:H"))
            start_col, end_col = image_range.split(":")
            row_height_px = points_to_pixels(ws.row_dimensions[row_number].height or cm_to_points(image_cfg["heightCm"]))
            padding = int(image_cfg.get("paddingPx", 8))
            marker = range_center_anchor_marker(
                ws,
                start_col,
                end_col,
                row_number,
                target_width,
                target_height,
                row_height_px,
                layout,
                padding_px=padding,
            )
            image.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(pixels_to_emu(target_width), pixels_to_emu(target_height)),
            )
        else:
            image.anchor = f"{anchor_col}{row_number}"
        ws.add_image(image)
    except Exception as exc:
        print(f"WARNING: 图片插入失败：{image_path} {exc}", file=sys.stderr)


def apply_page_fill_and_fonts(ws, configs: dict[str, dict[str, Any]], layout: dict[str, Any], max_row: int) -> None:
    page_fill = layout.get("pageFill", {})
    col_range = page_fill.get("columns", "A:I")
    start_col, end_col = col_range.split(":")
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    fill_color = page_fill.get("fill", "FFFFFF")
    default_font = configured_excel_font_name(configs, "chineseFont")
    default_size = layout["fonts"].get("defaultSize", 10)
    default_names = {None, "Calibri", "宋体", "SimSun"}
    for row in range(1, max_row + 1):
        for col in range(start_idx, end_idx + 1):
            cell = ws.cell(row, col)
            if cell.fill.fill_type is None:
                cell.fill = PatternFill("solid", fgColor=fill_color)
            if cell.font.name in default_names:
                cell.font = Font(name=default_font, size=default_size, color="000000")


def build_workbook(
    records: list[ReportRecord],
    configs: dict[str, dict[str, Any]],
    start: date,
    end: date,
    brands: list[str],
    tracked_brands: list[str],
    output_path: Path,
    data_quality_report: DataQualityReport,
    image_cache: ImageCache | None = None,
) -> None:
    layout = configs["excel_layout"]
    rules = configs["report_rules"]
    colors = layout["colors"]
    fonts = layout["fonts"]
    heights = layout.get("rowHeightsPt", layout.get("rowHeights", {}))
    chinese_font = configured_excel_font_name(configs, "chineseFont")
    latin_font = configured_excel_font_name(configs, "latinFont")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb._named_styles["Normal"].font = Font(name=chinese_font, size=fonts["defaultSize"], color="000000")
    ws = wb.active
    ws.title = layout["worksheetName"]
    ws.sheet_view.showGridLines = False

    column_widths = apply_column_widths(ws, layout, records)
    layout["_computedPriceWrapThresholdChars"] = max(1, int(column_widths.get("G", 19.0) * 1.8))

    ws.row_dimensions[2].height = heights["topSpacer"]
    style_range(ws, "B2:H2", chinese_font, fonts["defaultSize"], False, colors["white"], None, horizontal="centerContinuous")
    add_logo(ws, layout)
    write_across_range(ws, "B3:H3", format_title_date(start, end, layout["titleTemplate"]), chinese_font, fonts["titleSize"], False, colors["white"], None)
    ws.row_dimensions[3].height = 24
    style_range(ws, "B5:H5", chinese_font, fonts["introSize"], False, colors["white"], None, horizontal="left", wrap_text=False)
    ws["B5"] = layout["introText"]
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[5].height = heights["intro"]

    header_row = layout["summary"]["headerRow"]
    summary_cols = layout["summary"]["columns"]
    headers = layout["summary"]["headers"]
    for key in ("brand", "count", "category", "productName", "launchDate", "price", "remark"):
        col = summary_cols[key]
        cell = ws[f"{col}{header_row}"]
        cell.value = headers.get(key, "品牌" if key == "brand" else "")
        apply_base_style(cell, chinese_font, bold=True, fill=colors["headerFill"], border=border)
    ws.row_dimensions[header_row].height = heights["summaryHeader"]

    grouped: dict[str, list[ReportRecord]] = defaultdict(list)
    for record in records:
        grouped[record.brand].append(record)

    row = layout["summary"]["startRow"]
    displayed_brands = [brand for brand in brands if grouped.get(brand)]
    for brand in displayed_brands:
        brand_records = grouped[brand]
        start_row = row
        end_row = row + len(brand_records) - 1
        if start_row != end_row:
            ws.merge_cells(f"B{start_row}:B{end_row}")
            ws.merge_cells(f"C{start_row}:C{end_row}")
        ws[f"B{start_row}"] = brand
        ws[f"C{start_row}"] = len(brand_records)
        for cell_ref in (f"B{start_row}", f"C{start_row}"):
            apply_base_style(ws[cell_ref], chinese_font if cell_ref.startswith("B") else latin_font, bold=True, fill=colors["brandFill"] if cell_ref.startswith("B") else None, border=border)
        for rr in range(start_row, end_row + 1):
            for col in ("B", "C"):
                ws[f"{col}{rr}"].border = border

        for record in brand_records:
            ws[f"D{row}"] = record.category
            ws[f"E{row}"] = record.product_name
            ws[f"F{row}"] = f"{record.launch_date.month}月{record.launch_date.day}日" if record.launch_date else ""
            ws[f"F{row}"].number_format = "@"
            set_cell_price(ws[f"G{row}"], record.price, configs, wrap_after_slash=True)
            ws[f"H{row}"] = record.remark
            no_wrap_columns = layout["summary"].get("noWrapColumns", [])
            no_wrap_by_col = {"D": "category", "E": "productName"}
            for col in ("D", "E", "F", "G", "H"):
                wrap = no_wrap_by_col.get(col) not in no_wrap_columns
                apply_base_style(ws[f"{col}{row}"], chinese_font, border=border, wrap_text=wrap)
            price_display = price_with_slash_wrap(record.price, layout["_computedPriceWrapThresholdChars"])
            ws.row_dimensions[row].height = estimate_summary_row_height(record, price_display, column_widths, layout)
            row += 1

    note_row = row
    note_text = layout["trackedBrandsPrefix"] + "、".join(tracked_brands)
    note_font_size = fonts.get("trackedBrandsSize", 8)
    note_range = f"B{note_row}:H{note_row}"
    style_range(ws, note_range, chinese_font, note_font_size, False, colors["white"], None, horizontal="left", wrap_text=False)
    ws.merge_cells(note_range)
    ws[f"B{note_row}"] = note_text
    ws[f"B{note_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[note_row].height = float(heights.get("trackedBrands", 16.8))

    detail_row = note_row + layout["details"]["blankRowsAfterSummary"] + 1
    image_cache_manager = tempfile.TemporaryDirectory(prefix="weekly_report_images_") if image_cache is None else None
    image_cache_dir = Path(image_cache_manager.name) if image_cache_manager else Path(tempfile.gettempdir())
    try:
        for brand in displayed_brands:
            brand_records = grouped[brand]
            write_across_range(
                ws,
                f"B{detail_row}:H{detail_row}",
                layout["details"]["brandHeaderTemplate"].format(brand=brand, count=len(brand_records)),
                chinese_font,
                fonts["defaultSize"],
                True,
                colors["detailHeaderFill"],
                border,
            )
            ws.row_dimensions[detail_row].height = heights["detailHeader"]
            detail_row += 1

            for record in brand_records:
                row_values = {
                    "productName": record.product_name,
                    "series": record.series or "/",
                    "sellingPoint": record.selling_point,
                    "price": record.price,
                    "ingredients": record.ingredients,
                    "appearanceImages": "",
                }
                for key, label in layout["details"]["rowsPerProduct"]:
                    label_fill = colors["detailLabelFill"] if key in ("productName", "series") else colors["white"]
                    ws[f"B{detail_row}"] = label
                    apply_base_style(
                        ws[f"B{detail_row}"],
                        chinese_font,
                        bold=key == "productName",
                        fill=label_fill,
                        border=border,
                        wrap_text=layout["details"].get("labelWrapText", False),
                    )
                    value_range = f"C{detail_row}:H{detail_row}"
                    value_cell = ws[f"C{detail_row}"]
                    horizontal = "centerContinuous"
                    style_range(
                        ws,
                        value_range,
                        chinese_font,
                        fonts["defaultSize"],
                        key == "productName",
                        colors["detailLabelFill"] if key in ("productName", "series") else None,
                        border,
                        horizontal=horizontal,
                        wrap_text=True,
                        outer_border_only=True,
                    )
                    value_cell.value = format_price_for_output(record.price, configs, wrap_after_slash=False) if key == "price" else row_values.get(key, "")
                    value_cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

                    if key == "appearanceImages":
                        image_height = cm_to_points(float(layout["image"].get("heightCm", 4.0))) + float(heights.get("imagePadding", 10.0))
                        ws.row_dimensions[detail_row].height = max(float(layout["image"].get("rowHeightPt", 0.0)), image_height)
                        if record.image_urls:
                            image_path = download_image(record.image_urls[0], record.record_id, image_cache_dir, image_cache)
                            if image_path:
                                add_image(
                                    ws,
                                    image_path,
                                    detail_row,
                                    layout,
                                )
                            else:
                                data_quality_report.image_download_failures.append(f"{quality_record_label(record)}：产品外观图片下载失败")
                    else:
                        text = clean_price_text(record.price) if key == "price" else str(row_values.get(key, ""))
                        value_width = column_width_sum(column_widths, "C", "H")
                        ws.row_dimensions[detail_row].height = estimate_detail_row_height(key, text, value_width, layout)
                    detail_row += 1

        apply_page_fill_and_fonts(ws, configs, layout, max(detail_row, ws.max_row))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        if image_cache_manager:
            image_cache_manager.cleanup()


def default_output_path(start: date, end: date, layout: dict[str, Any]) -> Path:
    output_dir = ROOT / layout.get("outputDirectory", "outputs")
    return output_dir / f"竞品新品周报{start.isoformat()}_{end.isoformat()}.xlsx"


def resolve_non_overwriting_output_path(path: Path) -> Path:
    suffix = path.suffix
    stem = path.stem
    parent = path.parent
    if not parent.exists():
        return path
    escaped_stem = re.escape(stem)
    escaped_suffix = re.escape(suffix)
    pattern = re.compile(rf"^{escaped_stem}(?:_(\d+))?{escaped_suffix}$")
    max_suffix = 0
    for existing in parent.glob(f"{stem}*{suffix}"):
        match = pattern.match(existing.name)
        if not match:
            continue
        number = int(match.group(1)) if match.group(1) else 1
        max_suffix = max(max_suffix, number)
    if max_suffix == 0:
        return path
    return parent / f"{stem}_{max_suffix + 1}{suffix}"


def cleanup_legacy_image_cache(output_dir: Path) -> None:
    legacy_cache = output_dir / "_image_cache"
    if legacy_cache.exists() and legacy_cache.is_dir():
        shutil.rmtree(legacy_cache)


def main() -> int:
    args = parse_args()
    configs = load_configs()

    warnings = validate_config(configs)
    if args.validate_config:
        if warnings:
            print("配置校验完成，有警告：")
            for warning in warnings:
                print(f"- {warning}")
        else:
            print("配置校验通过。")
        return 0 if not [w for w in warnings if "缺少" in w or "必须" in w or "需要" in w] else 1

    if args.explain_config:
        print(explain_config(configs))
        if warnings:
            print("\n配置警告：")
            for warning in warnings:
                print(f"- {warning}")
        return 0

    start, end = resolve_date_window(args)
    brands = parse_brands(args.brands)
    business = parse_business(args.business, configs["report_rules"])
    table_fields = fetch_table_fields(configs)
    field_ids = resolve_field_ids(configs["field_mapping"], table_fields)
    default_brand_order = default_brand_option_order(configs, field_ids) if not brands else []
    if not brands and not default_brand_order:
        print("WARNING: 未从钉钉品牌字段读取到标签列表顺序，默认按本次记录首次出现顺序输出品牌。")
    raw_records = query_records(configs, field_ids, start, end)
    records = normalize_records(raw_records, configs, field_ids, brands, business, default_brand_order)
    if not brands:
        brands = effective_output_brands(records, default_brand_order)
    tracked_brands = fetch_tracked_brands(configs, business) if business else brands
    data_quality_report = collect_data_quality_report(records, configs["report_rules"])
    output_path = Path(args.output) if args.output else default_output_path(start, end, configs["excel_layout"])
    if not output_path.is_absolute():
        output_path = ROOT / output_path
    output_path = resolve_non_overwriting_output_path(output_path)
    build_workbook(records, configs, start, end, brands, tracked_brands, output_path, data_quality_report)
    cleanup_legacy_image_cache(output_path.parent)
    print_data_quality_warnings(data_quality_report)
    print(f"已生成：{output_path}")
    print(f"记录数：{len(records)}")
    print(f"数据提醒：缺失字段 {data_quality_report.missing_field_count} 处，缺图 {data_quality_report.image_issue_count} 条。")
    print(f"本次生成文件：{output_path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
